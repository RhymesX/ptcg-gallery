from __future__ import annotations

import hashlib
import json
import os
import re
import secrets
import sqlite3
import threading
import unicodedata
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

DEFAULT_EXCEL_NAME = "卡表.xlsx"
SEARCH_PREFERENCES_FILE_NAME = "search_preferences.json"
NICKNAME_EXCEL_NAME = "nicknames.xlsx"
DEFAULT_ACCOUNT_NAME = "RhymesX"
EXPECTED_HEADERS = {
    "商品名称": "product_name",
    "商品编号": "product_code",
    "卡牌编号": "card_code",
    "卡牌名称": "card_name",
    "类型": "card_type",
    "详细": "detail",
    "特殊": "special_text",
    "属性": "attribute",
    "稀有度": "rarity",
    "赛制": "regulation",
    "数量": "quantity",
    "备注": "note",
}
EXACT_CODE_PATTERN = re.compile(r"^\s*([A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)\s*-\s*([A-Za-z0-9]+)\s*$")
PROMO_PRODUCT_KEYWORD = "PROMO"
PROMO_NUMBERED_CARD_CODE_PATTERN = re.compile(r"^\s*(\d+)\s*/\s*([A-Za-z]+-[A-Za-z]+)\s*$", re.IGNORECASE)
PROMO_SERIES_CARD_CODE_PATTERN = re.compile(r"^\s*([A-Za-z]+-[A-Za-z]+)\s*$", re.IGNORECASE)
PROMO_YEAR_PATTERN = re.compile(r"(20\d{2})")
PROMO_RANK_PATTERN = re.compile(r"(冠军|亚军|殿军|\d+强)")
PROMO_CITIES = ("北京", "广州", "深圳", "上海", "苏州", "杭州")
DEFAULT_DECKS = ["电友", "龙柱", "铝钢龙", "多龙"]
DEFAULT_DECK_COLORS = {
    "电友": "#ffe95a",
    "龙柱": "#8b1f1f",
    "铝钢龙": "#c6c8ce",
    "多龙": "#f28c28",
}
DEFAULT_DECK_COLOR = "#9ca3af"
SUPPORTER_LABEL = "支援者"
LEGACY_SUPPORTER_LABEL = "\u5fd7\u613f\u8005"
LEGACY_DEFAULT_DECK_COLORS = {
    "龙柱": "#c9711f",
}
POKEMON_CATEGORY_KEYS = {"ordinary_pokemon", "pokemon_gx", "pokemon_v", "pokemon_ex", "radiant_pokemon"}
ATTRIBUTE_ORDER = ["草", "火", "水", "电", "超", "斗", "恶", "钢", "龙", "妖", "无"]
ATTRIBUTE_ALIASES = {
    "草": ("草",),
    "火": ("火", "炎"),
    "水": ("水",),
    "电": ("电", "雷"),
    "超": ("超",),
    "斗": ("斗",),
    "恶": ("恶",),
    "钢": ("钢", "金属"),
    "龙": ("龙",),
    "妖": ("妖", "妖精"),
    "无": ("无", "无色"),
}
ATTRIBUTE_COLOR_FALLBACKS = {
    "草": "#4caf50",
    "火": "#ef6c36",
    "水": "#2f80ed",
    "电": "#f5b700",
    "超": "#9b5de5",
    "斗": "#a35a1f",
    "恶": "#344054",
    "钢": "#7a8798",
    "龙": "#5a3fd1",
    "妖": "#ec6aa7",
    "无": "#9aa0a6",
}
ATTRIBUTE_ORDER_INDEX = {attribute: index for index, attribute in enumerate(ATTRIBUTE_ORDER)}
CARD_NAME_VARIANT_SUFFIXES = ("精灵球闪", "大师球闪", "球闪", "闪")
SPECIAL_SAME_NAME_KEYWORDS = ("神奇糖果", "超级球", "博士的研究", "老大的指令", "宝可梦捕捉器", "裁判", "巢穴球")
CATALOG_IDENTITY_FIELDS = (
    "product_code",
    "card_code",
    "card_name",
    "card_type",
    "detail",
    "special_text",
    "attribute",
    "rarity",
    "regulation",
)
CARD_CATEGORY_DEFINITIONS = [
    ("ordinary_pokemon", "普通的宝可梦"),
    ("pokemon_gx", "宝可梦GX"),
    ("pokemon_v", "宝可梦V"),
    ("pokemon_ex", "宝可梦ex"),
    ("radiant_pokemon", "光辉宝可梦"),
    ("item", "物品"),
    ("supporter", SUPPORTER_LABEL),
    ("stadium", "竞技场"),
    ("tool", "道具"),
    ("special_energy", "特殊能量"),
    ("basic_energy", "普通能量"),
]
DECK_DETAIL_SECTION_DEFINITIONS = [
    ("pokemon", "宝可梦", "left"),
    ("energy", "能量", "left"),
    ("item", "物品", "right"),
    ("supporter", SUPPORTER_LABEL, "right"),
    ("stadium", "竞技场", "right"),
    ("tool", "宝可梦道具", "right"),
    ("backup", "备卡", "full"),
]
DECK_DETAIL_SECTION_CATEGORY_MAP = {
    "ordinary_pokemon": "pokemon",
    "pokemon_gx": "pokemon",
    "pokemon_v": "pokemon",
    "pokemon_ex": "pokemon",
    "radiant_pokemon": "pokemon",
    "item": "item",
    "supporter": "supporter",
    "stadium": "stadium",
    "tool": "tool",
    "special_energy": "energy",
    "basic_energy": "energy",
    "backup": "backup",
}
BASIC_DECK_ENERGY_DEFINITIONS = [
    {"code": "GRA", "name": "基本草能量", "attribute": "草"},
    {"code": "FIR", "name": "基本火能量", "attribute": "火"},
    {"code": "WAT", "name": "基本水能量", "attribute": "水"},
    {"code": "LIG", "name": "基本雷能量", "attribute": "电"},
    {"code": "PSY", "name": "基本超能量", "attribute": "超"},
    {"code": "FIG", "name": "基本斗能量", "attribute": "斗"},
    {"code": "DAR", "name": "基本恶能量", "attribute": "恶"},
    {"code": "MET", "name": "基本钢能量", "attribute": "钢"},
]
BASIC_DECK_ENERGY_LOOKUP = {item["code"]: item for item in BASIC_DECK_ENERGY_DEFINITIONS}
BASIC_DECK_ENERGY_ORDER_INDEX = {item["code"]: index for index, item in enumerate(BASIC_DECK_ENERGY_DEFINITIONS)}
DEFAULT_SEARCH_PREFERENCES = {
    "selectedRegulations": [],
    "considerSameNameRegulation": False,
}
CATALOG_DB_SCHEMA_VERSION = 1
ACCOUNT_DB_SCHEMA_VERSION = 1
USER_SETTINGS_SEARCH_PREFERENCES_KEY = "search.preferences"


@dataclass(slots=True)
class AppPaths:
    root_dir: Path
    data_dir: Path
    db_path: Path
    accounts_dir: Path
    default_excel_path: Path


class ServiceError(ValueError):
    pass


class NotFoundError(ServiceError):
    pass


class ConflictError(ServiceError):
    pass


def build_paths(root_dir: str | os.PathLike[str]) -> AppPaths:
    root = Path(root_dir).resolve()
    data_dir = root / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    accounts_dir = data_dir / "accounts"
    accounts_dir.mkdir(parents=True, exist_ok=True)
    return AppPaths(
        root_dir=root,
        data_dir=data_dir,
        db_path=data_dir / "ptcg_gallery.db",
        accounts_dir=accounts_dir,
        default_excel_path=data_dir / DEFAULT_EXCEL_NAME,
    )


SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS cards (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_key TEXT NOT NULL UNIQUE,
    product_name TEXT DEFAULT '',
    product_code TEXT NOT NULL,
    card_code TEXT NOT NULL,
    card_name TEXT NOT NULL,
    card_type TEXT DEFAULT '',
    detail TEXT DEFAULT '',
    special_text TEXT DEFAULT '',
    attribute TEXT DEFAULT '',
    attribute_color TEXT DEFAULT '',
    rarity TEXT DEFAULT '',
    regulation TEXT DEFAULT '',
    note TEXT DEFAULT '',
    nickname TEXT DEFAULT '',
    show_nickname INTEGER NOT NULL DEFAULT 0,
    group_sort_order INTEGER NOT NULL DEFAULT 0,
    initial_quantity INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS accounts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    password_hash TEXT NOT NULL DEFAULT '',
    wx_openid TEXT NOT NULL DEFAULT '',
    sort_order INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS app_settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL DEFAULT '',
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS free_inventory (
    account_id INTEGER NOT NULL,
    card_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    PRIMARY KEY(account_id, card_id),
    FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE,
    FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS decks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    account_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    description TEXT NOT NULL DEFAULT '',
    color TEXT NOT NULL DEFAULT '',
    sort_order INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(account_id, name),
    FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS deck_cards (
    deck_id INTEGER NOT NULL,
    card_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    backup_quantity INTEGER NOT NULL DEFAULT 0 CHECK (backup_quantity >= 0),
    PRIMARY KEY(deck_id, card_id),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE,
    FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS deck_basic_energies (
    deck_id INTEGER NOT NULL,
    energy_code TEXT NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    PRIMARY KEY(deck_id, energy_code),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS holdings_group_orders (
    section_key TEXT NOT NULL,
    group_key TEXT NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY(section_key, group_key)
);

CREATE TABLE IF NOT EXISTS deck_section_orders (
    deck_id INTEGER NOT NULL,
    section_key TEXT NOT NULL,
    entry_key TEXT NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY(deck_id, section_key, entry_key),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_cards_code ON cards(product_code, card_code);
CREATE INDEX IF NOT EXISTS idx_cards_name ON cards(card_name);
CREATE INDEX IF NOT EXISTS idx_cards_regulation ON cards(regulation);
CREATE INDEX IF NOT EXISTS idx_deck_cards_card_id ON deck_cards(card_id);
CREATE INDEX IF NOT EXISTS idx_deck_cards_deck_id ON deck_cards(deck_id);

CREATE TABLE IF NOT EXISTS invite_codes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    used_by_account_id INTEGER DEFAULT NULL,
    used_at TEXT DEFAULT NULL,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    expires_at TEXT NOT NULL,
    FOREIGN KEY(used_by_account_id) REFERENCES accounts(id) ON DELETE SET NULL
);

CREATE INDEX IF NOT EXISTS idx_invite_codes_code ON invite_codes(code);
CREATE INDEX IF NOT EXISTS idx_invite_codes_expires ON invite_codes(expires_at);
"""

# 用户独立数据库的 schema（每个用户一个文件），不含 account_id 列
ACCOUNT_SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS free_inventory (
    card_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    PRIMARY KEY(card_id)
);

CREATE TABLE IF NOT EXISTS decks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT NOT NULL DEFAULT '',
    color TEXT NOT NULL DEFAULT '',
    sort_order INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS deck_cards (
    deck_id INTEGER NOT NULL,
    card_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    backup_quantity INTEGER NOT NULL DEFAULT 0 CHECK (backup_quantity >= 0),
    PRIMARY KEY(deck_id, card_id),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS deck_basic_energies (
    deck_id INTEGER NOT NULL,
    energy_code TEXT NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
    PRIMARY KEY(deck_id, energy_code),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS deck_section_orders (
    deck_id INTEGER NOT NULL,
    section_key TEXT NOT NULL,
    entry_key TEXT NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY(deck_id, section_key, entry_key),
    FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS holdings_group_orders (
    section_key TEXT NOT NULL,
    group_key TEXT NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY(section_key, group_key)
);

CREATE TABLE IF NOT EXISTS holdings_card_orders (
    card_id INTEGER NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY(card_id)
);

CREATE TABLE IF NOT EXISTS user_settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL DEFAULT '',
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""


class CardRepository:
    _account_local = threading.local()
    def __init__(self, db_path: str | os.PathLike[str], *, accounts_dir: str | os.PathLike[str] | None = None):
        self.db_path = str(db_path)
        self._init_admin_pass = ""
        self.search_preferences_path = Path(self.db_path).resolve().parent / SEARCH_PREFERENCES_FILE_NAME
        self.accounts_dir = str(accounts_dir) if accounts_dir else str(Path(self.db_path).resolve().parent / "accounts")
        Path(self.accounts_dir).mkdir(parents=True, exist_ok=True)
        self._accounts_initialized: set[int] = set()
        self.initialize()

    def set_init_admin_pass(self, password: str):
        self._init_admin_pass = str(password or "")

    def _configure_connection(self, conn: sqlite3.Connection):
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA busy_timeout = 5000")

    def _connect_sqlite(self, db_path: str | os.PathLike[str]) -> sqlite3.Connection:
        conn = sqlite3.connect(str(db_path))
        self._configure_connection(conn)
        return conn

    def _set_db_schema_version(self, conn: sqlite3.Connection, version: int):
        conn.execute(f"PRAGMA user_version = {int(version)}")

    @contextmanager
    def connect(self):
        """Legacy alias for the shared catalog database."""
        conn = self._connect_sqlite(self.db_path)
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    @contextmanager
    def connect_catalog(self):
        """连接共享的 catalog 数据库（cards, accounts, invite_codes 等）。"""
        conn = self._connect_sqlite(self.db_path)
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _account_db_path(self, account_id: int) -> str:
        return str(Path(self.accounts_dir) / f"{int(account_id)}.db")

    @contextmanager
    def connect_account(self, account_id: int):
        """连接指定用户的独立数据库，同时 ATTACH catalog 数据库为 shared。"""
        account_id = int(account_id)
        if account_id not in self._accounts_initialized:
            self._init_account_db(account_id)
            self._accounts_initialized.add(account_id)
        db_path = self._account_db_path(account_id)
        conn = self._connect_sqlite(db_path)
        conn.execute("ATTACH DATABASE ? AS shared", (self.db_path,))
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _init_account_db(self, account_id: int):
        """初始化指定用户的独立数据库（建表）。"""
        account_id = int(account_id)
        db_path = self._account_db_path(account_id)
        conn = self._connect_sqlite(db_path)
        try:
            conn.executescript(ACCOUNT_SCHEMA_SQL)
            self._set_db_schema_version(conn, ACCOUNT_DB_SCHEMA_VERSION)
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _delete_account_db(self, account_id: int):
        Path(self._account_db_path(account_id)).unlink(missing_ok=True)

    def initialize(self):
        with self.connect() as conn:
            conn.executescript(SCHEMA_SQL)
            self._set_db_schema_version(conn, CATALOG_DB_SCHEMA_VERSION)
            self._ensure_card_attribute_color_column(conn)
            self._ensure_card_group_sort_order_column(conn)
            self._ensure_card_show_nickname_column(conn)
            self._ensure_deck_color_column(conn)
            self._ensure_deck_sort_order_column(conn)
            self._ensure_deck_card_backup_quantity_column(conn)
            self._ensure_account_password_hash_column(conn)
            self._ensure_account_wx_openid_column(conn)
            self._ensure_account_storage(conn)
            self._normalize_supporter_wording(conn)
            self._normalize_deck_sort_order(conn)
        self._sync_nicknames()

    def _ensure_account_storage(self, conn: sqlite3.Connection):
        account_id = self._ensure_default_account(conn)
        self._ensure_free_inventory_account_schema(conn, account_id)
        self._ensure_decks_account_schema(conn, account_id)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_free_inventory_account ON free_inventory(account_id, card_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_decks_account_sort_order ON decks(account_id, sort_order, id)")
        self._ensure_default_account_password(conn)
        for row in conn.execute("SELECT id FROM accounts ORDER BY id ASC").fetchall():
            aid = int(row["id"])
            if aid not in self._accounts_initialized:
                self._init_account_db(aid)
                self._accounts_initialized.add(aid)

    def _ensure_default_account(self, conn: sqlite3.Connection) -> int:
        row = conn.execute("SELECT id FROM accounts WHERE name = ?", (DEFAULT_ACCOUNT_NAME,)).fetchone()
        if row is not None:
            return int(row["id"])
        next_sort_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM accounts").fetchone()[0]
        cursor = conn.execute(
            "INSERT INTO accounts(name, sort_order, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
            (DEFAULT_ACCOUNT_NAME, next_sort_order),
        )
        return int(cursor.lastrowid)

    def _ensure_account_password_hash_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(accounts)").fetchall()}
        if "password_hash" not in columns:
            conn.execute("ALTER TABLE accounts ADD COLUMN password_hash TEXT NOT NULL DEFAULT ''")

    def _ensure_account_wx_openid_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(accounts)").fetchall()}
        if "wx_openid" not in columns:
            conn.execute("ALTER TABLE accounts ADD COLUMN wx_openid TEXT NOT NULL DEFAULT ''")

    def _ensure_default_account_password(self, conn: sqlite3.Connection):
        if not self._init_admin_pass:
            return
        row = conn.execute("SELECT id, password_hash FROM accounts WHERE name = ?", (DEFAULT_ACCOUNT_NAME,)).fetchone()
        if row is None:
            return
        existing = normalize_text(row["password_hash"])
        if existing:
            return
        conn.execute(
            "UPDATE accounts SET password_hash = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (self.build_password_hash(self._init_admin_pass), int(row["id"])),
        )

    @classmethod
    def build_password_hash(cls, password: str) -> str:
        clean_password = str(password or "")
        if not clean_password:
            raise ServiceError("密码不能为空")
        salt = secrets.token_hex(16)
        digest = hashlib.sha256(f"{salt}:{clean_password}".encode("utf-8")).hexdigest()
        return f"{salt}${digest}"

    @classmethod
    def verify_password_hash(cls, stored_hash: str, password: str) -> bool:
        clean_hash = normalize_text(stored_hash)
        clean_password = str(password or "")
        if not clean_hash or not clean_password or "$" not in clean_hash:
            return False
        salt, digest = clean_hash.split("$", 1)
        expected = hashlib.sha256(f"{salt}:{clean_password}".encode("utf-8")).hexdigest()
        return secrets.compare_digest(digest, expected)

    @classmethod
    def set_request_account_id(cls, account_id: int | None):
        if account_id is None:
            if hasattr(cls._account_local, "account_id"):
                delattr(cls._account_local, "account_id")
            return
        cls._account_local.account_id = int(account_id)

    @classmethod
    def clear_request_account_id(cls):
        cls.set_request_account_id(None)

    def _resolve_account_id_from_context(self, conn: sqlite3.Connection) -> int:
        context_id = getattr(self._account_local, "account_id", None)
        if context_id is not None:
            exists = conn.execute("SELECT id FROM accounts WHERE id = ?", (int(context_id),)).fetchone()
            if exists is not None:
                return int(context_id)
        return self._ensure_default_account(conn)

    def _ensure_free_inventory_account_schema(self, conn: sqlite3.Connection, default_account_id: int):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(free_inventory)").fetchall()}
        if "account_id" in columns:
            return
        conn.execute("ALTER TABLE free_inventory RENAME TO free_inventory_legacy")
        conn.execute(
            """
            CREATE TABLE free_inventory (
                account_id INTEGER NOT NULL,
                card_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
                PRIMARY KEY(account_id, card_id),
                FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE,
                FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            "INSERT INTO free_inventory(account_id, card_id, quantity) SELECT ?, card_id, quantity FROM free_inventory_legacy",
            (default_account_id,),
        )
        conn.execute("DROP TABLE free_inventory_legacy")

    def _ensure_decks_account_schema(self, conn: sqlite3.Connection, default_account_id: int):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(decks)").fetchall()}
        if "account_id" in columns:
            return
        conn.execute("CREATE TEMP TABLE deck_cards_account_migration_backup AS SELECT * FROM deck_cards")
        conn.execute("CREATE TEMP TABLE deck_basic_energies_account_migration_backup AS SELECT * FROM deck_basic_energies")
        conn.execute("CREATE TEMP TABLE deck_section_orders_account_migration_backup AS SELECT * FROM deck_section_orders")
        conn.execute("ALTER TABLE decks RENAME TO decks_legacy")
        conn.execute(
            """
            CREATE TABLE decks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                color TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(account_id, name),
                FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            INSERT INTO decks(id, account_id, name, description, color, sort_order, created_at, updated_at)
            SELECT id, ?, name, description, color, sort_order, created_at, updated_at FROM decks_legacy
            """,
            (default_account_id,),
        )
        conn.execute("DROP TABLE decks_legacy")
        conn.execute("DROP TABLE deck_cards")
        conn.execute(
            """
            CREATE TABLE deck_cards (
                deck_id INTEGER NOT NULL,
                card_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
                backup_quantity INTEGER NOT NULL DEFAULT 0 CHECK (backup_quantity >= 0),
                PRIMARY KEY(deck_id, card_id),
                FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE,
                FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute("INSERT INTO deck_cards(deck_id, card_id, quantity, backup_quantity) SELECT deck_id, card_id, quantity, backup_quantity FROM deck_cards_account_migration_backup")
        conn.execute("DROP TABLE deck_basic_energies")
        conn.execute(
            """
            CREATE TABLE deck_basic_energies (
                deck_id INTEGER NOT NULL,
                energy_code TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
                PRIMARY KEY(deck_id, energy_code),
                FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute("INSERT INTO deck_basic_energies(deck_id, energy_code, quantity) SELECT deck_id, energy_code, quantity FROM deck_basic_energies_account_migration_backup")
        conn.execute("DROP TABLE deck_section_orders")
        conn.execute(
            """
            CREATE TABLE deck_section_orders (
                deck_id INTEGER NOT NULL,
                section_key TEXT NOT NULL,
                entry_key TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY(deck_id, section_key, entry_key),
                FOREIGN KEY(deck_id) REFERENCES decks(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute("INSERT INTO deck_section_orders(deck_id, section_key, entry_key, sort_order) SELECT deck_id, section_key, entry_key, sort_order FROM deck_section_orders_account_migration_backup")
        conn.execute("DROP TABLE deck_cards_account_migration_backup")
        conn.execute("DROP TABLE deck_basic_energies_account_migration_backup")
        conn.execute("DROP TABLE deck_section_orders_account_migration_backup")

    def _ensure_card_attribute_color_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(cards)").fetchall()}
        if "attribute_color" not in columns:
            conn.execute("ALTER TABLE cards ADD COLUMN attribute_color TEXT NOT NULL DEFAULT ''")

    def _ensure_card_group_sort_order_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(cards)").fetchall()}
        if "group_sort_order" not in columns:
            conn.execute("ALTER TABLE cards ADD COLUMN group_sort_order INTEGER NOT NULL DEFAULT 0")

    def _ensure_card_show_nickname_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(cards)").fetchall()}
        if "show_nickname" not in columns:
            conn.execute("ALTER TABLE cards ADD COLUMN show_nickname INTEGER NOT NULL DEFAULT 0")

    def _ensure_deck_color_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(decks)").fetchall()}
        if "color" not in columns:
            conn.execute("ALTER TABLE decks ADD COLUMN color TEXT NOT NULL DEFAULT ''")

    def _ensure_deck_sort_order_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(decks)").fetchall()}
        if "sort_order" not in columns:
            conn.execute("ALTER TABLE decks ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0")

    def _ensure_deck_card_backup_quantity_column(self, conn: sqlite3.Connection):
        columns = {row[1] for row in conn.execute("PRAGMA table_info(deck_cards)").fetchall()}
        if "backup_quantity" not in columns:
            conn.execute("ALTER TABLE deck_cards ADD COLUMN backup_quantity INTEGER NOT NULL DEFAULT 0")

    def _normalize_supporter_wording(self, conn: sqlite3.Connection):
        conn.execute(
            """
            UPDATE cards
            SET card_type = REPLACE(card_type, ?, ?),
                detail = REPLACE(detail, ?, ?),
                special_text = REPLACE(special_text, ?, ?),
                updated_at = CURRENT_TIMESTAMP
            WHERE INSTR(card_type, ?) > 0
               OR INSTR(detail, ?) > 0
               OR INSTR(special_text, ?) > 0
            """,
            (
                LEGACY_SUPPORTER_LABEL,
                SUPPORTER_LABEL,
                LEGACY_SUPPORTER_LABEL,
                SUPPORTER_LABEL,
                LEGACY_SUPPORTER_LABEL,
                SUPPORTER_LABEL,
                LEGACY_SUPPORTER_LABEL,
                LEGACY_SUPPORTER_LABEL,
                LEGACY_SUPPORTER_LABEL,
            ),
        )

    def _normalize_deck_sort_order(self, conn: sqlite3.Connection):
        rows = conn.execute(
            "SELECT id, account_id, sort_order FROM decks ORDER BY account_id ASC, sort_order ASC, id ASC"
        ).fetchall()
        if not rows:
            return

        account_rows: dict[int, list[sqlite3.Row]] = {}
        for row in rows:
            account_rows.setdefault(int(row["account_id"]), []).append(row)

        for account_rows_for_id in account_rows.values():
            sort_orders = [int(row["sort_order"]) for row in account_rows_for_id]
            if all(sort_order > 0 for sort_order in sort_orders):
                continue
            positive_sort_orders = [sort_order for sort_order in sort_orders if sort_order > 0]
            next_sort_order = max(positive_sort_orders, default=0) + 1
            for row in sorted(account_rows_for_id, key=lambda item: item["id"]):
                if int(row["sort_order"]) > 0:
                    continue
                conn.execute(
                    "UPDATE decks SET sort_order = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (next_sort_order, row["id"]),
                )
                next_sort_order += 1

    def get_current_account_id(self, conn: sqlite3.Connection) -> int:
        self._ensure_account_storage(conn)
        return self._resolve_account_id_from_context(conn)

    def _resolve_account_id(self) -> int:
        with self.connect_catalog() as conn:
            return self.get_current_account_id(conn)

    @contextmanager
    def connect_current_account(self):
        account_id = self._resolve_account_id()
        with self.connect_account(account_id) as conn:
            yield conn

    def list_accounts(self) -> dict[str, Any]:
        with self.connect_catalog() as conn:
            current_account_id = self.get_current_account_id(conn)
            rows = conn.execute(
                """
                SELECT a.id,
                       a.name,
                       a.sort_order,
                       a.created_at,
                       a.updated_at
                FROM accounts a
                ORDER BY a.sort_order ASC, a.id ASC
                """
            ).fetchall()
        accounts = [
            dict(row)
            | {
                "sortOrder": row["sort_order"],
                **self._account_db_stats(int(row["id"])),
                "isCurrent": int(row["id"]) == current_account_id,
            }
            for row in rows
        ]
        current = next((account for account in accounts if account["isCurrent"]), accounts[0] if accounts else None)
        return {"items": accounts, "current": current}

    def create_account(self, name: str, password: str = "123456") -> dict[str, Any]:
        clean_name = normalize_text(name)
        if not clean_name:
            raise ServiceError("账号名称不能为空")
        clean_password = (password or "").strip()
        if not clean_password:
            raise ServiceError("密码不能为空")
        if len(clean_password) < 4:
            raise ServiceError("密码至少需要 4 位")
        with self.connect() as conn:
            try:
                next_sort_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM accounts").fetchone()[0]
                cursor = conn.execute(
                    "INSERT INTO accounts(name, password_hash, sort_order, updated_at) VALUES (?, ?, ?, CURRENT_TIMESTAMP)",
                    (clean_name, self.build_password_hash(clean_password), next_sort_order),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError(f"账号“{clean_name}”已存在") from exc
            account_id = int(cursor.lastrowid)
        self._init_account_db(account_id)
        with self.connect_account(account_id) as account_conn:
            self._ensure_default_decks_for_account(account_conn)
        return self.list_accounts()

    def create_wechat_account(self, name: str, openid: str) -> dict[str, Any]:
        """为微信小程序用户创建无密码账号（通过 wx_openid 绑定）。"""
        clean_name = normalize_text(name)
        if not clean_name:
            raise ServiceError("账号名称不能为空")
        with self.connect_catalog() as conn:
            existing = conn.execute("SELECT id FROM accounts WHERE name = ?", (clean_name,)).fetchone()
            if existing is not None:
                clean_name = f"{clean_name}_{openid[-4:]}"
            try:
                next_sort = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM accounts").fetchone()[0]
                cursor = conn.execute(
                    "INSERT INTO accounts(name, password_hash, wx_openid, sort_order, updated_at) VALUES (?, '', ?, ?, CURRENT_TIMESTAMP)",
                    (clean_name, openid, next_sort),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError('账号"{}"已存在'.format(clean_name)) from exc
            account_id = int(cursor.lastrowid)
        self._init_account_db(account_id)
        with self.connect_account(account_id) as account_conn:
            self._ensure_default_decks_for_account(account_conn)
        return {"id": account_id, "name": clean_name}

    def get_account_by_wx_openid(self, openid: str) -> dict[str, Any] | None:
        """根据微信 openid 查找已绑定的账号。"""
        if not openid:
            return None
        with self.connect_catalog() as conn:
            row = conn.execute(
                "SELECT id, name FROM accounts WHERE wx_openid = ?",
                (openid,),
            ).fetchone()
        return dict(row) if row is not None else None

    def switch_account(self, account_id: int) -> dict[str, Any]:
        with self.connect_catalog() as conn:
            row = conn.execute("SELECT id FROM accounts WHERE id = ?", (int(account_id),)).fetchone()
            if row is None:
                raise NotFoundError("账号不存在")
            self.set_request_account_id(int(account_id))
        return self.list_accounts()

    def delete_account(self, account_id: int) -> dict[str, Any]:
        account_id = int(account_id)
        with self.connect_catalog() as conn:
            account_count = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
            if account_count <= 1:
                raise ServiceError("至少需要保留一个账号")
            row = conn.execute("SELECT id, name FROM accounts WHERE id = ?", (account_id,)).fetchone()
            if row is None:
                raise NotFoundError("账号不存在")
            if normalize_text(row["name"]) == normalize_text(DEFAULT_ACCOUNT_NAME):
                raise ServiceError("不能删除管理员账号")
            conn.execute("DELETE FROM invite_codes WHERE used_by_account_id = ?", (account_id,))
            conn.execute("DELETE FROM accounts WHERE id = ?", (account_id,))
        self._delete_account_db(account_id)
        return self.list_accounts()

    def get_account_by_name(self, name: str) -> dict[str, Any] | None:
        clean_name = normalize_text(name)
        if not clean_name:
            return None
        with self.connect() as conn:
            row = conn.execute(
                "SELECT id, name, password_hash FROM accounts WHERE name = ?",
                (clean_name,),
            ).fetchone()
        return dict(row) if row is not None else None

    def verify_account_credentials(self, name: str, password: str) -> dict[str, Any] | None:
        account = self.get_account_by_name(name)
        if account is None:
            return None
        if not self.verify_password_hash(str(account.get("password_hash", "")), password):
            return None
        return account

    def change_account_password(self, account_id: int, old_password: str, new_password: str) -> None:
        account_id = int(account_id)
        clean_new = (new_password or "").strip()
        if not clean_new:
            raise ServiceError("新密码不能为空")
        if len(clean_new) < 4:
            raise ServiceError("新密码至少需要 4 位")
        with self.connect() as conn:
            row = conn.execute(
                "SELECT id, password_hash FROM accounts WHERE id = ?", (account_id,),
            ).fetchone()
            if row is None:
                raise NotFoundError("账号不存在")
            if not self.verify_password_hash(str(row["password_hash"] or ""), old_password):
                raise ServiceError("原密码不正确")
            conn.execute(
                "UPDATE accounts SET password_hash = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (self.build_password_hash(clean_new), account_id),
            )

    def reset_account_password(self, account_id: int, new_password: str) -> None:
        """管理员直接重置任意账号密码，不需旧密码。"""
        account_id = int(account_id)
        clean_new = (new_password or "").strip()
        if not clean_new:
            raise ServiceError("新密码不能为空")
        if len(clean_new) < 4:
            raise ServiceError("新密码至少需要 4 位")
        with self.connect() as conn:
            row = conn.execute(
                "SELECT id, name FROM accounts WHERE id = ?", (account_id,),
            ).fetchone()
            if row is None:
                raise NotFoundError("账号不存在")
            if normalize_text(row["name"]) == normalize_text(DEFAULT_ACCOUNT_NAME):
                raise ServiceError("不能重置管理员账号的密码，请使用改密功能")
            conn.execute(
                "UPDATE accounts SET password_hash = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (self.build_password_hash(clean_new), account_id),
            )

    # ── 邀请码 ─────────────────────────────────────────────

    def generate_invite_code(self) -> dict[str, Any]:
        """管理员生成一个 24 小时有效的邀请码。"""
        import uuid
        code = str(uuid.uuid4())[:8].upper()
        with self.connect() as conn:
            conn.execute(
                "INSERT INTO invite_codes(code, expires_at) VALUES (?, datetime('now', '+1 day'))",
                (code,),
            )
        return self.list_invite_codes()

    def create_bind_code(self, account_id: int) -> str:
        """生成 6 位绑定码（5 分钟有效），供已登录 Web 用户绑定微信小程序使用。
        如果账号已绑定微信，抛出 ValueError。"""
        with self.connect_catalog() as conn:
            row = conn.execute(
                "SELECT wx_openid FROM accounts WHERE id = ?", (int(account_id),)
            ).fetchone()
            if row and (row["wx_openid"] or "").strip():
                raise ValueError("该账号已绑定微信，无需重复绑定")
        import random
        code = str(random.randint(100000, 999999))
        expires_at = str(int(__import__('time').time()) + 300)
        with self.connect_catalog() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO app_settings(key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                (f"bind_code:{code}", f"{int(account_id)}|{expires_at}"),
            )
        return code

    def consume_bind_code(self, code: str) -> int | None:
        """校验绑定码是否有效，有效则返回 account_id 并删除该码。返回 None 表示无效或过期。"""
        clean_code = normalize_text(code)
        if not clean_code:
            return None
        with self.connect_catalog() as conn:
            row = conn.execute(
                "SELECT value FROM app_settings WHERE key = ?",
                (f"bind_code:{clean_code}",),
            ).fetchone()
            if row is None:
                return None
            parts = (row["value"] or "").split("|")
            if len(parts) != 2:
                return None
            account_id = int(parts[0])
            expires_at = int(parts[1])
            if int(__import__('time').time()) > expires_at:
                conn.execute("DELETE FROM app_settings WHERE key = ?", (f"bind_code:{clean_code}",))
                return None
            conn.execute("DELETE FROM app_settings WHERE key = ?", (f"bind_code:{clean_code}",))
            return account_id

    def bind_wx_openid(self, account_id: int, openid: str) -> None:
        """将微信 openid 绑定到指定账号。"""
        with self.connect_catalog() as conn:
            conn.execute(
                "UPDATE accounts SET wx_openid = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (openid, int(account_id)),
            )

    def is_invite_required(self) -> bool:
        """读取注册邀请码开关，不存在时默认开启（返回 True）。"""
        with self.connect_catalog() as conn:
            row = conn.execute(
                "SELECT value FROM app_settings WHERE key = 'registration.require_invite'"
            ).fetchone()
            if row is None:
                return True
            return normalize_text(row["value"]).lower() in ("true", "1", "yes", "on")

    def set_invite_required(self, required: bool) -> None:
        """设置注册邀请码开关。"""
        value = "true" if required else "false"
        with self.connect_catalog() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO app_settings(key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                ("registration.require_invite", value),
            )

    def load_nicknames_from_excel(self) -> dict[str, tuple[str, bool]]:
        """从 data/nicknames.xlsx 加载昵称映射。返回 {(product_code, card_code, card_name): (nickname, show_nickname)}。"""
        excel_path = Path(self.db_path).resolve().parent / NICKNAME_EXCEL_NAME
        if not excel_path.exists():
            return {}
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {}
            header = [str(c or "").strip() for c in rows[0]]
            try:
                pc_idx = next(i for i, h in enumerate(header) if h and ("商品编号" in h or "product" in h.lower()))
                cc_idx = next(i for i, h in enumerate(header) if h and ("卡牌编号" in h or "card_code" in h.lower()))
                name_idx = next(i for i, h in enumerate(header) if h and ("卡牌名称" in h or "card_name" in h.lower()))
                nick_idx = next(i for i, h in enumerate(header) if h and ("昵称" in h or "nickname" in h.lower()))
                show_idx = next(i for i, h in enumerate(header) if h and ("显示" in h or "show" in h.lower()))
            except StopIteration:
                return {}
            nicknames: dict[str, tuple[str, bool]] = {}
            for row in rows[1:]:
                if not row:
                    continue
                pc = normalize_text(str(row[pc_idx] or ""))
                cc = normalize_text(str(row[cc_idx] or ""))
                name = str(row[name_idx] or "").strip()
                nick = str(row[nick_idx] or "").strip()
                show = str(row[show_idx] or "").strip() if show_idx < len(row) else "0"
                if pc and cc and name and nick:
                    key = "|".join([pc.upper(), cc.upper(), name])
                    nicknames[key] = (nick, show == "1")
            return nicknames
        finally:
            wb.close()

    def sync_nicknames_to_db(self, nicknames: dict[str, tuple[str, bool]]):
        """将昵称与显示开关写入 cards 表。"""
        with self.connect_catalog() as conn:
            for key, (nick, show) in nicknames.items():
                parts = key.split("|", 2)
                if len(parts) != 3:
                    continue
                pc, cc, name = parts
                conn.execute(
                    "UPDATE cards SET nickname = ?, show_nickname = ?, updated_at = CURRENT_TIMESTAMP WHERE UPPER(product_code) = ? AND UPPER(card_code) = ? AND TRIM(card_name) = ?",
                    (nick, int(show), pc, cc, name),
                )
            # 清除已不在 Excel 中的昵称
            if nicknames:
                conn.execute("UPDATE cards SET nickname = '', show_nickname = 0, updated_at = CURRENT_TIMESTAMP WHERE nickname != ''")

    def _sync_nicknames(self):
        """从 Excel 加载昵称并同步到数据库。"""
        nicknames = self.load_nicknames_from_excel()
        if nicknames:
            self.sync_nicknames_to_db(nicknames)

    def list_invite_codes(self) -> dict[str, Any]:
        """列出所有未使用的邀请码及其过期时间。"""
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT id, code, created_at, expires_at
                FROM invite_codes
                WHERE used_by_account_id IS NULL AND datetime(expires_at) > datetime('now')
                ORDER BY created_at DESC
                """
            ).fetchall()
        codes = [
            {"id": row["id"], "code": row["code"], "createdAt": row["created_at"], "expiresAt": row["expires_at"]}
            for row in rows
        ]
        return {"codes": codes}

    def validate_invite_code(self, code: str) -> bool:
        """仅校验邀请码是否有效（不消费）。"""
        clean_code = normalize_text(code).upper()
        if not clean_code:
            return False
        with self.connect() as conn:
            row = conn.execute(
                """
                SELECT id FROM invite_codes
                WHERE code = ? AND used_by_account_id IS NULL AND datetime(expires_at) > datetime('now')
                """,
                (clean_code,),
            ).fetchone()
            return row is not None

    def consume_invite_code(self, code: str, account_id: int) -> None:
        """消费邀请码（标记为已用）。仅在创建账号后调用。"""
        clean_code = normalize_text(code).upper()
        if not clean_code or not account_id:
            return
        with self.connect() as conn:
            conn.execute(
                "UPDATE invite_codes SET used_by_account_id = ?, used_at = datetime('now') WHERE code = ? AND used_by_account_id IS NULL",
                (int(account_id), clean_code),
            )

    def _ensure_default_decks_for_account(self, conn: sqlite3.Connection) -> list[str]:
        created: list[str] = []
        for deck_name in DEFAULT_DECKS:
            deck = conn.execute("SELECT id, color FROM decks WHERE name = ?", (deck_name,)).fetchone()
            default_color = DEFAULT_DECK_COLORS.get(deck_name, DEFAULT_DECK_COLOR)
            if deck is None:
                next_sort_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM decks").fetchone()[0]
                conn.execute(
                    "INSERT INTO decks(name, description, color, sort_order, updated_at) VALUES (?, '', ?, ?, CURRENT_TIMESTAMP)",
                    (deck_name, default_color, next_sort_order),
                )
                created.append(deck_name)
            elif not normalize_text(deck["color"]) or normalize_text(deck["color"]) == LEGACY_DEFAULT_DECK_COLORS.get(deck_name, ""):
                conn.execute(
                    "UPDATE decks SET color = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (default_color, deck["id"]),
                )
        return created

    def ensure_default_catalog(self, excel_path: str | os.PathLike[str]) -> dict[str, Any] | None:
        excel = Path(excel_path)
        if not excel.exists():
            return None
        with self.connect_catalog() as conn:
            card_count = conn.execute("SELECT COUNT(*) FROM cards").fetchone()[0]
        if card_count > 0:
            return None
        return self.import_catalog_from_excel(excel)

    def ensure_default_decks(self) -> list[str]:
        account_id = self._resolve_account_id()
        with self.connect_account(account_id) as conn:
            created = self._ensure_default_decks_for_account(conn)
        return created

    def stats(self) -> dict[str, Any]:
        account_id = self._resolve_account_id()
        with self.connect_account(account_id) as conn:
            card_catalog = conn.execute("SELECT COUNT(*) FROM shared.cards").fetchone()[0]
            total_free = conn.execute("SELECT COALESCE(SUM(quantity), 0) FROM free_inventory").fetchone()[0]
            total_in_decks = conn.execute("SELECT COALESCE(SUM(quantity), 0) FROM deck_cards").fetchone()[0]
            deck_count = conn.execute("SELECT COUNT(*) FROM decks").fetchone()[0]
        with self.connect_catalog() as conn:
            account_row = conn.execute("SELECT id, name FROM accounts WHERE id = ?", (account_id,)).fetchone()
        return {
            "catalogCount": card_catalog,
            "freeCount": total_free,
            "deckCount": deck_count,
            "inDeckCount": total_in_decks,
            "ownedCount": total_free + total_in_decks,
            "account": dict(account_row) if account_row else None,
        }

    def list_decks(self) -> list[dict[str, Any]]:
        with self.connect_current_account() as conn:
            rows = conn.execute(
                """
                SELECT d.id,
                       d.name,
                       d.description,
                       d.color,
                       d.sort_order,
                       d.created_at,
                       d.updated_at,
                       COALESCE((SELECT SUM(dc.quantity) FROM deck_cards dc WHERE dc.deck_id = d.id), 0)
                       + COALESCE((SELECT SUM(dbe.quantity) FROM deck_basic_energies dbe WHERE dbe.deck_id = d.id), 0) AS card_count
                FROM decks d
                ORDER BY d.sort_order ASC, d.id ASC
                """,
            ).fetchall()
        return [dict(row) | {"cardCount": row["card_count"], "sortOrder": row["sort_order"]} for row in rows]

    def holdings_report(self) -> dict[str, Any]:
        with self.connect_current_account() as conn:
            deck_rows = conn.execute(
                "SELECT id, name, description, color, sort_order FROM decks ORDER BY sort_order ASC, id ASC"
            ).fetchall()
            group_order_rows = conn.execute(
                "SELECT section_key, group_key, sort_order FROM holdings_group_orders WHERE sort_order > 0"
            ).fetchall()
            cards = conn.execute(
                self._search_select_sql() + " ORDER BY c.card_name COLLATE NOCASE ASC, c.product_code ASC, c.card_code ASC, c.id ASC"
            ).fetchall()
            deck_quantities_rows = conn.execute(
                """
                SELECT dc.card_id, d.name AS deck_name, dc.quantity
                FROM deck_cards dc
                JOIN decks d ON d.id = dc.deck_id
                WHERE dc.quantity > 0
                """
            ).fetchall()

        deck_names = [row["name"] for row in deck_rows]
        group_order_map = {
            (row["section_key"], row["group_key"]): int(row["sort_order"] or 0)
            for row in group_order_rows
        }
        deck_quantity_map: dict[int, dict[str, int]] = {}
        for row in deck_quantities_rows:
            deck_quantity_map.setdefault(row["card_id"], {})[row["deck_name"]] = row["quantity"]

        sections = [
            {"key": key, "title": title, "groups": []}
            for key, title in CARD_CATEGORY_DEFINITIONS
        ]
        section_lookup = {section["key"]: section for section in sections}
        group_lookup: dict[str, dict[str, Any]] = {}

        for row in cards:
            item = self._summary_from_row(row)
            item["categoryKey"], item["categoryTitle"] = classify_card(row)
            if item["ownedQuantity"] <= 0:
                continue
            item["deckQuantities"] = {
                deck_name: deck_quantity_map.get(item["id"], {}).get(deck_name, 0)
                for deck_name in deck_names
            }
            item["visibleDeckQuantity"] = sum(item["deckQuantities"].values())
            section = section_lookup[item["categoryKey"]]
            group_key = build_holdings_group_key(item["categoryKey"], item["cardName"])
            group = group_lookup.get(group_key)
            if group is None:
                group = {
                    "groupKey": group_key,
                    "groupSortOrder": group_order_map.get((section["key"], group_key), 0),
                    "groupBaseName": normalize_text(remove_card_name_variants(item["cardName"])) or item["cardName"],
                    "groupName": normalize_text(remove_card_name_variants(item["cardName"])) or item["cardName"],
                    "items": [],
                }
                group_lookup[group_key] = group
                section["groups"].append(group)
            group["items"].append(item)

        for section in sections:
            section["groups"] = [group for group in section["groups"] if group["items"]]
            for group in section["groups"]:
                group["groupName"] = format_holdings_group_name(section["key"], group["groupBaseName"], group["items"])
                group["items"].sort(key=lambda item: build_holdings_item_sort_key(section["key"], item))
            section["groups"].sort(
                key=lambda group: build_holdings_group_sort_key(section["key"], group)
            )
            for group in section["groups"]:
                group.pop("groupBaseName", None)

        return {
            "deckNames": deck_names,
            "decks": [dict(row) for row in deck_rows],
            "sections": [section for section in sections if section["groups"]],
        }

    def update_inventory_table_group_quantities(self, group_key: str, cards: list[dict[str, Any]]) -> dict[str, Any]:
        clean_group_key = normalize_text(group_key)
        if not clean_group_key:
            raise ServiceError("卡牌分组不能为空")
        if not isinstance(cards, list) or not cards:
            raise ServiceError("卡牌列表不能为空")

        with self.connect_current_account() as conn:
            deck_rows = conn.execute(
                "SELECT id, name FROM decks ORDER BY sort_order ASC, id ASC"
            ).fetchall()
            deck_id_to_name = {int(row["id"]): row["name"] for row in deck_rows}
            expected_deck_ids = set(deck_id_to_name)
            normalized_updates: list[tuple[int, int, dict[int, int], int]] = []
            seen_card_ids: set[int] = set()

            for order_index, item in enumerate(cards, start=1):
                if not isinstance(item, dict):
                    raise ServiceError("卡牌数据格式不正确")

                card_id = int(item.get("id", 0) or 0)
                if card_id <= 0:
                    raise ServiceError("卡牌 ID 不正确")
                if card_id in seen_card_ids:
                    raise ServiceError("卡牌不能重复提交")
                seen_card_ids.add(card_id)

                free_quantity = parse_non_negative_int(item.get("freeQuantity", 0), field_name="空闲数量")
                raw_deck_quantities = item.get("deckQuantities", [])
                if not isinstance(raw_deck_quantities, list):
                    raise ServiceError("卡组数量格式不正确")

                deck_quantities: dict[int, int] = {}
                for deck_entry in raw_deck_quantities:
                    if not isinstance(deck_entry, dict):
                        raise ServiceError("卡组数量格式不正确")
                    deck_id = int(deck_entry.get("deckId", 0) or 0)
                    if deck_id not in expected_deck_ids:
                        raise NotFoundError("卡组不存在")
                    if deck_id in deck_quantities:
                        raise ServiceError(f"卡组“{deck_id_to_name[deck_id]}”数量重复")
                    deck_quantities[deck_id] = parse_non_negative_int(deck_entry.get("quantity", 0), field_name="卡组数量")

                if set(deck_quantities) != expected_deck_ids:
                    missing_names = [deck_id_to_name[deck_id] for deck_id in deck_id_to_name if deck_id not in deck_quantities]
                    raise ServiceError(f"缺少卡组数量：{'、'.join(missing_names)}")

                row = conn.execute(
                    "SELECT id, card_name, card_type, detail, special_text, attribute FROM shared.cards WHERE id = ?",
                    (card_id,),
                ).fetchone()
                if row is None:
                    raise NotFoundError("卡牌不存在")

                category_key, _ = classify_card(row)
                current_group_key = build_holdings_group_key(category_key, row["card_name"])
                if current_group_key != clean_group_key:
                    raise ServiceError("提交的卡牌不属于当前分组")

                normalized_updates.append((card_id, free_quantity, deck_quantities, order_index))

            section_key = clean_group_key.split("::", 1)[0]
            if section_key != "basic_energy":
                for deck_id, deck_name in deck_id_to_name.items():
                    total_quantity = sum(deck_quantities[deck_id] for _card_id, _free_quantity, deck_quantities, _order_index in normalized_updates)
                    if total_quantity > 4:
                        raise ServiceError(f"卡组“{deck_name}”中的同名卡牌合计不能超过 4 张（基本能量除外）")

            for card_id, free_quantity, deck_quantities, order_index in normalized_updates:
                self._set_free_quantity(conn, card_id, free_quantity)
                for deck_id, quantity in deck_quantities.items():
                    self._set_deck_quantity(conn, deck_id, card_id, quantity)
                self._set_card_group_sort_order(conn, card_id, order_index)

        return {
            "groupKey": clean_group_key,
            "updatedCount": len(normalized_updates),
        }

    def reorder_inventory_table_groups(self, section_key: str, group_keys: list[str]) -> dict[str, Any]:
        clean_section_key = normalize_text(section_key)
        if clean_section_key not in {key for key, _title in CARD_CATEGORY_DEFINITIONS}:
            raise ServiceError("卡牌分类不正确")
        if not isinstance(group_keys, list) or not group_keys:
            raise ServiceError("卡牌分组顺序不能为空")

        report = self.holdings_report()
        section = next((item for item in report["sections"] if item["key"] == clean_section_key), None)
        if section is None:
            raise NotFoundError("卡牌分类不存在")

        current_group_keys = [normalize_text(group["groupKey"]) for group in section["groups"]]
        clean_group_keys = [normalize_text(group_key) for group_key in group_keys]
        if any(not group_key for group_key in clean_group_keys):
            raise ServiceError("卡牌分组顺序包含空值")
        if len(set(clean_group_keys)) != len(clean_group_keys):
            raise ServiceError("卡牌分组顺序包含重复项")
        if sorted(clean_group_keys) != sorted(current_group_keys) or len(clean_group_keys) != len(current_group_keys):
            raise ServiceError("卡牌分组顺序不完整或包含无效分组")

        with self.connect_current_account() as conn:
            conn.execute("DELETE FROM holdings_group_orders WHERE section_key = ?", (clean_section_key,))
            for sort_order, group_key in enumerate(clean_group_keys, start=1):
                conn.execute(
                    "INSERT INTO holdings_group_orders(section_key, group_key, sort_order) VALUES (?, ?, ?)",
                    (clean_section_key, group_key, sort_order),
                )

        return {
            "sectionKey": clean_section_key,
            "updatedCount": len(clean_group_keys),
        }

    def create_deck(self, name: str, description: str = "", color: str = "") -> dict[str, Any]:
        clean_name = normalize_text(name)
        if not clean_name:
            raise ServiceError("卡组名称不能为空")
        clean_description = normalize_text(description)
        clean_color = normalize_text(color) or DEFAULT_DECK_COLORS.get(clean_name, DEFAULT_DECK_COLOR)
        with self.connect_current_account() as conn:
            try:
                next_sort_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM decks").fetchone()[0]
                cursor = conn.execute(
                    "INSERT INTO decks(name, description, color, sort_order, updated_at) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
                    (clean_name, clean_description, clean_color, next_sort_order),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError(f"卡组“{clean_name}”已存在") from exc
            deck_id = cursor.lastrowid
        return self.get_deck(deck_id)

    def get_deck(self, deck_id: int) -> dict[str, Any]:
        with self.connect_current_account() as conn:
            row = conn.execute(
                """
                SELECT d.id,
                       d.name,
                       d.description,
                       d.color,
                       d.created_at,
                       d.updated_at,
                       COALESCE((SELECT SUM(dc.quantity) FROM deck_cards dc WHERE dc.deck_id = d.id), 0)
                       + COALESCE((SELECT SUM(dbe.quantity) FROM deck_basic_energies dbe WHERE dbe.deck_id = d.id), 0) AS card_count
                FROM decks d
                  WHERE d.id = ?
                """,
                  (deck_id,),
            ).fetchone()
        if row is None:
            raise NotFoundError("卡组不存在")
        result = dict(row)
        result["cardCount"] = row["card_count"]
        return result

    def get_deck_detail(self, deck_id: int) -> dict[str, Any]:
        deck = self.get_deck(deck_id)
        with self.connect_current_account() as conn:
            rows = conn.execute(
                """
                SELECT c.id,
                       c.source_key,
                       c.product_name,
                       c.product_code,
                       c.card_code,
                       c.card_name,
                       c.card_type,
                       c.detail,
                       c.special_text,
                       c.attribute,
                      c.attribute_color,
                       c.rarity,
                       c.regulation,
                       c.note,
                       c.nickname,
                       c.show_nickname,
                                             COALESCE(hco.sort_order, c.group_sort_order, 0) AS group_sort_order,
                       COALESCE(fi.quantity, 0) AS free_quantity,
                                             dc.quantity AS deck_quantity,
                                             COALESCE(dc.backup_quantity, 0) AS backup_quantity
                FROM deck_cards dc
                                JOIN shared.cards c ON c.id = dc.card_id
                                LEFT JOIN holdings_card_orders hco ON hco.card_id = c.id
                                LEFT JOIN free_inventory fi ON fi.card_id = c.id
                                WHERE dc.deck_id = ?
                ORDER BY c.card_name COLLATE NOCASE ASC, c.product_code ASC, c.card_code ASC, c.rarity ASC, c.id ASC
                """,
                                (deck_id,),
            ).fetchall()
            section_order_rows = conn.execute(
                """
                SELECT section_key, entry_key, sort_order
                FROM deck_section_orders
                WHERE deck_id = ? AND sort_order > 0
                ORDER BY section_key ASC, sort_order ASC, entry_key ASC
                """,
                (deck_id,),
            ).fetchall()
            basic_energy_rows = conn.execute(
                """
                SELECT energy_code, quantity
                FROM deck_basic_energies
                WHERE deck_id = ?
                ORDER BY energy_code ASC
                """,
                (deck_id,),
            ).fetchall()
        cards: list[dict[str, Any]] = []
        backup_cards: list[dict[str, Any]] = []
        backup_card_count = 0
        for row in rows:
            summary = self._summary_from_row(row)
            category_key, category_title = classify_card(row)
            same_name_group_key = build_holdings_group_key(category_key, row["card_name"])
            same_name_group_base_name = normalize_text(remove_card_name_variants(row["card_name"])) or row["card_name"]
            total_quantity = int(row["deck_quantity"])
            backup_quantity = max(0, min(int(row["backup_quantity"] or 0), total_quantity))
            main_quantity = total_quantity - backup_quantity

            if main_quantity > 0:
                main_entry = dict(summary)
                main_entry["categoryKey"] = category_key
                main_entry["categoryTitle"] = category_title
                main_entry["deckEntryType"] = "catalog_card"
                main_entry["deckQuantity"] = main_quantity
                main_entry["totalDeckQuantity"] = total_quantity
                main_entry["currentBackupQuantity"] = backup_quantity
                main_entry["backupCap"] = total_quantity
                main_entry["showBackupControl"] = True
                main_entry["deckEntryKind"] = "main"
                main_entry["sameNameGroupKey"] = same_name_group_key
                main_entry["sameNameCategoryKey"] = category_key
                main_entry["sameNameGroupBaseName"] = same_name_group_base_name
                cards.append(main_entry)

            if backup_quantity > 0:
                backup_entry = dict(summary)
                backup_entry["categoryKey"] = "backup"
                backup_entry["categoryTitle"] = "备卡"
                backup_entry["deckEntryType"] = "catalog_card"
                backup_entry["deckQuantity"] = backup_quantity
                backup_entry["totalDeckQuantity"] = total_quantity
                backup_entry["currentBackupQuantity"] = backup_quantity
                backup_entry["backupCap"] = total_quantity
                backup_entry["showBackupControl"] = main_quantity <= 0
                backup_entry["isBackup"] = True
                backup_entry["deckEntryKind"] = "backup"
                backup_entry["sameNameGroupKey"] = same_name_group_key
                backup_entry["sameNameCategoryKey"] = category_key
                backup_entry["sameNameGroupBaseName"] = same_name_group_base_name
                backup_cards.append(backup_entry)
                backup_card_count += backup_quantity

        basic_energy_settings = build_deck_basic_energy_settings(basic_energy_rows)
        basic_energy_cards = build_deck_basic_energy_cards(deck_id, basic_energy_settings)
        all_cards = cards + basic_energy_cards + backup_cards
        section_order_map = {
            (normalize_text(row["section_key"]), normalize_text(row["entry_key"])): int(row["sort_order"] or 0)
            for row in section_order_rows
        }
        main_card_count = deck["cardCount"] - backup_card_count
        return deck | {
            "cards": all_cards,
            "basicEnergies": basic_energy_settings,
            "mainCardCount": main_card_count,
            "backupCardCount": backup_card_count,
            "deckChecks": build_deck_checks(all_cards, main_card_count, backup_card_count),
            "sections": build_deck_detail_sections(all_cards, section_order_map),
        }

    def update_deck_basic_energies(self, deck_id: int, items: list[dict[str, Any]]) -> dict[str, Any]:
        if not isinstance(items, list):
            raise ServiceError("基础能量数据格式不正确")

        with self.connect_current_account() as conn:
            self._ensure_deck_exists(conn, deck_id)
            quantities = {definition["code"]: 0 for definition in BASIC_DECK_ENERGY_DEFINITIONS}
            seen_codes: set[str] = set()

            for item in items:
                if not isinstance(item, dict):
                    raise ServiceError("基础能量数据格式不正确")
                code = normalize_text(item.get("code", "")).upper()
                if code not in BASIC_DECK_ENERGY_LOOKUP:
                    raise ServiceError(f"不支持的基础能量编号：{code or '-'}")
                if code in seen_codes:
                    raise ServiceError(f"基础能量编号重复：{code}")
                seen_codes.add(code)
                quantities[code] = parse_non_negative_int(item.get("quantity", 0), field_name=f"{code} 数量")

            for code, quantity in quantities.items():
                self._set_deck_basic_energy_quantity(conn, deck_id, code, quantity)

        return self.get_deck_detail(deck_id)

    def update_deck(self, deck_id: int, name: str, description: str = "", color: str = "") -> dict[str, Any]:
        clean_name = normalize_text(name)
        if not clean_name:
            raise ServiceError("卡组名称不能为空")
        clean_description = normalize_text(description)
        clean_color = normalize_text(color) or DEFAULT_DECK_COLORS.get(clean_name, DEFAULT_DECK_COLOR)
        with self.connect_current_account() as conn:
            exists = conn.execute("SELECT id FROM decks WHERE id = ?", (deck_id,)).fetchone()
            if exists is None:
                raise NotFoundError("卡组不存在")
            try:
                conn.execute(
                    "UPDATE decks SET name = ?, description = ?, color = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (clean_name, clean_description, clean_color, deck_id),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError(f"卡组“{clean_name}”已存在") from exc
        return self.get_deck(deck_id)

    def delete_deck(self, deck_id: int):
        with self.connect_current_account() as conn:
            deck = conn.execute("SELECT id FROM decks WHERE id = ?", (deck_id,)).fetchone()
            if deck is None:
                raise NotFoundError("卡组不存在")
            deck_cards = conn.execute(
                "SELECT card_id, quantity FROM deck_cards WHERE deck_id = ?",
                (deck_id,),
            ).fetchall()
            for row in deck_cards:
                current_free = self._get_free_quantity(conn, row["card_id"])
                self._set_free_quantity(conn, row["card_id"], current_free + int(row["quantity"]))
            conn.execute("DELETE FROM decks WHERE id = ?", (deck_id,))

    def reorder_decks(self, deck_ids: list[int]) -> list[dict[str, Any]]:
        clean_ids = [int(deck_id) for deck_id in deck_ids]
        if not clean_ids:
            raise ServiceError("卡组顺序不能为空")

        with self.connect_current_account() as conn:
            rows = conn.execute("SELECT id FROM decks ORDER BY id ASC").fetchall()
            existing_ids = [row["id"] for row in rows]
            if sorted(clean_ids) != sorted(existing_ids) or len(clean_ids) != len(existing_ids):
                raise ServiceError("卡组顺序不完整或包含无效卡组")

            for index, deck_id in enumerate(clean_ids, start=1):
                conn.execute(
                    "UPDATE decks SET sort_order = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (index, deck_id),
                )

        return self.list_decks()

    def reorder_deck_detail_group(self, deck_id: int, group_key: str, card_ids: list[int]) -> dict[str, Any]:
        clean_group_key = normalize_text(group_key)
        if not clean_group_key:
            raise ServiceError("卡牌分组不能为空")
        if not isinstance(card_ids, list) or not card_ids:
            raise ServiceError("卡牌顺序不能为空")

        with self.connect_current_account() as conn:
            self._ensure_deck_exists(conn, deck_id)
            rows = conn.execute(
                """
                SELECT c.id,
                       c.card_name,
                       c.card_type,
                       c.detail,
                       c.special_text
                FROM deck_cards dc
            JOIN shared.cards c ON c.id = dc.card_id
                WHERE dc.deck_id = ? AND dc.quantity > 0
                """,
                (deck_id,),
            ).fetchall()
            if not rows:
                raise NotFoundError("卡组中没有可排序的卡牌")

            rows_by_id = {int(row["id"]): row for row in rows}
            group_rows = [
                row
                for row in rows
                if build_holdings_group_key(classify_card(row)[0], row["card_name"]) == clean_group_key
            ]
            if not group_rows:
                raise NotFoundError("卡牌分组不存在")

            clean_card_ids: list[int] = []
            seen_card_ids: set[int] = set()
            for raw_card_id in card_ids:
                card_id = int(raw_card_id or 0)
                if card_id <= 0:
                    raise ServiceError("卡牌顺序包含无效卡牌")
                if card_id in seen_card_ids:
                    raise ServiceError("卡牌顺序包含重复项")
                seen_card_ids.add(card_id)

                row = rows_by_id.get(card_id)
                if row is None:
                    raise ServiceError("卡牌顺序不完整或包含无效卡牌")

                row_category_key, _ = classify_card(row)
                if build_holdings_group_key(row_category_key, row["card_name"]) != clean_group_key:
                    raise ServiceError("提交的卡牌不属于当前分组")
                clean_card_ids.append(card_id)

            current_group_ids = [int(row["id"]) for row in group_rows]
            if sorted(clean_card_ids) != sorted(current_group_ids) or len(clean_card_ids) != len(current_group_ids):
                raise ServiceError("卡牌顺序不完整或包含无效卡牌")

            for order_index, card_id in enumerate(clean_card_ids, start=1):
                self._set_card_group_sort_order(conn, card_id, order_index)

        return self.get_deck_detail(deck_id)

    def reorder_deck_detail_section(self, deck_id: int, section_key: str, entry_keys: list[str]) -> dict[str, Any]:
        clean_section_key = normalize_text(section_key)
        allowed_section_keys = {key for key, _title, _column in DECK_DETAIL_SECTION_DEFINITIONS}
        if clean_section_key not in allowed_section_keys:
            raise ServiceError("卡牌分区不存在")
        if not isinstance(entry_keys, list) or not entry_keys:
            raise ServiceError("卡牌顺序不能为空")

        detail = self.get_deck_detail(deck_id)
        section = next((item for item in detail["sections"] if item["key"] == clean_section_key), None)
        if section is None:
            raise NotFoundError("卡牌分区不存在")

        clean_entry_keys: list[str] = []
        seen_entry_keys: set[str] = set()
        for raw_entry_key in entry_keys:
            entry_key = normalize_text(raw_entry_key)
            if not entry_key:
                raise ServiceError("卡牌顺序包含无效卡牌")
            if entry_key in seen_entry_keys:
                raise ServiceError("卡牌顺序包含重复项")
            seen_entry_keys.add(entry_key)
            clean_entry_keys.append(entry_key)

        current_entry_keys = [normalize_text(item.get("deckSectionEntryKey", "")) for item in section["items"]]
        if sorted(clean_entry_keys) != sorted(current_entry_keys) or len(clean_entry_keys) != len(current_entry_keys):
            raise ServiceError("卡牌顺序不完整或包含无效卡牌")

        with self.connect_current_account() as conn:
            self._ensure_deck_exists(conn, deck_id)
            conn.execute(
                "DELETE FROM deck_section_orders WHERE deck_id = ? AND section_key = ?",
                (deck_id, clean_section_key),
            )
            for order_index, entry_key in enumerate(clean_entry_keys, start=1):
                conn.execute(
                    "INSERT INTO deck_section_orders(deck_id, section_key, entry_key, sort_order) VALUES (?, ?, ?, ?)",
                    (deck_id, clean_section_key, entry_key, order_index),
                )

        return self.get_deck_detail(deck_id)

    def list_search_regulations(self) -> list[str]:
        with self.connect_catalog() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT TRIM(regulation) AS regulation
                FROM cards
                WHERE TRIM(COALESCE(regulation, '')) <> ''
                ORDER BY regulation COLLATE NOCASE ASC
                """
            ).fetchall()
        return [row["regulation"] for row in rows]

    def get_search_preferences(self) -> dict[str, Any]:
        with self.connect_current_account() as conn:
            payload = self._get_user_setting_json(conn, USER_SETTINGS_SEARCH_PREFERENCES_KEY)
        return sanitize_search_preferences_payload(payload)

    def update_search_preferences(self, selected_regulations: list[str], consider_same_name_regulation: bool) -> dict[str, Any]:
        preferences = sanitize_search_preferences_payload(
            {
                "selectedRegulations": selected_regulations,
                "considerSameNameRegulation": consider_same_name_regulation,
            }
        )
        with self.connect_current_account() as conn:
            self._set_user_setting_json(conn, USER_SETTINGS_SEARCH_PREFERENCES_KEY, preferences)
        return preferences

    def search_cards(
        self,
        query: str,
        limit: int = 100,
        regulations: list[str] | None = None,
        consider_same_name_regulation: bool = False,
    ) -> list[dict[str, Any]]:
        clean_query = normalize_text(query)
        if not clean_query:
            return []

        selected_regulations: list[str] = []
        seen_regulations: set[str] = set()
        for regulation in regulations or []:
            clean_regulation = normalize_text(regulation)
            if not clean_regulation or clean_regulation in seen_regulations:
                continue
            seen_regulations.add(clean_regulation)
            selected_regulations.append(clean_regulation)

        with self.connect_current_account() as conn:
            rows = conn.execute(
                self._search_select_sql() + " ORDER BY c.card_name COLLATE NOCASE ASC, c.id ASC"
            ).fetchall()

        items = [self._summary_from_row(row) for row in rows]
        if selected_regulations:
            allowed_regulations = set(selected_regulations)
            if consider_same_name_regulation:
                same_name_regulations: dict[str, set[str]] = {}
                for item in items:
                    regulation = normalize_text(item.get("regulation", ""))
                    if not regulation:
                        continue
                    # 仅训练家（含支援者）、能量参与同名赛制组；宝可梦不设组，直接按自身赛制筛
                    card_type = normalize_text(item.get("cardType", "") or "")
                    if card_type in ("训练家", "支援者", "能量"):
                        same_name_regulations.setdefault(build_search_item_same_name_key(item), set()).add(regulation)
                # 过滤：训练家/能量按同名组赛制; 宝可梦只要求自身赛制符合
                items = [
                    item
                    for item in items
                    if (
                        normalize_text(item.get("cardType", "") or "") in ("训练家", "支援者", "能量")
                        and same_name_regulations.get(build_search_item_same_name_key(item), set()) & allowed_regulations
                    ) or (
                        normalize_text(item.get("cardType", "") or "") not in ("训练家", "支援者", "能量")
                        and normalize_text(item.get("regulation", "")) in allowed_regulations
                    )
                ]
            else:
                items = [item for item in items if normalize_text(item.get("regulation", "")) in allowed_regulations]

        exact_match = EXACT_CODE_PATTERN.match(clean_query)
        if exact_match:
            normalized_query = f"{normalize_text(exact_match.group(1)).upper()}-{normalize_text(exact_match.group(2)).upper()}"
            matched = [item for item in items if normalized_query in build_card_exact_search_keys(item)]
            return matched[:limit]

        query_text = clean_query.casefold()
        matched = [item for item in items if query_matches_card(item, query_text)]
        return matched[:limit]

    def get_card(self, card_id: int) -> dict[str, Any]:
        with self.connect_current_account() as conn:
            row = conn.execute(
                self._search_select_sql() + " WHERE c.id = ?",
                (card_id,),
            ).fetchone()
            if row is None:
                raise NotFoundError("卡牌不存在")
            decks = conn.execute(
                """
                SELECT d.id AS deck_id, d.name AS deck_name, dc.quantity, COALESCE(dc.backup_quantity, 0) AS backup_quantity
                FROM deck_cards dc
                JOIN decks d ON d.id = dc.deck_id
                WHERE dc.card_id = ? AND dc.quantity > 0
                ORDER BY d.name COLLATE NOCASE ASC
                """,
                (card_id,),
            ).fetchall()
        card = self._summary_from_row(row)
        card["deckBreakdown"] = []
        for item in decks:
            entry = {
                "deckId": item["deck_id"],
                "deckName": item["deck_name"],
                "quantity": item["quantity"],
            }
            if int(item["backup_quantity"] or 0) > 0:
                entry["backupQuantity"] = item["backup_quantity"]
            card["deckBreakdown"].append(entry)
        return card

    def adjust_free_inventory(self, card_id: int, delta: int) -> dict[str, Any]:
        if delta == 0:
            raise ServiceError("数量变更不能为 0")
        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            current = self._get_free_quantity(conn, card_id)
            target = current + delta
            if target < 0:
                raise ServiceError("空闲库存不足，无法减少")
            self._set_free_quantity(conn, card_id, target)
            self._ensure_sort_order_on_first_add(conn, card_id)
        return self.get_card(card_id)

    def set_free_inventory_quantity(self, card_id: int, quantity: int) -> dict[str, Any]:
        clean_quantity = parse_non_negative_int(quantity, field_name="空闲库存")
        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            self._set_free_quantity(conn, card_id, clean_quantity)
            self._ensure_sort_order_on_first_add(conn, card_id)
        return self.get_card(card_id)

    def add_to_deck(self, card_id: int, deck_id: int, amount: int, consume_free: bool = False) -> dict[str, Any]:
        amount = parse_positive_int(amount, field_name="数量")
        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            self._ensure_deck_exists(conn, deck_id)
            if consume_free:
                current_free = self._get_free_quantity(conn, card_id)
                if current_free < amount:
                    raise ServiceError("空闲库存不足，无法转入卡组")
                self._set_free_quantity(conn, card_id, current_free - amount)
            current_deck = self._get_deck_quantity(conn, deck_id, card_id)
            self._validate_deck_same_name_limit_for_card(conn, deck_id, card_id, current_deck + amount)
            self._set_deck_quantity(conn, deck_id, card_id, current_deck + amount)
        return self.get_card(card_id)

    def remove_from_deck(self, card_id: int, deck_id: int, amount: int, back_to_free: bool = False) -> dict[str, Any]:
        amount = parse_positive_int(amount, field_name="数量")
        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            self._ensure_deck_exists(conn, deck_id)
            current_deck = self._get_deck_quantity(conn, deck_id, card_id)
            if current_deck < amount:
                raise ServiceError("卡组中的库存不足，无法减少")
            self._set_deck_quantity(conn, deck_id, card_id, current_deck - amount)
            if back_to_free:
                current_free = self._get_free_quantity(conn, card_id)
                self._set_free_quantity(conn, card_id, current_free + amount)
        return self.get_card(card_id)

    def update_deck_backup_quantity(self, deck_id: int, card_id: int, quantity: int) -> dict[str, Any]:
        backup_quantity = parse_non_negative_int(quantity, field_name="备卡数量")
        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            self._ensure_deck_exists(conn, deck_id)
            current_quantity = self._get_deck_quantity(conn, deck_id, card_id)
            if current_quantity <= 0:
                raise NotFoundError("卡组中不存在这张卡")
            if backup_quantity > current_quantity:
                raise ServiceError("备卡数量不能超过卡组中的总数量")
            self._set_deck_backup_quantity(conn, deck_id, card_id, backup_quantity)
        return self.get_deck_detail(deck_id)

    def apply_deck_card_quantity_action(
        self,
        deck_id: int,
        card_id: int,
        entry_type: str,
        mode: str,
        target_quantity: int,
    ) -> dict[str, Any]:
        clean_entry_type = normalize_text(entry_type).lower()
        if clean_entry_type not in {"main", "backup"}:
            raise ServiceError("卡牌条目类型不正确")

        clean_mode = normalize_text(mode).lower()
        if clean_mode not in {"add_direct", "add_from_free", "back_to_free", "remove"}:
            raise ServiceError("数量操作类型不正确")

        clean_target_quantity = parse_non_negative_int(target_quantity, field_name="目标数量")

        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)
            self._ensure_deck_exists(conn, deck_id)

            total_quantity = self._get_deck_quantity(conn, deck_id, card_id)
            if total_quantity <= 0:
                raise NotFoundError("卡组中不存在这张卡")

            backup_quantity = self._get_deck_backup_quantity(conn, deck_id, card_id)
            main_quantity = total_quantity - backup_quantity
            current_entry_quantity = backup_quantity if clean_entry_type == "backup" else main_quantity

            if current_entry_quantity <= 0:
                if clean_entry_type == "backup":
                    raise NotFoundError("卡组备卡中不存在这张卡")
                raise NotFoundError("卡组主牌中不存在这张卡")

            delta = clean_target_quantity - current_entry_quantity
            if delta == 0:
                raise ServiceError("目标数量与当前数量相同，无需调整")
            if clean_mode in {"add_direct", "add_from_free"} and delta < 0:
                raise ServiceError("目标数量更小，请使用“转回空闲到此数量”或“直接删到此数量”")
            if clean_mode in {"back_to_free", "remove"} and delta > 0:
                raise ServiceError("目标数量更大，请使用“直接补到此数量”或“从空闲补到此数量”")

            amount = abs(delta)
            current_free = self._get_free_quantity(conn, card_id)
            if clean_mode == "add_from_free":
                if current_free < amount:
                    raise ServiceError("空闲库存不足，无法转入卡组")
                self._set_free_quantity(conn, card_id, current_free - amount)
            elif clean_mode == "back_to_free":
                self._set_free_quantity(conn, card_id, current_free + amount)

            new_total_quantity = total_quantity + delta
            new_backup_quantity = backup_quantity + delta if clean_entry_type == "backup" else backup_quantity
            if new_total_quantity < 0 or new_backup_quantity < 0 or new_backup_quantity > new_total_quantity:
                raise ServiceError("调整后的卡组数量不合法")

            self._validate_deck_same_name_limit_for_card(conn, deck_id, card_id, new_total_quantity)
            self._set_deck_quantity(conn, deck_id, card_id, new_total_quantity, backup_quantity=new_backup_quantity)

        return self.get_deck_detail(deck_id)

    def move_deck_cards_to_free(self, deck_id: int) -> dict[str, Any]:
        """将卡组中所有卡牌转回空闲库存。"""
        with self.connect_current_account() as conn:
            self._ensure_deck_exists(conn, deck_id)
            rows = conn.execute(
                "SELECT card_id, quantity FROM deck_cards WHERE deck_id = ? AND quantity > 0",
                (deck_id,),
            ).fetchall()
            for row in rows:
                current_free = self._get_free_quantity(conn, row["card_id"])
                self._set_free_quantity(conn, row["card_id"], current_free + int(row["quantity"]))
            conn.execute("DELETE FROM deck_cards WHERE deck_id = ?", (deck_id,))
        return self.get_deck_detail(deck_id)

    def adjust_total_quantity(self, card_id: int, delta: int) -> dict[str, Any]:
        if delta == 0:
            raise ServiceError("数量变更不能为 0")

        with self.connect_current_account() as conn:
            self._ensure_card_exists(conn, card_id)

            if delta > 0:
                current_free = self._get_free_quantity(conn, card_id)
                self._set_free_quantity(conn, card_id, current_free + delta)
            else:
                remaining = abs(delta)
                current_free = self._get_free_quantity(conn, card_id)
                free_removed = min(current_free, remaining)
                if free_removed > 0:
                    self._set_free_quantity(conn, card_id, current_free - free_removed)
                    remaining -= free_removed

                if remaining > 0:
                    deck_rows = conn.execute(
                        """
                        SELECT dc.deck_id, dc.quantity
                        FROM deck_cards dc
                        JOIN decks d ON d.id = dc.deck_id
                        WHERE dc.card_id = ? AND dc.quantity > 0
                        ORDER BY d.sort_order ASC, d.id ASC
                        """,
                        (card_id,),
                    ).fetchall()
                    total_in_decks = sum(int(row["quantity"]) for row in deck_rows)
                    if total_in_decks < remaining:
                        raise ServiceError("总持有不足，无法减少")

                    for row in deck_rows:
                        if remaining <= 0:
                            break
                        current_deck = int(row["quantity"])
                        reduction = min(current_deck, remaining)
                        self._set_deck_quantity(conn, row["deck_id"], card_id, current_deck - reduction)
                        remaining -= reduction

        return self.get_card(card_id)

    def delete_card(self, card_id: int):
        for account_db_path in sorted(Path(self.accounts_dir).glob("*.db")):
            account_conn = self._connect_sqlite(account_db_path)
            try:
                account_conn.execute("DELETE FROM free_inventory WHERE card_id = ?", (card_id,))
                account_conn.execute("DELETE FROM deck_cards WHERE card_id = ?", (card_id,))
                account_conn.execute("DELETE FROM holdings_card_orders WHERE card_id = ?", (card_id,))
                account_conn.commit()
            except Exception:
                account_conn.rollback()
                raise
            finally:
                account_conn.close()

        with self.connect_catalog() as conn:
            self._ensure_card_exists(conn, card_id)
            conn.execute("DELETE FROM cards WHERE id = ?", (card_id,))

    # ── 退标 ──────────────────────────────────────────────────────

    def preview_retire_by_regulation(
        self, regulation: str, skip_same_name: bool = True, include_deck_cards: bool = True
    ) -> dict[str, Any]:
        """预览指定赛制下当前用户持有的所有卡牌，用于退标确认。"""
        target = regulation.strip().upper()
        if not target:
            return {"regulation": "", "decks": [], "cards": [], "totalCount": 0, "totalQuantity": 0}

        with self.connect_current_account() as conn:
            cards = conn.execute(self._search_select_sql(), ()).fetchall()
            deck_rows = conn.execute(
                "SELECT id, name, color FROM decks ORDER BY sort_order ASC, id ASC"
            ).fetchall()
            dc_rows = conn.execute(
                """
                SELECT dc.card_id, dc.deck_id, d.name AS deck_name, dc.quantity
                FROM deck_cards dc
                JOIN decks d ON d.id = dc.deck_id
                WHERE dc.quantity > 0
                """
            ).fetchall()

        deck_map: dict[int, dict[str, Any]] = {}
        for row in deck_rows:
            deck_map[row["id"]] = {"id": row["id"], "name": row["name"], "color": row["color"] or "#9ca3af"}

        # (card_id, deck_id) → quantity
        deck_card_qty: dict[tuple[int, int], int] = {}
        for row in dc_rows:
            deck_card_qty[(row["card_id"], row["deck_id"])] = row["quantity"]

        # 1) 收集候选卡：regulation 匹配 + 有库存
        candidates: list[dict[str, Any]] = []
        for row in cards:
            row_reg = normalize_text(row["regulation"]).upper()
            if row_reg != target:
                continue
            free_qty = row["free_quantity"] or 0
            deck_qty = row["deck_quantity"] or 0
            total_qty = free_qty + deck_qty
            if total_qty <= 0:
                continue
            if not include_deck_cards and deck_qty > 0:
                continue
            item = self._summary_from_row(row)
            item["categoryKey"], item["categoryTitle"] = classify_card(row)
            item["freeQuantity"] = free_qty
            item["deckQuantity"] = deck_qty
            item["ownedQuantity"] = total_qty
            item["regulation"] = row_reg
            candidates.append(item)

        # 2) 同名保护
        if skip_same_name and candidates:
            with self.connect_catalog() as cat_conn:
                other_names = {
                    normalize_text(remove_card_name_variants(r["card_name"]))
                    for r in cat_conn.execute(
                        "SELECT DISTINCT card_name FROM cards WHERE TRIM(COALESCE(regulation,'')) <> ?",
                        (target,),
                    ).fetchall()
                }
            candidates = [
                c for c in candidates
                if is_pokemon_category_key(c["categoryKey"])
                or normalize_text(remove_card_name_variants(c["cardName"])) not in other_names
            ]

        # 3) deck 分布明细
        for c in candidates:
            cid = c["id"]
            c["deckBreakdown"] = [
                {"deckId": did, "deckName": deck_map[did]["name"], "quantity": qty}
                for (card_id, did), qty in deck_card_qty.items()
                if card_id == cid
            ]

        total_qty = sum(c["ownedQuantity"] for c in candidates)
        return {
            "regulation": target,
            "decks": list(deck_map.values()),
            "cards": candidates,
            "totalCount": len(candidates),
            "totalQuantity": total_qty,
        }

    def execute_retire_cards(self, card_ids: list[int]) -> int:
        """删除当前用户库存中的指定卡牌（不删共享目录数据）。返回删除的卡牌种数。"""
        if not card_ids:
            return 0
        placeholders = ",".join("?" for _ in card_ids)
        with self.connect_current_account() as conn:
            dc = conn.execute(
                f"SELECT DISTINCT deck_id FROM deck_cards WHERE card_id IN ({placeholders})",
                tuple(card_ids),
            ).fetchall()
            conn.execute(f"DELETE FROM free_inventory WHERE card_id IN ({placeholders})", tuple(card_ids))
            conn.execute(f"DELETE FROM deck_cards WHERE card_id IN ({placeholders})", tuple(card_ids))
            conn.execute(f"DELETE FROM holdings_card_orders WHERE card_id IN ({placeholders})", tuple(card_ids))
            # 清理没有卡牌的 deck section orders
            for row in dc:
                remaining = conn.execute(
                    "SELECT COUNT(*) AS cnt FROM deck_cards WHERE deck_id = ?", (row["deck_id"],)
                ).fetchone()
                if remaining and remaining["cnt"] == 0:
                    conn.execute("DELETE FROM deck_section_orders WHERE deck_id = ?", (row["deck_id"],))
            conn.commit()
        return len(card_ids)

    def import_catalog_from_excel(self, excel_path: str | os.PathLike[str]) -> dict[str, Any]:
        excel = Path(excel_path)
        if not excel.exists():
            raise NotFoundError(f"未找到 Excel 文件：{excel}")
        workbook = load_workbook(excel, data_only=True)
        try:
            sheet = workbook.worksheets[1] if len(workbook.worksheets) >= 2 else workbook.worksheets[0]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                raise ServiceError("Excel 首行为空，无法识别表头")
            index_map = build_header_index_map(header_row)
            missing = [key for key in ("商品编号", "卡牌编号", "卡牌名称") if key not in index_map]
            if missing:
                raise ServiceError(f"Excel 缺少必要列：{'、'.join(missing)}")
            attribute_index = index_map.get("属性")

            created = 0
            updated = 0
            skipped = 0
            new_cards: list[tuple[int, str, str, str]] = []
            with self.connect_current_account() as conn:
                cards_table = self._catalog_cards_table(conn)
                for row in sheet.iter_rows(min_row=2):
                    values = tuple(cell.value for cell in row)
                    if values is None or all(value is None or normalize_text(str(value)) == "" for value in values):
                        continue
                    record = row_to_record(values, index_map)
                    if attribute_index is not None and attribute_index < len(row):
                        record["attribute_color"] = extract_excel_fill_color(row[attribute_index])
                    else:
                        record["attribute_color"] = ""
                    if not record["product_code"] or not record["card_code"] or not record["card_name"]:
                        skipped += 1
                        continue
                    source_key = build_source_key(record)
                    existing = conn.execute(f"SELECT id FROM {cards_table} WHERE source_key = ?", (source_key,)).fetchone()
                    if existing is None:
                        existing = self._find_card_by_catalog_identity(conn, record)
                    if existing is None:
                        cursor = conn.execute(
                            f"""
                            INSERT INTO {cards_table}(
                                source_key, product_name, product_code, card_code, card_name,
                                card_type, detail, special_text, attribute, attribute_color, rarity, regulation,
                                note, nickname, initial_quantity, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                            """,
                            (
                                source_key,
                                record["product_name"],
                                record["product_code"],
                                record["card_code"],
                                record["card_name"],
                                record["card_type"],
                                record["detail"],
                                record["special_text"],
                                record["attribute"],
                                record["attribute_color"],
                                record["rarity"],
                                record["regulation"],
                                record["note"],
                                record["nickname"],
                                record["quantity"],
                            ),
                        )
                        card_id = cursor.lastrowid
                        self._set_free_quantity(conn, card_id, record["quantity"])
                        new_cards.append((card_id, record["card_name"], record["product_code"], record["card_code"]))
                        created += 1
                    else:
                        card_id = existing["id"]
                        conn.execute(
                            f"""
                            UPDATE {cards_table}
                            SET source_key = ?,
                                product_name = ?,
                                product_code = ?,
                                card_code = ?,
                                card_name = ?,
                                card_type = ?,
                                detail = ?,
                                special_text = ?,
                                attribute = ?,
                                attribute_color = ?,
                                rarity = ?,
                                regulation = ?,
                                note = ?,
                                nickname = ?,
                                initial_quantity = ?,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE id = ?
                            """,
                            (
                                source_key,
                                record["product_name"],
                                record["product_code"],
                                record["card_code"],
                                record["card_name"],
                                record["card_type"],
                                record["detail"],
                                record["special_text"],
                                record["attribute"],
                                record["attribute_color"],
                                record["rarity"],
                                record["regulation"],
                                record["note"],
                                record["nickname"],
                                record["quantity"],
                                card_id,
                            ),
                        )
                        if self._get_free_quantity(conn, card_id) <= 0:
                            self._set_free_quantity(conn, card_id, record["quantity"])
                        updated += 1
            return {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "path": str(excel),
                "sheetName": sheet.title,
                "newCardIds": [c[0] for c in new_cards],
                "_new_cards": new_cards,  # 供路由层传递给爬虫（(id, name, pc, cc)）
            }
        finally:
            workbook.close()

    def export_state(self) -> dict[str, Any]:
        account_id = self._resolve_account_id()
        with self.connect_account(account_id) as conn:
            cards = conn.execute(
                self._search_select_sql() + " ORDER BY c.product_code ASC, c.card_code ASC, c.id ASC"
            ).fetchall()
            deck_rows = conn.execute(
                "SELECT id, name, description, color, sort_order FROM decks ORDER BY sort_order ASC, id ASC"
            ).fetchall()
            deck_cards = conn.execute(
                """
                SELECT deck_id, card_id, quantity, COALESCE(backup_quantity, 0) AS backup_quantity
                FROM deck_cards
                WHERE quantity > 0
                """
            ).fetchall()
            deck_basic_energy_rows = conn.execute(
                """
                SELECT deck_id, energy_code, quantity
                FROM deck_basic_energies
                WHERE quantity > 0
                ORDER BY deck_id ASC, energy_code ASC
                """
            ).fetchall()
            group_order_rows = conn.execute(
                """
                SELECT section_key, group_key, sort_order
                FROM holdings_group_orders
                WHERE sort_order > 0
                ORDER BY section_key ASC, sort_order ASC, group_key ASC
                """
            ).fetchall()
        with self.connect_catalog() as conn:
            account_row = conn.execute("SELECT id, name FROM accounts WHERE id = ?", (account_id,)).fetchone()

        deck_map: dict[int, dict[str, Any]] = {row["id"]: dict(row) for row in deck_rows}
        card_deck_map: dict[int, list[dict[str, Any]]] = {}
        deck_basic_energy_quantity_map: dict[int, list[dict[str, Any]]] = {}
        for row in deck_cards:
            entry = {
                "deckId": row["deck_id"],
                "deckName": deck_map[row["deck_id"]]["name"],
                "quantity": row["quantity"],
                "backupQuantity": row["backup_quantity"],
            }
            card_deck_map.setdefault(row["card_id"], []).append(entry)
        for row in deck_basic_energy_rows:
            deck_basic_energy_quantity_map.setdefault(row["deck_id"], []).append(
                {"energy_code": row["energy_code"], "quantity": row["quantity"]}
            )

        payload_cards = []
        for row in cards:
            summary = self._summary_from_row(row)
            payload_cards.append(
                {
                    "sourceKey": summary["sourceKey"],
                    "productCode": summary["productCode"],
                    "cardCode": summary["cardCode"],
                    "cardName": summary["cardName"],
                    "groupSortOrder": summary["groupSortOrder"],
                    "freeQuantity": summary["freeQuantity"],
                    "deckQuantities": card_deck_map.get(summary["id"], []),
                }
            )
        return {
            "version": 1,
            "exportedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "account": dict(account_row) if account_row else None,
            "decks": [
                dict(row) | {
                    "sortOrder": row["sort_order"],
                    "basicEnergies": [
                        {"code": item["code"], "name": item["name"], "quantity": item["quantity"]}
                        for item in build_deck_basic_energy_settings(deck_basic_energy_quantity_map.get(row["id"], []))
                        if item["quantity"] > 0
                    ],
                }
                for row in deck_rows
            ],
            "groupOrders": [
                {
                    "sectionKey": row["section_key"],
                    "groupKey": row["group_key"],
                    "sortOrder": row["sort_order"],
                }
                for row in group_order_rows
            ],
            "cards": payload_cards,
        }

    def import_state(self, payload: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise ServiceError("状态文件格式不正确")
        cards = payload.get("cards")
        decks = payload.get("decks")
        group_orders = payload.get("groupOrders", [])
        if not isinstance(cards, list) or not isinstance(decks, list):
            raise ServiceError("状态文件缺少 cards 或 decks 数组")

        imported_cards = 0
        skipped_cards = 0
        with self.connect_current_account() as conn:
            account_deck_ids = [row["id"] for row in conn.execute("SELECT id FROM decks").fetchall()]
            for deck_id in account_deck_ids:
                conn.execute("DELETE FROM deck_cards WHERE deck_id = ?", (deck_id,))
                conn.execute("DELETE FROM deck_basic_energies WHERE deck_id = ?", (deck_id,))
            conn.execute("DELETE FROM free_inventory")
            conn.execute("DELETE FROM holdings_group_orders")
            conn.execute("DELETE FROM holdings_card_orders")
            conn.execute("DELETE FROM decks")

            deck_name_to_id: dict[str, int] = {}
            for index, deck in enumerate(decks, start=1):
                name = normalize_text((deck or {}).get("name", ""))
                if not name:
                    continue
                description = normalize_text((deck or {}).get("description", ""))
                color = normalize_text((deck or {}).get("color", "")) or DEFAULT_DECK_COLOR
                sort_order = int((deck.get("sortOrder", deck.get("sort_order", index)) if isinstance(deck, dict) else index) or index)
                cursor = conn.execute(
                    "INSERT INTO decks(name, description, color, sort_order, updated_at) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
                    (name, description, color, sort_order),
                )
                deck_id = cursor.lastrowid
                deck_name_to_id[name] = deck_id
                for energy_entry in (deck.get("basicEnergies", []) if isinstance(deck, dict) else []):
                    if not isinstance(energy_entry, dict):
                        continue
                    code = normalize_text(energy_entry.get("code", "")).upper()
                    if code not in BASIC_DECK_ENERGY_LOOKUP:
                        continue
                    quantity = max(0, int(energy_entry.get("quantity", 0) or 0))
                    self._set_deck_basic_energy_quantity(conn, deck_id, code, quantity)

            for item in cards:
                if not isinstance(item, dict):
                    skipped_cards += 1
                    continue
                source_key = normalize_text(item.get("sourceKey", ""))
                row = None
                if source_key:
                    row = conn.execute("SELECT id FROM shared.cards WHERE source_key = ?", (source_key,)).fetchone()
                if row is None:
                    fallback = conn.execute(
                        "SELECT id FROM shared.cards WHERE product_code = ? AND card_code = ? AND card_name = ? LIMIT 1",
                        (
                            normalize_text(item.get("productCode", "")),
                            normalize_text(item.get("cardCode", "")),
                            normalize_text(item.get("cardName", "")),
                        ),
                    ).fetchone()
                    row = fallback
                if row is None:
                    skipped_cards += 1
                    continue
                card_id = row["id"]
                free_quantity = int(item.get("freeQuantity", 0) or 0)
                group_sort_order = int(item.get("groupSortOrder", 0) or 0)
                self._set_card_group_sort_order(conn, card_id, max(0, group_sort_order))
                self._set_free_quantity(conn, card_id, max(0, free_quantity))
                for deck_entry in item.get("deckQuantities", []):
                    if not isinstance(deck_entry, dict):
                        continue
                    deck_name = normalize_text(deck_entry.get("deckName", ""))
                    quantity = int(deck_entry.get("quantity", 0) or 0)
                    backup_quantity = int(deck_entry.get("backupQuantity", 0) or 0)
                    deck_id = deck_name_to_id.get(deck_name)
                    if deck_id and quantity > 0:
                        self._set_deck_quantity(conn, deck_id, card_id, quantity, backup_quantity=backup_quantity)
                imported_cards += 1

            if isinstance(group_orders, list):
                for entry in group_orders:
                    if not isinstance(entry, dict):
                        continue
                    section_key = normalize_text(entry.get("sectionKey", ""))
                    group_key = normalize_text(entry.get("groupKey", ""))
                    if section_key not in {key for key, _title in CARD_CATEGORY_DEFINITIONS} or not group_key:
                        continue
                    sort_order = max(0, int(entry.get("sortOrder", 0) or 0))
                    if sort_order <= 0:
                        continue
                    conn.execute(
                        "INSERT INTO holdings_group_orders(section_key, group_key, sort_order) VALUES (?, ?, ?) "
                        "ON CONFLICT(section_key, group_key) DO UPDATE SET sort_order = excluded.sort_order",
                        (section_key, group_key, sort_order),
                    )
        return {
            "importedCards": imported_cards,
            "skippedCards": skipped_cards,
            "deckCount": len(decks),
        }

    def _search_select_sql(self) -> str:
        return f"""
        SELECT c.id,
               c.source_key,
               c.product_name,
               c.product_code,
               c.card_code,
               c.card_name,
               c.card_type,
               c.detail,
               c.special_text,
               c.attribute,
               c.attribute_color,
               c.rarity,
               c.regulation,
               c.note,
               c.nickname,
               c.show_nickname,
             COALESCE(hco.sort_order, c.group_sort_order, 0) AS group_sort_order,
               COALESCE(fi.quantity, 0) AS free_quantity,
               COALESCE(deck_totals.deck_quantity, 0) AS deck_quantity
         FROM shared.cards c
         LEFT JOIN holdings_card_orders hco ON hco.card_id = c.id
         LEFT JOIN free_inventory fi ON fi.card_id = c.id
        LEFT JOIN (
            SELECT dc.card_id, SUM(dc.quantity) AS deck_quantity
            FROM deck_cards dc
            GROUP BY dc.card_id
        ) deck_totals ON deck_totals.card_id = c.id
        """

    # ── 纯库存导出/导入 ─────────────────────────────────────

    def export_inventory(self) -> dict[str, Any]:
        account_id = self._resolve_account_id()
        with self.connect_account(account_id) as conn:
            deck_rows = conn.execute(
                "SELECT id, name FROM decks ORDER BY sort_order ASC, id ASC"
            ).fetchall()
            cards = conn.execute(
                self._search_select_sql()
                + " ORDER BY c.product_code ASC, c.card_code ASC, c.id ASC"
            ).fetchall()
        with self.connect_catalog() as conn:
            account_row = conn.execute("SELECT id, name FROM accounts WHERE id = ?", (account_id,)).fetchone()

        payload_cards = []
        for row in cards:
            summary = self._summary_from_row(row)
            if summary["freeQuantity"] <= 0 and summary["deckQuantity"] <= 0:
                continue
            payload_cards.append({
                "sourceKey": summary["sourceKey"],
                "productCode": summary["productCode"],
                "cardCode": summary["cardCode"],
                "cardName": summary["cardName"],
                "freeQuantity": summary["freeQuantity"],
                "deckQuantity": summary["deckQuantity"],
            })
        return {
            "version": 1,
            "exportedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "account": dict(account_row) if account_row else None,
            "deckNames": [r["name"] for r in deck_rows],
            "cards": payload_cards,
        }

    def import_inventory(self, payload: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise ServiceError("库存文件格式不正确")
        cards = payload.get("cards")
        if not isinstance(cards, list):
            raise ServiceError("库存文件缺少 cards 数组")
        imported = 0
        skipped = 0
        with self.connect_current_account() as conn:
            for item in cards:
                if not isinstance(item, dict):
                    skipped += 1; continue
                source_key = normalize_text(item.get("sourceKey", ""))
                row = None
                if source_key:
                    row = conn.execute("SELECT id FROM shared.cards WHERE source_key = ?", (source_key,)).fetchone()
                if row is None:
                    row = conn.execute(
                        "SELECT id FROM shared.cards WHERE product_code = ? AND card_code = ? AND card_name = ? LIMIT 1",
                        (normalize_text(item.get("productCode", "")),
                         normalize_text(item.get("cardCode", "")),
                         normalize_text(item.get("cardName", ""))),
                    ).fetchone()
                if row is None:
                    skipped += 1; continue
                card_id = row["id"]
                free_qty = max(0, int(item.get("freeQuantity", 0) or 0))
                self._set_free_quantity(conn, card_id, free_qty)
                imported += 1
        return {"importedCards": imported, "skippedCards": skipped}

    def _summary_from_row(self, row: sqlite3.Row) -> dict[str, Any]:
        free_quantity = row["free_quantity"]
        deck_quantity = row["deck_quantity"]
        display_product_code = resolve_display_product_code(
            row["product_code"],
            row["card_code"],
            row["product_name"],
            row["card_name"],
        )
        display_card_code = format_holdings_card_code(row["card_code"])
        return {
            "id": row["id"],
            "sourceKey": row["source_key"],
            "productName": row["product_name"],
            "productCode": row["product_code"],
            "displayProductCode": display_product_code,
            "cardCode": row["card_code"],
            "displayCardCode": display_card_code,
            "displayCode": format_card_display_code(
                row["product_code"],
                row["card_code"],
                row["product_name"],
                row["card_name"],
            ),
            "cardName": row["card_name"],
            "cardType": row["card_type"],
            "detail": row["detail"],
            "special": row["special_text"],
            "attribute": row["attribute"],
            "attributeColor": normalize_hex_color(row["attribute_color"]),
            "rarity": row["rarity"],
            "regulation": row["regulation"],
            "note": row["note"],
            "groupSortOrder": row["group_sort_order"] if "group_sort_order" in row.keys() else 0,
            "freeQuantity": free_quantity,
            "deckQuantity": deck_quantity,
            "ownedQuantity": free_quantity + deck_quantity,
            "nickname": row["nickname"] if "nickname" in row.keys() else "",
            "showNickname": bool(row["show_nickname"]) if "show_nickname" in row.keys() else False,
        }

    def _find_card_by_catalog_identity(self, conn: sqlite3.Connection, record: dict[str, Any]) -> sqlite3.Row | None:
        identity_parts = build_catalog_identity_parts(record)
        return conn.execute(
                        f"""
            SELECT id
                        FROM {self._catalog_cards_table(conn)}
            WHERE UPPER(TRIM(product_code)) = ?
              AND UPPER(TRIM(card_code)) = ?
              AND TRIM(card_name) = ?
              AND TRIM(card_type) = ?
              AND TRIM(detail) = ?
              AND TRIM(special_text) = ?
              AND TRIM(attribute) = ?
              AND TRIM(rarity) = ?
              AND TRIM(regulation) = ?
            LIMIT 1
            """,
            identity_parts,
        ).fetchone()

    def _account_db_stats(self, account_id: int) -> dict[str, int]:
        account_db_path = Path(self._account_db_path(account_id))
        if not account_db_path.exists():
            return {"freeCount": 0, "inDeckCount": 0, "deckCount": 0}
        conn = self._connect_sqlite(account_db_path)
        try:
            free_count = int(conn.execute("SELECT COALESCE(SUM(quantity), 0) FROM free_inventory").fetchone()[0])
            in_deck_count = int(conn.execute("SELECT COALESCE(SUM(quantity), 0) FROM deck_cards").fetchone()[0])
            deck_count = int(conn.execute("SELECT COUNT(*) FROM decks").fetchone()[0])
        finally:
            conn.close()
        return {"freeCount": free_count, "inDeckCount": in_deck_count, "deckCount": deck_count}

    def _has_shared_catalog(self, conn: sqlite3.Connection) -> bool:
        return any(row[1] == "shared" for row in conn.execute("PRAGMA database_list").fetchall())

    def _catalog_cards_table(self, conn: sqlite3.Connection) -> str:
        return "shared.cards" if self._has_shared_catalog(conn) else "cards"

    def _table_has_column(self, conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
        return column_name in {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}

    def _get_user_setting_json(self, conn: sqlite3.Connection, key: str) -> Any:
        row = conn.execute("SELECT value FROM user_settings WHERE key = ?", (key,)).fetchone()
        if row is None:
            return None
        try:
            return json.loads(str(row["value"] or ""))
        except json.JSONDecodeError:
            return None

    def _set_user_setting_json(self, conn: sqlite3.Connection, key: str, value: Any):
        conn.execute(
            "INSERT INTO user_settings(key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = CURRENT_TIMESTAMP",
            (key, json.dumps(value, ensure_ascii=False)),
        )

    def _get_card_group_sort_order(self, conn: sqlite3.Connection, card_id: int) -> int:
        row = conn.execute("SELECT sort_order FROM holdings_card_orders WHERE card_id = ?", (card_id,)).fetchone()
        if row is not None:
            return int(row["sort_order"] or 0)
        row = conn.execute(
            f"SELECT group_sort_order FROM {self._catalog_cards_table(conn)} WHERE id = ?",
            (card_id,),
        ).fetchone()
        return int(row["group_sort_order"] or 0) if row is not None else 0

    def _set_card_group_sort_order(self, conn: sqlite3.Connection, card_id: int, sort_order: int):
        clean_sort_order = max(0, int(sort_order))
        if clean_sort_order <= 0:
            conn.execute("DELETE FROM holdings_card_orders WHERE card_id = ?", (card_id,))
            return
        conn.execute(
            "INSERT INTO holdings_card_orders(card_id, sort_order) VALUES (?, ?) "
            "ON CONFLICT(card_id) DO UPDATE SET sort_order = excluded.sort_order",
            (card_id, clean_sort_order),
        )

    def _ensure_card_exists(self, conn: sqlite3.Connection, card_id: int):
        row = conn.execute(f"SELECT id FROM {self._catalog_cards_table(conn)} WHERE id = ?", (card_id,)).fetchone()
        if row is None:
            raise NotFoundError("卡牌不存在")

    def _ensure_deck_exists(self, conn: sqlite3.Connection, deck_id: int):
        if self._table_has_column(conn, "decks", "account_id"):
            account_id = self.get_current_account_id(conn)
            row = conn.execute("SELECT id FROM decks WHERE id = ? AND account_id = ?", (deck_id, account_id)).fetchone()
        else:
            row = conn.execute("SELECT id FROM decks WHERE id = ?", (deck_id,)).fetchone()
        if row is None:
            raise NotFoundError("卡组不存在")

    def _get_free_quantity(self, conn: sqlite3.Connection, card_id: int) -> int:
        if self._table_has_column(conn, "free_inventory", "account_id"):
            account_id = self.get_current_account_id(conn)
            row = conn.execute("SELECT quantity FROM free_inventory WHERE account_id = ? AND card_id = ?", (account_id, card_id)).fetchone()
        else:
            row = conn.execute("SELECT quantity FROM free_inventory WHERE card_id = ?", (card_id,)).fetchone()
        return int(row["quantity"]) if row else 0

    def _set_free_quantity(self, conn: sqlite3.Connection, card_id: int, quantity: int):
        quantity = max(0, int(quantity))
        if self._table_has_column(conn, "free_inventory", "account_id"):
            account_id = self.get_current_account_id(conn)
            conn.execute(
                """
                INSERT INTO free_inventory(account_id, card_id, quantity) VALUES (?, ?, ?)
                ON CONFLICT(account_id, card_id) DO UPDATE SET quantity = excluded.quantity
                """,
                (account_id, card_id, quantity),
            )
        else:
            conn.execute(
                """
                INSERT INTO free_inventory(card_id, quantity) VALUES (?, ?)
                ON CONFLICT(card_id) DO UPDATE SET quantity = excluded.quantity
                """,
                (card_id, quantity),
            )

    def _ensure_sort_order_on_first_add(self, conn: sqlite3.Connection, card_id: int):
        """仅在卡牌首次有库存时（old_qty=0, new_qty>0）自动分配排序。"""
        old_qty = self._get_free_quantity(conn, card_id)
        if old_qty > 0:
            return
        if self._get_free_quantity(conn, card_id) <= 0:
            return
        card = conn.execute(
            f"SELECT card_type, detail, special_text, product_code, attribute, card_name FROM {self._catalog_cards_table(conn)} WHERE id = ?",
            (card_id,),
        ).fetchone()
        if not card:
            return
        ct = normalize_text(card["card_type"])
        if ct not in ("训练家", "支援者", "能量", "宝可梦"):
            return
        rows = conn.execute(
            f"SELECT c.id, c.card_type, c.detail, c.special_text, c.card_name, COALESCE(hco.sort_order, c.group_sort_order, 0) AS group_sort_order, c.attribute, c.product_code FROM {self._catalog_cards_table(conn)} c LEFT JOIN holdings_card_orders hco ON hco.card_id = c.id"
        ).fetchall()
        self._assign_sort_order(conn, card_id, card, rows)

    def _get_deck_quantity(self, conn: sqlite3.Connection, deck_id: int, card_id: int) -> int:
        row = conn.execute(
            "SELECT quantity FROM deck_cards WHERE deck_id = ? AND card_id = ?",
            (deck_id, card_id),
        ).fetchone()
        return int(row["quantity"]) if row else 0

    def _get_deck_backup_quantity(self, conn: sqlite3.Connection, deck_id: int, card_id: int) -> int:
        row = conn.execute(
            "SELECT backup_quantity FROM deck_cards WHERE deck_id = ? AND card_id = ?",
            (deck_id, card_id),
        ).fetchone()
        return int(row["backup_quantity"]) if row else 0

    def _assign_sort_order(
        self,
        conn: sqlite3.Connection,
        card_id: int,
        card: sqlite3.Row,
        all_rows: list[sqlite3.Row],
    ):
        """首次添加时分配排序：宝可梦按属性+发售顺序插入；训练家/能量直接追加到末尾。"""
        import bisect
        card_category_key, _ = classify_card(card)
        release_idx = _get_release_index(card["product_code"])
        attr_idx = _attribute_sort_index(card["attribute"])

        # 训练家/能量：不分配 gso（保持 0），仅处理组间排序
        if not is_pokemon_category_key(card_category_key):
            group_key = build_holdings_group_key(card_category_key, card["card_name"])
            self._insert_holdings_group_order(conn, card_category_key, group_key, attr_idx, release_idx)
            return

        # 宝可梦：按属性+发售顺序插入
        sort_key = attr_idx * 10000 + release_idx

        # 需要同 category、同属性、已有库存的卡
        existing: list[tuple[int, int]] = []
        for r in all_rows:
            if r["id"] == card_id:
                continue
            r_cat_key, _ = classify_card(r)
            if r_cat_key != card_category_key:
                continue
            if _attribute_sort_index(r["attribute"]) != attr_idx:
                continue
            if self._get_free_quantity(conn, int(r["id"])) <= 0:
                continue
            rk = attr_idx * 10000 + _get_release_index(r["product_code"])
            existing.append((r["id"], rk))

        if not existing:
            self._set_card_group_sort_order(conn, card_id, 1)
        else:
            existing.sort(key=lambda x: x[1])
            insert_pos = bisect.bisect_right([e[1] for e in existing], sort_key)
            # Shift later cards
            for i in range(insert_pos, len(existing)):
                self._set_card_group_sort_order(conn, existing[i][0], self._get_card_group_sort_order(conn, existing[i][0]) + 1)
            self._set_card_group_sort_order(conn, card_id, insert_pos + 1)

        # 同时更新 holdings_group_orders：按属性+发售顺序插入组位置
        group_key = build_holdings_group_key(card_category_key, card["card_name"])
        self._insert_holdings_group_order(conn, card_category_key, group_key, attr_idx, release_idx)

    def _insert_holdings_group_order(
        self,
        conn: sqlite3.Connection,
        section_key: str,
        new_group_key: str,
        new_attr_idx: int,
        new_release_idx: int,
    ):
        """首次添加卡牌时，若该组不存在则追加到 section 末尾。
        不动任何已有组的顺序。"""
        existing = conn.execute(
            "SELECT sort_order FROM holdings_group_orders WHERE section_key = ? AND group_key = ?",
            (section_key, new_group_key),
        ).fetchone()
        if existing:
            return  # 已存在，不动

        # 取当前最大 sort_order
        max_so = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) FROM holdings_group_orders WHERE section_key = ?",
            (section_key,),
        ).fetchone()[0]
        conn.execute(
            "INSERT INTO holdings_group_orders (section_key, group_key, sort_order) VALUES (?, ?, ?)",
            (section_key, new_group_key, int(max_so) + 1),
        )

    def _validate_deck_same_name_limit_for_card(
        self,
        conn: sqlite3.Connection,
        deck_id: int,
        card_id: int,
        target_quantity: int,
    ):
        card_row = conn.execute(
            f"SELECT id, card_name, card_type, detail, special_text FROM {self._catalog_cards_table(conn)} WHERE id = ?",
            (card_id,),
        ).fetchone()
        if card_row is None:
            raise NotFoundError("卡牌不存在")

        category_key, _ = classify_card(card_row)
        if category_key == "basic_energy":
            return

        group_key = build_holdings_group_key(category_key, card_row["card_name"])
        display_name = normalize_text(remove_card_name_variants(card_row["card_name"])) or card_row["card_name"]
        candidate_rows = conn.execute(
            f"SELECT id, card_name, card_type, detail, special_text FROM {self._catalog_cards_table(conn)}"
        ).fetchall()

        total_quantity = 0
        for row in candidate_rows:
            row_category_key, _ = classify_card(row)
            if row_category_key != category_key:
                continue
            if build_holdings_group_key(row_category_key, row["card_name"]) != group_key:
                continue
            if int(row["id"]) == int(card_id):
                total_quantity += max(0, int(target_quantity))
            else:
                total_quantity += self._get_deck_quantity(conn, deck_id, int(row["id"]))

        if total_quantity > 4:
            deck_row = conn.execute("SELECT name FROM decks WHERE id = ?", (deck_id,)).fetchone()
            deck_name = deck_row["name"] if deck_row else str(deck_id)
            raise ServiceError(f"卡组“{deck_name}”中的同名卡牌合计不能超过 4 张（基本能量除外）：{display_name}")

    def _set_deck_quantity(
        self,
        conn: sqlite3.Connection,
        deck_id: int,
        card_id: int,
        quantity: int,
        backup_quantity: int | None = None,
    ):
        quantity = int(quantity)
        if quantity <= 0:
            conn.execute("DELETE FROM deck_cards WHERE deck_id = ? AND card_id = ?", (deck_id, card_id))
            return
        if backup_quantity is None:
            backup_quantity = self._get_deck_backup_quantity(conn, deck_id, card_id)
        backup_quantity = max(0, min(int(backup_quantity), quantity))
        conn.execute(
            """
            INSERT INTO deck_cards(deck_id, card_id, quantity, backup_quantity) VALUES (?, ?, ?, ?)
            ON CONFLICT(deck_id, card_id) DO UPDATE SET quantity = excluded.quantity,
                                                    backup_quantity = excluded.backup_quantity
            """,
            (deck_id, card_id, quantity, backup_quantity),
        )

    def _set_deck_backup_quantity(self, conn: sqlite3.Connection, deck_id: int, card_id: int, quantity: int):
        current_quantity = self._get_deck_quantity(conn, deck_id, card_id)
        if current_quantity <= 0:
            raise NotFoundError("卡组中不存在这张卡")
        clean_quantity = max(0, min(int(quantity), current_quantity))
        conn.execute(
            "UPDATE deck_cards SET backup_quantity = ? WHERE deck_id = ? AND card_id = ?",
            (clean_quantity, deck_id, card_id),
        )

    def _set_deck_basic_energy_quantity(self, conn: sqlite3.Connection, deck_id: int, energy_code: str, quantity: int):
        clean_code = normalize_text(energy_code).upper()
        quantity = int(quantity)
        if quantity <= 0:
            conn.execute(
                "DELETE FROM deck_basic_energies WHERE deck_id = ? AND energy_code = ?",
                (deck_id, clean_code),
            )
            return
        conn.execute(
            """
            INSERT INTO deck_basic_energies(deck_id, energy_code, quantity) VALUES (?, ?, ?)
            ON CONFLICT(deck_id, energy_code) DO UPDATE SET quantity = excluded.quantity
            """,
            (deck_id, clean_code, quantity),
        )


def classify_card(row: sqlite3.Row | dict[str, Any]) -> tuple[str, str]:
    card_type = normalize_text(row["card_type"] if isinstance(row, sqlite3.Row) else row.get("card_type"))
    detail = normalize_text(row["detail"] if isinstance(row, sqlite3.Row) else row.get("detail"))
    card_name = normalize_text(row["card_name"] if isinstance(row, sqlite3.Row) else row.get("card_name"))
    special_text = normalize_text(row["special_text"] if isinstance(row, sqlite3.Row) else row.get("special_text"))
    text = normalize_supporter_wording_text(" ".join(part for part in [card_type, detail, card_name, special_text] if part))

    # 硬编码：以下卡牌始终归类为"道具"（宝可梦道具/Pokémon Tool），与表格中 detail 字段无关
    card_base_name = remove_card_name_variants(card_name)
    if card_base_name in ("勇气护符", "学习装置", "不服输头带"):
        return "tool", "道具"

    if "宝可梦GX" in text:
        return "pokemon_gx", "宝可梦GX"
    if "宝可梦V" in text:
        return "pokemon_v", "宝可梦V"
    if "宝可梦ex" in text:
        return "pokemon_ex", "宝可梦ex"
    if "光辉宝可梦" in text:
        return "radiant_pokemon", "光辉宝可梦"
    if "特殊能量" in text:
        return "special_energy", "特殊能量"
    if "普通能量" in text:
        return "basic_energy", "普通能量"
    if "物品" in text:
        return "item", "物品"
    if SUPPORTER_LABEL in text:
        return "supporter", SUPPORTER_LABEL
    if "竞技场" in text:
        return "stadium", "竞技场"
    if "道具" in text:
        return "tool", "道具"
    if "宝可梦" in text:
        return "ordinary_pokemon", "普通的宝可梦"
    if "能量" in text:
        return "basic_energy", "普通能量"
    return "ordinary_pokemon", "普通的宝可梦"


def is_pokemon_category_key(category_key: str) -> bool:
    return category_key in POKEMON_CATEGORY_KEYS


def extract_holdings_attributes(attribute_value: Any) -> list[str]:
    text = normalize_text(attribute_value)
    if not text:
        return []

    matched_attributes = [
        canonical
        for canonical in ATTRIBUTE_ORDER
        if any(alias in text for alias in ATTRIBUTE_ALIASES.get(canonical, (canonical,)))
    ]
    if matched_attributes:
        return matched_attributes
    return [text]


def collect_group_holdings_attributes(items: list[dict[str, Any]]) -> list[str]:
    attributes: list[str] = []
    seen: set[str] = set()
    for item in items:
        for attribute in extract_holdings_attributes(item.get("attribute", "")):
            if attribute in seen:
                continue
            seen.add(attribute)
            attributes.append(attribute)
    return attributes


def holdings_attribute_sort_index(attribute_value: Any) -> int:
    attributes = extract_holdings_attributes(attribute_value)
    if not attributes:
        return len(ATTRIBUTE_ORDER)
    return ATTRIBUTE_ORDER_INDEX.get(attributes[0], len(ATTRIBUTE_ORDER))


def format_holdings_group_name(category_key: str, base_name: str, items: list[dict[str, Any]]) -> str:
    if not is_pokemon_category_key(category_key):
        return base_name

    attributes = collect_group_holdings_attributes(items)
    if not attributes:
        return base_name
    return f"{base_name} {'/'.join(attributes)}"


def build_holdings_group_sort_key(category_key: str, group: dict[str, Any]) -> tuple[Any, ...]:
    first_item = group["items"][0]
    group_sort_order = int(group.get("groupSortOrder") or 0)
    order_key = group_sort_order if group_sort_order > 0 else 10_000
    base_key = (
        group["groupBaseName"],
        first_item["productCode"],
        first_item["cardCode"],
        first_item["rarity"],
    )
    if not is_pokemon_category_key(category_key):
        return (order_key, *base_key)

    attributes = collect_group_holdings_attributes(group["items"])
    attribute_sort_index = ATTRIBUTE_ORDER_INDEX.get(attributes[0], len(ATTRIBUTE_ORDER)) if attributes else len(ATTRIBUTE_ORDER)
    return (order_key, attribute_sort_index, group["groupBaseName"], *base_key[1:])


def build_holdings_item_sort_key(category_key: str, item: dict[str, Any]) -> tuple[Any, ...]:
    group_sort_order = int(item.get("groupSortOrder") or 0)
    order_key = group_sort_order if group_sort_order > 0 else 10_000
    base_key = (item["productCode"], item["cardCode"], item["rarity"], item["cardName"], item["id"])
    if not is_pokemon_category_key(category_key):
        return (order_key, *base_key)
    return (order_key, holdings_attribute_sort_index(item.get("attribute", "")), *base_key)


def build_holdings_group_key(category_key: str, card_name: str) -> str:
    return f"{category_key}::{normalize_card_group_key(card_name)}"


def normalize_card_group_key(card_name: str) -> str:
    return remove_card_name_variants(card_name)


def format_holdings_product_code(product_code: str) -> str:
    text = normalize_text(product_code)
    if re.fullmatch(r'151C[1-4]', text, flags=re.IGNORECASE):
        return '151C'
    # 规范化 alias（CSEC1→CSEC, CSVE1pC2→CSVE1pC 等）
    from .mikmoe_source import normalize_product_code
    return normalize_product_code(text)


def format_holdings_card_code(card_code: str) -> str:
    text = normalize_text(card_code)
    return text.split('/', 1)[0].strip()


def resolve_display_product_code(product_code: str, card_code: str, product_name: str = "", card_name: str = "") -> str:
    raw_product_code = normalize_text(product_code)
    if not is_promo_product_code(raw_product_code):
        return format_holdings_product_code(raw_product_code)

    numbered_match = PROMO_NUMBERED_CARD_CODE_PATTERN.fullmatch(normalize_text(card_code))
    if numbered_match:
        return f"{numbered_match.group(2).upper()}-{numbered_match.group(1)}"

    series_match = PROMO_SERIES_CARD_CODE_PATTERN.fullmatch(normalize_text(card_code))
    if series_match and "特典" in normalize_text(product_name):
        series_code = series_match.group(1).upper()
        candy_alias = build_promo_candy_alias(product_name, card_name)
        if candy_alias:
            return candy_alias
        if "超级球" in normalize_text(card_name):
            return f"{series_code}{normalize_text(card_name)}"

    return format_holdings_product_code(raw_product_code)


def format_card_display_code(product_code: str, card_code: str, product_name: str = "", card_name: str = "") -> str:
    display_product_code = resolve_display_product_code(product_code, card_code, product_name, card_name)
    if is_promo_product_code(product_code):
        return display_product_code
    display_card_code = format_holdings_card_code(card_code)
    return f"{display_product_code}-{display_card_code}" if display_card_code else display_product_code


def is_promo_product_code(product_code: str) -> bool:
    return PROMO_PRODUCT_KEYWORD in normalize_text(product_code).upper()


def build_promo_candy_alias(product_name: str, card_name: str) -> str:
    clean_card_name = normalize_text(card_name)
    if "神奇糖果" not in clean_card_name:
        return ""

    year = extract_promo_year(product_name)
    city = extract_promo_city(product_name)
    if not year or not city:
        return ""

    rank = extract_promo_rank(card_name)
    suffix = f"{rank}糖" if rank else "糖"
    return f"{year}{city}{suffix}"


def extract_promo_year(product_name: str) -> str:
    match = PROMO_YEAR_PATTERN.search(normalize_text(product_name))
    return match.group(1) if match else ""


def extract_promo_city(product_name: str) -> str:
    clean_product_name = normalize_text(product_name)
    for city in PROMO_CITIES:
        if city in clean_product_name:
            return city
    return ""


def extract_promo_rank(card_name: str) -> str:
    match = PROMO_RANK_PATTERN.search(normalize_text(card_name))
    return match.group(1) if match else ""


def build_card_exact_search_keys(item: dict[str, Any]) -> set[str]:
    keys = {
        normalize_text(item.get("displayCode", "")).upper(),
        f"{normalize_text(item.get('productCode', '')).upper()}-{normalize_text(item.get('displayCardCode', '')).upper()}",
    }
    return {key for key in keys if key and key != "-"}


def build_card_search_texts(item: dict[str, Any]) -> set[str]:
    texts = {
        normalize_text(item.get("cardName", "")),
        normalize_text(item.get("displayProductCode", "")),
        normalize_text(item.get("displayCode", "")),
        normalize_text(item.get("productCode", "")),
        normalize_text(item.get("nickname", "")),
    }

    for extra_text in build_additional_product_search_texts(item):
        texts.add(extra_text)
    return {text for text in texts if text}


def build_additional_product_search_texts(item: dict[str, Any]) -> set[str]:
    raw_product_code = normalize_text(item.get("productCode", ""))
    raw_card_code = normalize_text(item.get("cardCode", ""))
    raw_product_name = normalize_text(item.get("productName", ""))
    raw_card_name = normalize_text(item.get("cardName", ""))
    texts: set[str] = set()

    if not is_promo_product_code(raw_product_code):
        return texts

    texts.add(resolve_display_product_code(raw_product_code, raw_card_code, raw_product_name, raw_card_name))
    numbered_match = PROMO_NUMBERED_CARD_CODE_PATTERN.fullmatch(raw_card_code)
    if numbered_match:
        texts.add(f"{numbered_match.group(2).upper()}-{numbered_match.group(1)}")
        return texts

    candy_alias = build_promo_candy_alias(raw_product_name, raw_card_name)
    if candy_alias:
        texts.add(candy_alias)
        year = extract_promo_year(raw_product_name)
        city = extract_promo_city(raw_product_name)
        rank = extract_promo_rank(raw_card_name)
        if city:
            texts.add(f"{city}糖")
        if year and city:
            texts.add(f"{year}{city}糖")
        if city and rank:
            texts.add(f"{city}{rank}糖")
        return texts

    series_match = PROMO_SERIES_CARD_CODE_PATTERN.fullmatch(raw_card_code)
    if series_match and "特典" in raw_product_name and "超级球" in raw_card_name:
        texts.add(f"{series_match.group(1).upper()}{raw_card_name}")

    return texts


def build_search_item_same_name_key(item: dict[str, Any]) -> str:
    category_key, _ = classify_card(
        {
            "card_type": item.get("cardType", ""),
            "detail": item.get("detail", ""),
            "card_name": item.get("cardName", ""),
            "special_text": item.get("special", ""),
        }
    )
    return build_holdings_group_key(category_key, item.get("cardName", ""))


def sanitize_search_preferences_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return dict(DEFAULT_SEARCH_PREFERENCES)

    selected_regulations: list[str] = []
    seen_regulations: set[str] = set()
    raw_regulations = payload.get("selectedRegulations", [])
    if isinstance(raw_regulations, list):
        for regulation in raw_regulations:
            clean_regulation = normalize_text(regulation)
            if not clean_regulation or clean_regulation in seen_regulations:
                continue
            seen_regulations.add(clean_regulation)
            selected_regulations.append(clean_regulation)

    return {
        "selectedRegulations": selected_regulations,
        "considerSameNameRegulation": bool(payload.get("considerSameNameRegulation", False)),
    }


def query_matches_card(item: dict[str, Any], query_text: str) -> bool:
    return any(query_text in text.casefold() for text in build_card_search_texts(item))


def build_deck_checks(cards: list[dict[str, Any]], main_card_count: int, backup_card_count: int) -> dict[str, Any]:
    check_items: list[dict[str, Any]] = []
    regulation_counts: dict[str, int] = {}
    same_name_groups: dict[str, dict[str, Any]] = {}
    seen_catalog_ids: set[int] = set()

    for card in cards:
        if card.get("deckEntryType") != "catalog_card":
            continue
        try:
            card_id = int(card.get("id", 0) or 0)
        except (TypeError, ValueError):
            continue
        if card_id in seen_catalog_ids:
            continue
        seen_catalog_ids.add(card_id)

        quantity = max(0, int(card.get("totalDeckQuantity", card.get("deckQuantity", 0)) or 0))
        regulation = normalize_text(card.get("regulation", "")) or "未填写"
        regulation_counts[regulation] = regulation_counts.get(regulation, 0) + quantity

        category_key = normalize_text(card.get("sameNameCategoryKey") or card.get("categoryKey", ""))
        if category_key == "basic_energy":
            continue
        group_key = normalize_text(card.get("sameNameGroupKey")) or build_holdings_group_key(category_key, card.get("cardName", ""))
        group_name = normalize_text(card.get("sameNameGroupName") or card.get("sameNameGroupBaseName")) or normalize_text(remove_card_name_variants(card.get("cardName", ""))) or card.get("cardName", "")
        group = same_name_groups.setdefault(group_key, {"name": group_name, "quantity": 0})
        group["quantity"] += quantity

    deck_size_status = "pass" if main_card_count == 60 else "warning"
    check_items.append(
        {
            "key": "deck_size",
            "label": "主牌数量",
            "status": deck_size_status,
            "message": f"主牌当前 {main_card_count} 张，标准卡组应为 60 张。" if deck_size_status != "pass" else "主牌数量正好 60 张。",
        }
    )

    over_limit_groups = [group for group in same_name_groups.values() if int(group["quantity"]) > 4]
    check_items.append(
        {
            "key": "same_name_limit",
            "label": "同名限制",
            "status": "pass" if not over_limit_groups else "error",
            "message": "没有发现同名超过 4 张的非基本能量。" if not over_limit_groups else "；".join(
                f"{group['name']} {group['quantity']} 张" for group in over_limit_groups
            ),
        }
    )

    if backup_card_count > 0:
        check_items.append(
            {
                "key": "backup",
                "label": "备卡",
                "status": "info",
                "message": f"当前有 {backup_card_count} 张备卡，已从主牌数量检查中排除。",
            }
        )

    return {
        "ok": all(item["status"] in {"pass", "info"} for item in check_items),
        "items": check_items,
        "regulationBreakdown": [
            {"regulation": name, "quantity": quantity}
            for name, quantity in sorted(regulation_counts.items(), key=lambda item: item[0])
        ],
    }


def build_deck_basic_energy_settings(rows: list[sqlite3.Row | dict[str, Any]]) -> list[dict[str, Any]]:
    quantity_map = {
        normalize_text(row["energy_code"] if isinstance(row, sqlite3.Row) else row.get("energy_code", "")).upper(): int(
            row["quantity"] if isinstance(row, sqlite3.Row) else row.get("quantity", 0) or 0
        )
        for row in rows
    }
    return [
        {
            "code": definition["code"],
            "name": definition["name"],
            "attribute": definition["attribute"],
            "quantity": quantity_map.get(definition["code"], 0),
        }
        for definition in BASIC_DECK_ENERGY_DEFINITIONS
    ]


def build_deck_basic_energy_cards(deck_id: int, settings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    for setting in settings:
        quantity = int(setting.get("quantity", 0) or 0)
        if quantity <= 0:
            continue
        attribute = normalize_text(setting.get("attribute", ""))
        cards.append(
            {
                "id": f"basic-energy:{deck_id}:{setting['code']}",
                "sourceKey": f"basic-energy:{deck_id}:{setting['code']}",
                "productName": "",
                "productCode": setting["code"],
                "displayProductCode": setting["code"],
                "cardCode": "",
                "displayCardCode": "",
                "displayCode": setting["code"],
                "cardName": setting["name"],
                "cardType": "普通能量",
                "detail": "普通能量",
                "special": "",
                "attribute": attribute,
                "attributeColor": normalizeColorForAttribute(attribute),
                "rarity": "",
                "regulation": "",
                "note": "",
                "freeQuantity": 0,
                "deckQuantity": quantity,
                "ownedQuantity": quantity,
                "totalDeckQuantity": quantity,
                "currentBackupQuantity": 0,
                "backupCap": 0,
                "showBackupControl": False,
                "categoryKey": "basic_energy",
                "categoryTitle": "普通能量",
                "deckEntryType": "basic_energy",
                "deckEntryKind": "main",
            }
        )
    return cards


def build_deck_detail_sections(
    cards: list[dict[str, Any]],
    section_order_map: dict[tuple[str, str], int] | None = None,
) -> list[dict[str, Any]]:
    sections = [
        {"key": key, "title": title, "column": column, "items": [], "groups": []}
        for key, title, column in DECK_DETAIL_SECTION_DEFINITIONS
    ]
    section_lookup = {section["key"]: section for section in sections}
    section_order_map = section_order_map or {}

    for card in cards:
        category_key = normalize_text(card.get("categoryKey", ""))
        section_key = DECK_DETAIL_SECTION_CATEGORY_MAP.get(category_key, "pokemon")
        section = section_lookup[section_key]
        group_key = normalize_text(card.get("sameNameGroupKey")) or f"{section_key}::{card.get('deckEntryType', 'card')}::{card.get('id', '')}"
        entry_key = build_deck_detail_entry_key(card)
        card["sameNameGroupKey"] = group_key
        card["deckSectionKey"] = section_key
        card["deckSectionEntryKey"] = entry_key
        card["deckSectionSortOrder"] = int(section_order_map.get((section_key, entry_key), 0) or 0)
        section["items"].append(card)

    for section in sections:
        if not section["items"]:
            continue

        grouped_items: dict[str, list[dict[str, Any]]] = {}
        for item in section["items"]:
            grouped_items.setdefault(item["sameNameGroupKey"], []).append(item)

        for group_items in grouped_items.values():
            group_category_key = normalize_text(group_items[0].get("sameNameCategoryKey") or group_items[0].get("categoryKey") or section["key"])
            group_base_name = normalize_text(group_items[0].get("sameNameGroupBaseName")) or normalize_text(remove_card_name_variants(group_items[0].get("cardName", ""))) or group_items[0].get("cardName", "")
            group_name = format_holdings_group_name(group_category_key, group_base_name, group_items)
            ordered_group_items = sorted(group_items, key=lambda item: build_deck_detail_group_item_sort_key(group_category_key, item))
            for index, item in enumerate(ordered_group_items, start=1):
                item["sameNameGroupName"] = group_name
                item["sameNameGroupSize"] = len(group_items)
                item["sameNameGroupIndex"] = index

        section["items"].sort(key=lambda item: build_deck_detail_item_sort_key(section["key"], item))

    return [section for section in sections if section["items"]]


def build_deck_detail_entry_key(card: dict[str, Any]) -> str:
    deck_entry_type = normalize_text(card.get("deckEntryType", ""))
    if deck_entry_type == "basic_energy":
        code = normalize_text(card.get("displayCode", "") or card.get("productCode", "")).upper()
        return f"basic_energy::{code}"

    entry_kind = normalize_text(card.get("deckEntryKind", "")) or ("backup" if card.get("isBackup") else "main")
    return f"{deck_entry_type or 'card'}::{normalize_text(card.get('id', ''))}::{entry_kind}"


def build_deck_detail_item_sort_key(section_key: str, item: dict[str, Any]) -> tuple[Any, ...]:
    section_sort_order = int(item.get("deckSectionSortOrder") or 0)
    order_key = section_sort_order if section_sort_order > 0 else 10_000
    group_category_key = normalize_text(item.get("sameNameCategoryKey") or item.get("categoryKey") or section_key)
    group_base_name = normalize_text(item.get("sameNameGroupBaseName")) or normalize_text(remove_card_name_variants(item.get("cardName", ""))) or item.get("cardName", "")
    if group_category_key == "basic_energy" and item.get("deckEntryType") == "basic_energy":
        energy_order = BASIC_DECK_ENERGY_ORDER_INDEX.get(normalize_text(item.get("displayCode", "")).upper(), 10_000)
        group_sort_key = (len(ATTRIBUTE_ORDER), 1, energy_order, group_base_name)
    elif not is_pokemon_category_key(group_category_key):
        group_sort_key = (len(ATTRIBUTE_ORDER), 0, group_base_name)
    else:
        attributes = extract_holdings_attributes(item.get("attribute", ""))
        attribute_sort_index = ATTRIBUTE_ORDER_INDEX.get(attributes[0], len(ATTRIBUTE_ORDER)) if attributes else len(ATTRIBUTE_ORDER)
        group_sort_key = (attribute_sort_index, 0, group_base_name)
    entry_kind_order = 1 if normalize_text(item.get("deckEntryKind", "")) == "backup" or item.get("isBackup") else 0
    return (
        order_key,
        *group_sort_key,
        *build_deck_detail_group_item_sort_key(group_category_key, item),
        entry_kind_order,
    )


def build_deck_detail_group_item_sort_key(group_category_key: str, item: dict[str, Any]) -> tuple[Any, ...]:
    group_sort_order = int(item.get("groupSortOrder") or 0)
    order_key = group_sort_order if group_sort_order > 0 else 10_000
    base_key = (item["productCode"], item["cardCode"], item["rarity"], item["cardName"], item["id"])
    if not is_pokemon_category_key(group_category_key):
        return (order_key, *base_key)
    return (order_key, holdings_attribute_sort_index(item.get("attribute", "")), *base_key)


def build_deck_detail_group_sort_key(section_key: str, group: dict[str, Any]) -> tuple[Any, ...]:
    first_item = group["items"][0]
    group_category_key = normalize_text(first_item.get("sameNameCategoryKey") or first_item.get("categoryKey") or section_key)
    base_key = (group["groupBaseName"], first_item["productCode"], first_item["cardCode"], first_item["rarity"])
    if group_category_key == "basic_energy" and first_item.get("deckEntryType") == "basic_energy":
        energy_order = BASIC_DECK_ENERGY_ORDER_INDEX.get(normalize_text(first_item.get("displayCode", "")).upper(), 10_000)
        return (len(ATTRIBUTE_ORDER), 1, energy_order, *base_key)
    if not is_pokemon_category_key(group_category_key):
        return (len(ATTRIBUTE_ORDER), 0, *base_key)

    attributes = collect_group_holdings_attributes(group["items"])
    attribute_sort_index = ATTRIBUTE_ORDER_INDEX.get(attributes[0], len(ATTRIBUTE_ORDER)) if attributes else len(ATTRIBUTE_ORDER)
    return (attribute_sort_index, 0, *base_key)


def normalizeColorForAttribute(attribute: str) -> str:
    attributes = extract_holdings_attributes(attribute)
    if not attributes:
        return ""
    return normalize_hex_color(ATTRIBUTE_COLOR_FALLBACKS.get(attributes[0], ""))


def remove_card_name_variants(card_name: str) -> str:
    text = normalize_text(card_name)
    for keyword in SPECIAL_SAME_NAME_KEYWORDS:
        if keyword in text:
            return keyword
    for suffix in CARD_NAME_VARIANT_SUFFIXES:
        text = text.replace(suffix, "")
    text = "".join(ch for ch in text if not is_unicode_punctuation(ch))
    return re.sub(r"\s+", "", text)


def is_unicode_punctuation(ch: str) -> bool:
    return bool(ch) and unicodedata.category(ch).startswith(("P", "S"))


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_supporter_wording_text(value: Any) -> str:
    return normalize_text(value).replace(LEGACY_SUPPORTER_LABEL, SUPPORTER_LABEL)


def normalize_hex_color(value: Any) -> str:
    text = normalize_text(value).lstrip("#")
    if len(text) == 8:
        text = text[2:]
    if re.fullmatch(r"[0-9a-fA-F]{6}", text):
        return f"#{text.lower()}"
    return ""


def extract_excel_fill_color(cell: Any) -> str:
    fill = getattr(cell, "fill", None)
    if fill is None or not getattr(fill, "fill_type", None):
        return ""

    start_color = getattr(fill, "start_color", None)
    if start_color is None:
        return ""

    color_type = getattr(start_color, "type", None)
    if color_type == "rgb":
        return normalize_hex_color(getattr(start_color, "rgb", ""))

    if color_type == "indexed":
        try:
            indexed = int(getattr(start_color, "indexed", -1))
        except (TypeError, ValueError):
            return ""
        if 0 <= indexed < len(COLOR_INDEX):
            return normalize_hex_color(COLOR_INDEX[indexed])

    return ""



def parse_positive_int(value: Any, field_name: str = "数量") -> int:
    try:
        amount = int(value)
    except (TypeError, ValueError) as exc:
        raise ServiceError(f"{field_name}必须是正整数") from exc
    if amount <= 0:
        raise ServiceError(f"{field_name}必须大于 0")
    return amount


def parse_non_negative_int(value: Any, field_name: str = "数量") -> int:
    try:
        amount = int(value)
    except (TypeError, ValueError) as exc:
        raise ServiceError(f"{field_name}必须是大于等于 0 的整数") from exc
    if amount < 0:
        raise ServiceError(f"{field_name}不能小于 0")
    return amount



def build_header_index_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for index, header in enumerate(header_row):
        name = normalize_text(header)
        if name in EXPECTED_HEADERS:
            index_map[name] = index
    return index_map



def row_to_record(values: tuple[Any, ...], index_map: dict[str, int]) -> dict[str, Any]:
    def get(header_name: str) -> str:
        idx = index_map.get(header_name)
        if idx is None or idx >= len(values):
            return ""
        return normalize_text(values[idx])

    quantity_text = get("数量")
    try:
        quantity = int(float(quantity_text)) if quantity_text else 0
    except ValueError:
        quantity = 0

    record = {
        "product_name": get("商品名称"),
        "product_code": get("商品编号"),
        "card_code": get("卡牌编号"),
        "card_name": get("卡牌名称"),
        "card_type": normalize_supporter_wording_text(get("类型")),
        "detail": normalize_supporter_wording_text(get("详细")),
        "special_text": normalize_supporter_wording_text(get("特殊")),
        "attribute": get("属性"),
        "rarity": get("稀有度"),
        "regulation": get("赛制"),
        "quantity": max(quantity, 0),
        "note": get("备注"),
        "nickname": "",
    }
    return record



def build_source_key(record: dict[str, Any]) -> str:
    return "||".join(build_catalog_identity_parts(record))


def build_catalog_identity_parts(record: dict[str, Any]) -> tuple[str, ...]:
    parts: list[str] = []
    for field_name in CATALOG_IDENTITY_FIELDS:
        value = normalize_text(record.get(field_name, ""))
        if field_name in {"product_code", "card_code"}:
            value = value.upper()
        parts.append(value)
    return tuple(parts)



def dumps_state(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════
# 首次添加宝可梦时自动排序辅助函数
# ═══════════════════════════════════════════════════════════════

_release_index_cache: dict[str, int] | None = None
_release_index_lock = __import__('threading').Lock()


def _load_release_index():
    global _release_index_cache
    with _release_index_lock:
        if _release_index_cache is not None:
            return
        _release_index_cache = {}
        try:
            from pathlib import Path
            excel_path = Path(__file__).resolve().parent.parent / "data" / "卡表.xlsx"
            if not excel_path.exists():
                return
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.worksheets[0]
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                pc = str(row[1] or '').strip() if len(row) > 1 else ''
                if pc and pc not in _release_index_cache:
                    _release_index_cache[pc] = i
            wb.close()
        except Exception:
            pass


def _get_release_index(product_code: str) -> int:
    _load_release_index()
    return (_release_index_cache or {}).get(product_code.strip(), 9999)


def _attribute_sort_index(attribute_value: str) -> int:
    text = (attribute_value or '').strip()
    if not text:
        return len(ATTRIBUTE_ORDER)
    for canonical in ATTRIBUTE_ORDER:
        if any(alias in text for alias in ATTRIBUTE_ALIASES.get(canonical, (canonical,))):
            return ATTRIBUTE_ORDER_INDEX.get(canonical, len(ATTRIBUTE_ORDER))
    return len(ATTRIBUTE_ORDER)

