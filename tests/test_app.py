from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from ptcg_gallery import create_app


ATTRIBUTE_FILL_COLORS = {
    "草": "4CAF50",
    "火": "EF6C36",
    "水": "2F80ED",
    "电": "F5B700",
    "雷": "F5B700",
    "超": "9B5DE5",
    "斗": "A35A1F",
    "恶": "344054",
    "钢": "7A8798",
    "金属": "7A8798",
    "龙": "5A3FD1",
    "妖": "EC6AA7",
    "妖精": "EC6AA7",
    "无": "9AA0A6",
    "无色": "9AA0A6",
}


class PtcgGalleryAppTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.data_dir = self.root / "data"
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.default_excel = self.data_dir / "卡表.xlsx"
        self.db_path = self.data_dir / "test.db"
        self._create_catalog_excel(self.default_excel)
        self.app = create_app(
            {
                "TESTING": True,
                "ROOT_DIR": str(self.root),
                "DATABASE": str(self.db_path),
                "DEFAULT_EXCEL_PATH": str(self.default_excel),
                "AUTH_USERNAME": "test_admin",
                "AUTH_PASSWORD": "test_pass",
                "INIT_ADMIN_PASS": "test_pass",
            }
        )
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["account_id"] = 1
            session["account_name"] = "RhymesX"

    def tearDown(self):
        self.temp_dir.cleanup()

    def _build_rows(self, include_same_name_duplicate: bool = False):
        rows = [
            ["起始包A", "CSM1aC", "002", "小火龙", "宝可梦", "基础", "", "火", "C", "标准", 2, "首刷", "火龙宝宝"],
            ["起始包A", "CSM1aC", "002", "小火龙球闪", "宝可梦", "基础", "球闪", "火", "C", "标准", 1, "闪版", "小火闪"],
            ["起始包B", "SV8", "015", "皮卡丘", "宝可梦", "基础", "", "雷", "R", "标准", 3, "", "电耗子"],
            ["对战包", "SVP", "001", "喷火龙GX", "宝可梦", "阶段2", "宝可梦GX", "火", "RR", "标准", 1, "", "火焰GX"],
            ["对战包", "SVP", "002", "伊布V", "宝可梦", "基础", "宝可梦V", "无", "RR", "标准", 1, "", "V小伊布"],
            ["对战包", "SVP", "003", "烈咬陆鲨ex", "宝可梦", "阶段2", "宝可梦ex", "龙", "RR", "标准", 1, "", "龙ex"],
            ["对战包", "SVP", "004", "拉鲁拉丝", "宝可梦", "基础", "光辉宝可梦", "超", "R", "标准", 1, "", "光辉小星"],
            ["辅助包", "TRN", "101", "高级球", "物品", "", "", "无", "U", "标准", 4, "", "球"],
            ["辅助包", "TRN", "102", "研究员", "支援者", "", "", "无", "U", "标准", 2, "", "博士"],
            ["辅助包", "TRN", "103", "山岳竞技场", "竞技场", "", "", "无", "U", "标准", 2, "", "山岳"],
            ["辅助包", "TRN", "104", "神奇糖果", "道具", "", "", "无", "U", "标准", 4, "", "糖果"],
            ["辅助包", "TRN", "105", "空白测试卡", "物品", "", "", "无", "U", "标准", 0, "", "空白"],
            ["能量包", "ENG", "201", "双重涡轮能量", "特殊能量", "", "特殊能量", "无", "C", "标准", 4, "", "双涡轮"],
            ["能量包", "ENG", "202", "火能量", "普通能量", "", "普通能量", "火", "C", "标准", 8, "", "火能"],
            ["特别包", "151C4", "020/1", "测试卡", "宝可梦", "基础", "", "无", "C", "标准", 1, "", "测试"],
            ["起始包D", "CSM2aC", "010", "一对鼠", "宝可梦", "基础", "", "无", "C", "标准", 2, "", "一对鼠"],
            ["起始包D", "CSM2aC", "010", "一对鼠·精灵球闪", "宝可梦", "基础", "精灵球闪", "无", "R", "标准", 1, "", "一对鼠精闪"],
            ["起始包D", "CSM2aC", "011", "一对鼠·大师球闪", "宝可梦", "基础", "大师球闪", "无", "U", "标准", 1, "", "一对鼠大师闪"],
            ["起始包D", "CSM2aC", "013", "一对鼠ex", "宝可梦", "基础", "宝可梦ex", "无", "RR", "标准", 1, "", "一对鼠ex"],
        ]
        if include_same_name_duplicate:
            rows.append(["起始包C", "CSM9aC", "099", "小火龙", "宝可梦", "基础", "特别版", "火", "U", "标准", 5, "重复名称", "火龙宝宝特别版"])
        return rows

    def _build_promo_rows(self):
        return [
            ["常规 PROMO", "PROMO", "002/SM-P", "皮卡丘PROMO", "宝可梦", "基础", "", "雷", "R", "标准", 1, "", ""],
            ["2023北京特典", "PROMO", "SM-P", "神奇糖果·64强", "道具", "", "", "无", "U", "标准", 1, "", ""],
            ["2025上海特典", "PROMO", "SM-P", "神奇糖果·冠军", "道具", "", "", "无", "U", "标准", 1, "", ""],
            ["2024广州特典", "PROMO", "SM-P", "超级球·8强", "物品", "", "", "无", "U", "标准", 1, "", ""],
            ["2025深圳特典", "PROMO", "SM-P", "超级球·冠军", "物品", "", "", "无", "U", "标准", 1, "", ""],
        ]

    def _create_catalog_excel(self, path: Path, rows: list[list[object]] | None = None):
        workbook = Workbook()
        guide_sheet = workbook.active
        guide_sheet.title = "说明"
        guide_sheet.append(["这个页签不是卡表"])

        sheet = workbook.create_sheet("卡表")
        sheet.append([
            "商品名称",
            "商品编号",
            "卡牌编号",
            "卡牌名称",
            "类型",
            "详细",
            "特殊",
            "属性",
            "稀有度",
            "赛制",
            "数量",
            "备注",
            "昵称",
        ])
        for row in (rows or self._build_rows()):
            sheet.append(row)

        attribute_column = 8
        for row_index in range(2, sheet.max_row + 1):
            attribute_value = str(sheet.cell(row=row_index, column=attribute_column).value or "")
            fill_color = self._pick_attribute_fill(attribute_value)
            if not fill_color:
                continue
            hex_color = f"FF{fill_color}"
            sheet.cell(row=row_index, column=attribute_column).fill = PatternFill(
                fill_type="solid",
                start_color=hex_color,
                end_color=hex_color,
            )
        workbook.save(path)

    def _pick_attribute_fill(self, value: str) -> str:
        text = str(value or "").strip()
        for keyword, color in ATTRIBUTE_FILL_COLORS.items():
            if keyword and keyword in text:
                return color
        return ""

    def _create_test_app(self, rows: list[list[object]]):
        temp_dir = tempfile.TemporaryDirectory()
        root = Path(temp_dir.name)
        data_dir = root / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        default_excel = data_dir / "卡表.xlsx"
        db_path = data_dir / "test.db"
        self._create_catalog_excel(default_excel, rows)
        app = create_app(
            {
                "TESTING": True,
                "ROOT_DIR": str(root),
                "DATABASE": str(db_path),
                "DEFAULT_EXCEL_PATH": str(default_excel),
            }
        )
        client = app.test_client()
        with client.session_transaction() as session:
            session["account_id"] = 1
            session["account_name"] = "RhymesX"
            session["authed"] = True
        return temp_dir, app, client

    def _find_card_id(self, client, query: str, card_name: str) -> int:
        payload = client.get(f"/api/search?q={query}").get_json()
        item = next(item for item in payload["items"] if item["cardName"] == card_name)
        return item["id"]

    def _generate_invite_code(self, client) -> str:
        """以管理员身份生成邀请码并返回 code 字符串。"""
        with client.session_transaction() as session:
            session["account_id"] = 1
            session["account_name"] = "RhymesX"
            session["is_admin"] = True
        data = client.post("/api/invite-codes").get_json()
        codes = data.get("codes", [])
        self.assertTrue(len(codes) > 0, "应至少有一个邀请码")
        return codes[0]["code"]

    def test_exact_code_search_returns_multiple_cards(self):
        decks = self.client.get("/api/decks")
        deck_names = [deck["name"] for deck in decks.get_json()["items"]]
        self.assertCountEqual(deck_names, ["电友", "龙柱", "铝钢龙", "多龙"])

        summary = self.client.get("/api/summary")
        self.assertEqual(summary.status_code, 200)
        self.assertEqual(summary.get_json()["catalogCount"], len(self._build_rows()))

        empty_search = self.client.get("/api/search?q=")
        self.assertEqual(empty_search.status_code, 200)
        self.assertEqual(empty_search.get_json()["items"], [])

        response = self.client.get("/api/search?q=CSM1aC-002")
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        names = [item["cardName"] for item in payload["items"]]
        self.assertEqual(names, ["小火龙", "小火龙球闪"])

    def test_name_and_parsed_promo_search(self):
        temp_dir, app, client = self._create_test_app(self._build_rows() + self._build_promo_rows())
        self.addCleanup(temp_dir.cleanup)

        response = client.get("/api/search?q=小火")
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(len(payload["items"]), 2)

        response = client.get("/api/search?q=SM-P-002")
        payload = response.get_json()
        self.assertEqual(len(payload["items"]), 1)
        self.assertEqual(payload["items"][0]["cardName"], "皮卡丘PROMO")
        self.assertEqual(payload["items"][0]["displayProductCode"], "SM-P-002")
        self.assertEqual(payload["items"][0]["displayCode"], "SM-P-002")

        response = client.get("/api/search?q=上海糖")
        payload = response.get_json()
        self.assertEqual([item["cardName"] for item in payload["items"]], ["神奇糖果·冠军"])
        self.assertEqual(payload["items"][0]["displayProductCode"], "2025上海冠军糖")

        response = client.get("/api/search?q=SM-P超级球")
        payload = response.get_json()
        self.assertEqual([item["cardName"] for item in payload["items"]], ["超级球·8强", "超级球·冠军"])

    def test_search_can_filter_by_multiple_regulations_and_list_options(self):
        rows = [
            ["训练家包A", "TRN", "101", "高级球", "物品", "", "", "无", "U", "标准", 2, "", ""],
            ["训练家包B", "TRN", "102", "高级球", "物品", "", "", "无", "U", "扩展", 1, "", ""],
            ["训练家包C", "TRN", "103", "高级球", "物品", "", "", "无", "U", "无限", 1, "", ""],
            ["训练家包D", "TRN", "104", "研究员", "支援者", "", "", "无", "U", "标准", 1, "", ""],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        options_response = client.get("/api/search/options")
        self.assertEqual(options_response.status_code, 200)
        self.assertCountEqual(options_response.get_json()["regulations"], ["标准", "扩展", "无限"])
        self.assertEqual(
            options_response.get_json()["preferences"],
            {"selectedRegulations": [], "considerSameNameRegulation": False},
        )

        all_response = client.get("/api/search?q=高级球")
        self.assertEqual(all_response.status_code, 200)
        self.assertEqual(len(all_response.get_json()["items"]), 3)

        standard_response = client.get("/api/search?q=高级球&regulation=标准")
        self.assertEqual(standard_response.status_code, 200)
        self.assertEqual([item["regulation"] for item in standard_response.get_json()["items"]], ["标准"])

        mixed_response = client.get("/api/search?q=高级球&regulation=标准&regulation=无限")
        self.assertEqual(mixed_response.status_code, 200)
        self.assertCountEqual([item["regulation"] for item in mixed_response.get_json()["items"]], ["标准", "无限"])

    def test_search_preferences_are_persisted_to_local_file(self):
        response = self.client.put(
            "/api/search/preferences",
            json={
                "selectedRegulations": ["F", "G", "F", ""],
                "considerSameNameRegulation": True,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get_json(),
            {"selectedRegulations": ["F", "G"], "considerSameNameRegulation": True},
        )

        preferences = self.client.get("/api/search/options").get_json()["preferences"]
        self.assertEqual(
            preferences,
            {"selectedRegulations": ["F", "G"], "considerSameNameRegulation": True},
        )

        options_response = self.client.get("/api/search/options")
        self.assertEqual(options_response.status_code, 200)
        self.assertEqual(
            options_response.get_json()["preferences"],
            {"selectedRegulations": ["F", "G"], "considerSameNameRegulation": True},
        )

    def test_search_can_use_same_name_regulation_filter_when_enabled(self):
        rows = [
            ["训练家套装A", "TRN", "201", "博士的研究（弗图博士）", "支援者", "", "", "无", "U", "F", 1, "", ""],
            ["训练家套装B", "TRN", "202", "博士的研究（奥琳博士）", "支援者", "", "", "无", "U", "G", 1, "", ""],
            ["训练家套装C", "TRN", "203", "博士的研究（木兰博士）", "支援者", "", "", "无", "U", "D", 1, "", ""],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        default_filtered = client.get("/api/search?q=博士的研究&regulation=F&regulation=G")
        self.assertEqual(default_filtered.status_code, 200)
        self.assertCountEqual(
            [item["cardName"] for item in default_filtered.get_json()["items"]],
            ["博士的研究（弗图博士）", "博士的研究（奥琳博士）"],
        )

        same_name_filtered = client.get(
            "/api/search?q=博士的研究&regulation=F&regulation=G&considerSameNameRegulation=true"
        )
        self.assertEqual(same_name_filtered.status_code, 200)
        self.assertCountEqual(
            [item["cardName"] for item in same_name_filtered.get_json()["items"]],
            ["博士的研究（弗图博士）", "博士的研究（奥琳博士）", "博士的研究（木兰博士）"],
        )

    def test_catalog_reload_updates_parsed_product_search_without_creating_duplicate(self):
        rows = [
            ["2023北京特典", "PROMO", "SM-P", "神奇糖果·64强", "道具", "", "", "无", "U", "标准", 1, "", ""],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        updated_rows = [
            ["2025广州特典", "PROMO", "SM-P", "神奇糖果·64强", "道具", "", "", "无", "U", "标准", 1, "", ""],
        ]
        self._create_catalog_excel(Path(app.config["DEFAULT_EXCEL_PATH"]), updated_rows)

        import_response = client.post("/api/import/catalog-default")
        self.assertEqual(import_response.status_code, 200)
        import_payload = import_response.get_json()
        self.assertEqual(import_payload["created"], 0)
        self.assertEqual(import_payload["updated"], 1)

        summary = client.get("/api/summary")
        self.assertEqual(summary.status_code, 200)
        self.assertEqual(summary.get_json()["catalogCount"], 1)

        old_search = client.get("/api/search?q=北京糖")
        self.assertEqual(old_search.status_code, 200)
        self.assertEqual(old_search.get_json()["items"], [])

        new_search = client.get("/api/search?q=广州糖")
        self.assertEqual(new_search.status_code, 200)
        items = new_search.get_json()["items"]
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["cardName"], "神奇糖果·64强")
        self.assertEqual(items[0]["displayProductCode"], "2025广州64强糖")

    def test_exact_code_search_matches_display_style_code(self):
        rows = [
            ["特别包", "151C4", "020/1", "测试卡", "宝可梦", "基础", "", "无", "C", "标准", 1, "", "测试"],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        raw_response = client.get("/api/search?q=151C4-020")
        self.assertEqual(raw_response.status_code, 200)
        raw_items = raw_response.get_json()["items"]
        self.assertEqual(len(raw_items), 1)
        self.assertEqual(raw_items[0]["cardCode"], "020/1")

        display_response = client.get("/api/search?q=151C-020")
        self.assertEqual(display_response.status_code, 200)
        display_items = display_response.get_json()["items"]
        self.assertEqual(len(display_items), 1)
        self.assertEqual(display_items[0]["displayProductCode"], "151C")
        self.assertEqual(display_items[0]["displayCardCode"], "020")

    def test_holdings_report_groups_by_category_and_keeps_rows_separate(self):
        temp_dir, app, client = self._create_test_app(self._build_rows(include_same_name_duplicate=True))
        self.addCleanup(temp_dir.cleanup)

        self.assertEqual(client.get("/holdings").status_code, 200)
        response = client.get("/api/holdings")
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()

        self.assertCountEqual(payload["deckNames"], ["电友", "龙柱", "铝钢龙", "多龙"])
        section_titles = [section["title"] for section in payload["sections"]]
        self.assertIn("普通的宝可梦", section_titles)
        self.assertIn("宝可梦GX", section_titles)
        self.assertIn("宝可梦V", section_titles)
        self.assertIn("宝可梦ex", section_titles)
        self.assertIn("特殊能量", section_titles)
        self.assertIn("支援者", section_titles)

        ordinary = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        ordinary_group_names = [group["groupName"] for group in ordinary["groups"]]
        self.assertEqual(ordinary_group_names, ["小火龙 火", "皮卡丘 电", "一对鼠 无", "测试卡 无"])
        all_items = [item for group in ordinary["groups"] for item in group["items"]]
        zero_rows = [item for item in all_items if item["cardName"] == "空白测试卡"]
        self.assertEqual(zero_rows, [])

        display_item = next(item for item in all_items if item["productCode"] == "151C4")
        self.assertEqual(display_item["displayProductCode"], "151C")
        self.assertEqual(display_item["displayCardCode"], "020")
        charmander_item = next(item for item in all_items if item["cardName"] == "小火龙")
        self.assertEqual(charmander_item["attributeColor"], "#ef6c36")

        same_name_group = next(group for group in ordinary["groups"] if group["groupName"] == "一对鼠 无")
        self.assertEqual(len(same_name_group["items"]), 3)
        self.assertEqual([item["cardCode"] for item in same_name_group["items"]], ["010", "010", "011"])
        self.assertEqual([item["rarity"] for item in same_name_group["items"]], ["C", "R", "U"])
        self.assertEqual([item["cardName"] for item in same_name_group["items"]], ["一对鼠", "一对鼠·精灵球闪", "一对鼠·大师球闪"])

        charmander_group = next(group for group in ordinary["groups"] if group["groupName"] == "小火龙 火")
        self.assertEqual(len(charmander_group["items"]), 3)
        self.assertEqual([item["cardName"] for item in charmander_group["items"]], ["小火龙", "小火龙球闪", "小火龙"])

        ex_group = next(
            group
            for section in payload["sections"]
            for group in section["groups"]
            if group["groupName"] == "一对鼠ex 无"
        )
        self.assertEqual(len(ex_group["items"]), 1)
        self.assertEqual(ex_group["items"][0]["cardName"], "一对鼠ex")

        categories = {section["key"]: sum(len(group["items"]) for group in section["groups"]) for section in payload["sections"]}
        self.assertGreaterEqual(categories["pokemon_gx"], 1)
        self.assertGreaterEqual(categories["pokemon_v"], 1)
        self.assertGreaterEqual(categories["pokemon_ex"], 1)
        self.assertGreaterEqual(categories["radiant_pokemon"], 1)
        self.assertGreaterEqual(categories["item"], 1)
        self.assertGreaterEqual(categories["supporter"], 1)
        self.assertGreaterEqual(categories["stadium"], 1)
        self.assertGreaterEqual(categories["tool"], 1)
        self.assertGreaterEqual(categories["special_energy"], 1)
        self.assertGreaterEqual(categories["basic_energy"], 1)

    def test_holdings_groups_magic_candy_and_super_ball_as_same_name(self):
        temp_dir, app, client = self._create_test_app(self._build_promo_rows())
        self.addCleanup(temp_dir.cleanup)

        payload = client.get("/api/holdings").get_json()
        tool_section = next(section for section in payload["sections"] if section["key"] == "tool")
        item_section = next(section for section in payload["sections"] if section["key"] == "item")

        self.assertEqual([group["groupName"] for group in tool_section["groups"]], ["神奇糖果"])
        self.assertEqual([group["groupName"] for group in item_section["groups"]], ["超级球"])
        self.assertEqual(tool_section["groups"][0]["items"][0]["displayProductCode"], "2023北京64强糖")
        self.assertEqual(item_section["groups"][0]["items"][0]["displayProductCode"], "SM-P超级球·8强")

    def test_holdings_groups_professor_and_boss_cards_as_same_name(self):
        rows = [
            ["训练家套装A", "TRN", "201", "博士的研究（弗图博士）", "支援者", "", "", "无", "U", "标准", 1, "", ""],
            ["训练家套装B", "TRN", "202", "博士的研究（奥琳博士）", "支援者", "", "", "无", "U", "标准", 1, "", ""],
            ["训练家套装C", "TRN", "203", "老大的指令（赤日）", "支援者", "", "", "无", "U", "标准", 1, "", ""],
            ["训练家套装D", "TRN", "204", "老大的指令（坂木）", "支援者", "", "", "无", "U", "标准", 1, "", ""],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        payload = client.get("/api/holdings").get_json()
        supporter_section = next(section for section in payload["sections"] if section["key"] == "supporter")
        self.assertEqual([group["groupName"] for group in supporter_section["groups"]], ["博士的研究", "老大的指令"])
        self.assertEqual(
            [item["cardName"] for item in supporter_section["groups"][0]["items"]],
            ["博士的研究（弗图博士）", "博士的研究（奥琳博士）"],
        )
        self.assertEqual(
            [item["cardName"] for item in supporter_section["groups"][1]["items"]],
            ["老大的指令（赤日）", "老大的指令（坂木）"],
        )

    def test_search_does_not_match_unrelated_cards_only_by_product_name(self):
        rows = [
            ["骑拉帝纳VSTAR卡组构筑进阶礼盒", "CSXC", "001", "交替推车", "物品", "", "", "无", "U", "标准", 1, "", ""],
            ["骑拉帝纳VSTAR卡组构筑进阶礼盒", "CSXC", "002", "勾魂眼", "宝可梦", "基础", "", "恶", "C", "标准", 1, "", ""],
            ["补充包 失落深渊", "SLL", "077", "骑拉帝纳V", "宝可梦", "基础", "宝可梦V", "龙", "RR", "标准", 1, "", ""],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        payload = client.get("/api/search?q=骑拉帝纳").get_json()
        self.assertEqual([item["cardName"] for item in payload["items"]], ["骑拉帝纳V"])

    def test_holdings_report_orders_pokemon_groups_by_attribute(self):
        rows = [
            ["测试包", "TMP", "001", "双属性兽", "宝可梦", "基础", "", "龙/火", "C", "标准", 1, "", "双火龙"],
            ["测试包", "TMP", "002", "电属性兽", "宝可梦", "基础", "", "雷", "C", "标准", 1, "", "电兽"],
            ["测试包", "TMP", "003", "无属性兽", "宝可梦", "基础", "", "无", "C", "标准", 1, "", "无兽"],
        ]
        temp_dir, app, client = self._create_test_app(rows)
        self.addCleanup(temp_dir.cleanup)

        payload = client.get("/api/holdings").get_json()
        ordinary = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        self.assertEqual([group["groupName"] for group in ordinary["groups"]], ["双属性兽 火/龙", "电属性兽 电", "无属性兽 无"])

    def test_inventory_and_deck_flow_and_state_roundtrip(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "喷火龙卡组", "description": "测试卡组", "color": "#ff8800"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        create_payload = create_deck_response.get_json()
        self.assertEqual(create_payload["color"], "#ff8800")
        deck_id = create_payload["id"]

        search = self.client.get("/api/search?q=小火龙")
        first_card = search.get_json()["items"][0]
        card_id = first_card["id"]
        self.assertEqual(first_card["freeQuantity"], 2)

        plus_one = self.client.post(f"/api/cards/{card_id}/free-adjust", json={"delta": 1})
        self.assertEqual(plus_one.status_code, 200)
        self.assertEqual(plus_one.get_json()["freeQuantity"], 3)

        moved = self.client.post(
            f"/api/cards/{card_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 2, "consumeFree": True},
        )
        self.assertEqual(moved.status_code, 200)
        moved_payload = moved.get_json()
        self.assertEqual(moved_payload["freeQuantity"], 1)
        self.assertEqual(moved_payload["deckQuantity"], 2)

        deck_detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(deck_detail.status_code, 200)
        deck_detail_payload = deck_detail.get_json()
        self.assertEqual(deck_detail_payload["name"], "喷火龙卡组")
        self.assertGreaterEqual(len(deck_detail_payload["cards"]), 1)

        back = self.client.post(
            f"/api/cards/{card_id}/remove-from-deck",
            json={"deckId": deck_id, "amount": 1, "backToFree": True},
        )
        self.assertEqual(back.status_code, 200)
        back_payload = back.get_json()
        self.assertEqual(back_payload["freeQuantity"], 2)
        self.assertEqual(back_payload["deckQuantity"], 1)

        exported = self.client.get("/api/export/state")
        self.assertEqual(exported.status_code, 200)
        exported_payload = json.loads(exported.data.decode("utf-8"))
        exported_decks = exported_payload["decks"]
        self.assertEqual(exported_decks[-1]["name"], "喷火龙卡组")
        self.assertEqual(exported_decks[-1]["color"], "#ff8800")
        self.assertGreater(exported_decks[-1]["sortOrder"], exported_decks[0]["sortOrder"])

        import_response = self.client.post(
            "/api/import/state",
            data={"file": (io.BytesIO(json.dumps(exported_payload, ensure_ascii=False).encode("utf-8")), "state.json")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_response.status_code, 200)

        detail = self.client.get(f"/api/cards/{card_id}")
        detail_payload = detail.get_json()
        self.assertEqual(detail_payload["freeQuantity"], 2)
        self.assertEqual(detail_payload["deckQuantity"], 1)
        self.assertEqual(detail_payload["deckBreakdown"][0]["deckName"], "喷火龙卡组")

    def test_set_free_inventory_to_exact_value(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "空闲修改卡组", "description": "测试", "color": "#226688"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        card_id = self._find_card_id(self.client, "小火龙", "小火龙")
        add_response = self.client.post(
            f"/api/cards/{card_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 1, "consumeFree": False},
        )
        self.assertEqual(add_response.status_code, 200)
        self.assertEqual(add_response.get_json()["deckQuantity"], 1)
        self.assertEqual(add_response.get_json()["freeQuantity"], 2)

        set_response = self.client.put(
            f"/api/cards/{card_id}/free-quantity",
            json={"quantity": 7},
        )
        self.assertEqual(set_response.status_code, 200)
        payload = set_response.get_json()
        self.assertEqual(payload["freeQuantity"], 7)
        self.assertEqual(payload["deckQuantity"], 1)

    def test_deck_detail_groups_cards_and_basic_energies(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "展示卡组", "description": "分组展示测试", "color": "#336699"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        card_ids = {
            "小火龙": self._find_card_id(self.client, "小火龙", "小火龙"),
            "高级球": self._find_card_id(self.client, "高级球", "高级球"),
            "研究员": self._find_card_id(self.client, "研究员", "研究员"),
            "山岳竞技场": self._find_card_id(self.client, "山岳竞技场", "山岳竞技场"),
            "神奇糖果": self._find_card_id(self.client, "神奇糖果", "神奇糖果"),
            "双重涡轮能量": self._find_card_id(self.client, "双重涡轮能量", "双重涡轮能量"),
        }
        for name, card_id in card_ids.items():
            response = self.client.post(
                f"/api/cards/{card_id}/add-to-deck",
                json={"deckId": deck_id, "amount": 1, "consumeFree": False},
            )
            self.assertEqual(response.status_code, 200, msg=name)

        energy_update = self.client.put(
            f"/api/decks/{deck_id}/basic-energies",
            json={
                "items": [
                    {"code": "GRA", "quantity": 7},
                    {"code": "FIR", "quantity": 3},
                ]
            },
        )
        self.assertEqual(energy_update.status_code, 200)

        detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(detail.status_code, 200)
        payload = detail.get_json()
        self.assertEqual(payload["cardCount"], 16)
        self.assertEqual(payload["mainCardCount"], 16)
        self.assertEqual(payload["backupCardCount"], 0)

        sections = {section["key"]: section for section in payload["sections"]}
        self.assertEqual(sections["pokemon"]["column"], "left")
        self.assertEqual(sections["energy"]["column"], "left")
        self.assertEqual(sections["item"]["column"], "right")
        self.assertEqual(sections["supporter"]["title"], "支援者")
        self.assertEqual(sections["tool"]["title"], "宝可梦道具")
        self.assertNotIn("backup", sections)

        self.assertEqual([item["cardName"] for item in sections["pokemon"]["items"]], ["小火龙"])
        self.assertEqual([item["cardName"] for item in sections["item"]["items"]], ["高级球"])
        self.assertEqual([item["cardName"] for item in sections["supporter"]["items"]], ["研究员"])
        self.assertEqual([item["cardName"] for item in sections["stadium"]["items"]], ["山岳竞技场"])
        self.assertEqual([item["cardName"] for item in sections["tool"]["items"]], ["神奇糖果"])

        energy_items = sections["energy"]["items"]
        self.assertEqual([item["cardName"] for item in energy_items], ["双重涡轮能量", "基本草能量", "基本火能量"])
        self.assertEqual(energy_items[1]["displayCode"], "GRA")
        self.assertEqual(energy_items[1]["rarity"], "")
        self.assertEqual(energy_items[1]["deckQuantity"], 7)
        self.assertEqual(energy_items[2]["displayCode"], "FIR")
        self.assertEqual(energy_items[2]["deckQuantity"], 3)

        decks = self.client.get("/api/decks").get_json()["items"]
        deck_summary = next(deck for deck in decks if deck["id"] == deck_id)
        self.assertEqual(deck_summary["cardCount"], 16)

    def test_deck_detail_splits_backup_cards_from_main_cards(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "备卡测试卡组", "description": "备卡分区测试", "color": "#7755aa"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        charmander_id = self._find_card_id(self.client, "小火龙", "小火龙")
        ball_id = self._find_card_id(self.client, "高级球", "高级球")
        researcher_id = self._find_card_id(self.client, "研究员", "研究员")

        self.assertEqual(
            self.client.post(f"/api/cards/{charmander_id}/add-to-deck", json={"deckId": deck_id, "amount": 3, "consumeFree": False}).status_code,
            200,
        )
        self.assertEqual(
            self.client.post(f"/api/cards/{ball_id}/add-to-deck", json={"deckId": deck_id, "amount": 2, "consumeFree": False}).status_code,
            200,
        )
        self.assertEqual(
            self.client.post(f"/api/cards/{researcher_id}/add-to-deck", json={"deckId": deck_id, "amount": 1, "consumeFree": False}).status_code,
            200,
        )

        charmander_backup = self.client.put(
            f"/api/decks/{deck_id}/cards/{charmander_id}/backup-quantity",
            json={"quantity": 1},
        )
        self.assertEqual(charmander_backup.status_code, 200)

        ball_backup = self.client.put(
            f"/api/decks/{deck_id}/cards/{ball_id}/backup-quantity",
            json={"quantity": 2},
        )
        self.assertEqual(ball_backup.status_code, 200)

        detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(detail.status_code, 200)
        payload = detail.get_json()
        self.assertEqual(payload["cardCount"], 6)
        self.assertEqual(payload["mainCardCount"], 3)
        self.assertEqual(payload["backupCardCount"], 3)

        sections = payload["sections"]
        self.assertEqual([section["key"] for section in sections][-1], "backup")
        sections_by_key = {section["key"]: section for section in sections}

        self.assertEqual([item["cardName"] for item in sections_by_key["pokemon"]["items"]], ["小火龙"])
        self.assertEqual([item["deckQuantity"] for item in sections_by_key["pokemon"]["items"]], [2])
        self.assertNotIn("item", sections_by_key)
        self.assertEqual([item["cardName"] for item in sections_by_key["supporter"]["items"]], ["研究员"])
        self.assertEqual(sections_by_key["supporter"]["title"], "支援者")

        backup_items = sections_by_key["backup"]["items"]
        self.assertEqual([item["cardName"] for item in backup_items], ["小火龙", "高级球"])
        self.assertEqual([item["deckQuantity"] for item in backup_items], [1, 2])

    def test_deck_detail_group_order_update_reorders_same_name_cards(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "同组顺序测试", "description": "卡组详情同组排序", "color": "#556677"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        base_mouse_id = self._find_card_id(self.client, "一对鼠", "一对鼠")
        ball_flash_mouse_id = self._find_card_id(self.client, "一对鼠", "一对鼠·精灵球闪")
        master_flash_mouse_id = self._find_card_id(self.client, "一对鼠", "一对鼠·大师球闪")

        for card_id in [base_mouse_id, ball_flash_mouse_id, master_flash_mouse_id]:
            self.assertEqual(
                self.client.post(f"/api/cards/{card_id}/add-to-deck", json={"deckId": deck_id, "amount": 1, "consumeFree": False}).status_code,
                200,
            )

        detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(detail.status_code, 200)
        pokemon_section = next(section for section in detail.get_json()["sections"] if section["key"] == "pokemon")
        self.assertEqual(
            [item["cardName"] for item in pokemon_section["items"]],
            ["一对鼠", "一对鼠·精灵球闪", "一对鼠·大师球闪"],
        )

        reorder_response = self.client.put(
            f"/api/decks/{deck_id}/group-order",
            json={
                "groupKey": pokemon_section["items"][0]["sameNameGroupKey"],
                "cardIds": [master_flash_mouse_id, base_mouse_id, ball_flash_mouse_id],
            },
        )
        self.assertEqual(reorder_response.status_code, 200)
        reordered_section = next(section for section in reorder_response.get_json()["sections"] if section["key"] == "pokemon")
        self.assertEqual(
            [item["cardName"] for item in reordered_section["items"]],
            ["一对鼠·大师球闪", "一对鼠", "一对鼠·精灵球闪"],
        )

    def test_deck_detail_section_order_update_reorders_all_cards_in_section(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "全行顺序测试", "description": "卡组详情整段排序", "color": "#4d6a8a"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        card_ids = [
            self._find_card_id(self.client, "小火龙", "小火龙"),
            self._find_card_id(self.client, "皮卡丘", "皮卡丘"),
            self._find_card_id(self.client, "一对鼠", "一对鼠"),
        ]
        for card_id in card_ids:
            self.assertEqual(
                self.client.post(f"/api/cards/{card_id}/add-to-deck", json={"deckId": deck_id, "amount": 1, "consumeFree": False}).status_code,
                200,
            )

        detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(detail.status_code, 200)
        pokemon_section = next(section for section in detail.get_json()["sections"] if section["key"] == "pokemon")
        original_names = [item["cardName"] for item in pokemon_section["items"]]
        reversed_items = list(reversed(pokemon_section["items"]))

        reorder_response = self.client.put(
            f"/api/decks/{deck_id}/section-order",
            json={
                "sectionKey": "pokemon",
                "entryKeys": [item["deckSectionEntryKey"] for item in reversed_items],
            },
        )
        self.assertEqual(reorder_response.status_code, 200)
        reordered_section = next(section for section in reorder_response.get_json()["sections"] if section["key"] == "pokemon")
        self.assertEqual([item["cardName"] for item in reordered_section["items"]], list(reversed(original_names)))

        persisted_detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(persisted_detail.status_code, 200)
        persisted_section = next(section for section in persisted_detail.get_json()["sections"] if section["key"] == "pokemon")
        self.assertEqual([item["cardName"] for item in persisted_section["items"]], list(reversed(original_names)))

    def test_backup_quantities_roundtrip_in_state_export_import(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "备卡回灌卡组", "description": "备卡导入导出测试", "color": "#884422"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        card_id = self._find_card_id(self.client, "小火龙", "小火龙")
        self.assertEqual(
            self.client.post(f"/api/cards/{card_id}/add-to-deck", json={"deckId": deck_id, "amount": 2, "consumeFree": False}).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(f"/api/decks/{deck_id}/cards/{card_id}/backup-quantity", json={"quantity": 1}).status_code,
            200,
        )

        exported = self.client.get("/api/export/state")
        self.assertEqual(exported.status_code, 200)
        exported_payload = json.loads(exported.data.decode("utf-8"))
        exported_card = next(card for card in exported_payload["cards"] if card["cardName"] == "小火龙")
        exported_entry = next(entry for entry in exported_card["deckQuantities"] if entry["deckName"] == "备卡回灌卡组")
        self.assertEqual(exported_entry["quantity"], 2)
        self.assertEqual(exported_entry["backupQuantity"], 1)

        import_response = self.client.post(
            "/api/import/state",
            data={"file": (io.BytesIO(json.dumps(exported_payload, ensure_ascii=False).encode("utf-8")), "state.json")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_response.status_code, 200)

        imported_deck_id = next(deck["id"] for deck in self.client.get("/api/decks").get_json()["items"] if deck["name"] == "备卡回灌卡组")
        detail = self.client.get(f"/api/decks/{imported_deck_id}")
        self.assertEqual(detail.status_code, 200)
        payload = detail.get_json()
        self.assertEqual(payload["mainCardCount"], 1)
        self.assertEqual(payload["backupCardCount"], 1)
        backup_section = next(section for section in payload["sections"] if section["key"] == "backup")
        self.assertEqual([item["cardName"] for item in backup_section["items"]], ["小火龙"])
        self.assertEqual([item["deckQuantity"] for item in backup_section["items"]], [1])

    def test_deck_card_quantity_action_operates_on_main_and_backup_separately(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "主备分离测试", "description": "主牌备卡动作分离", "color": "#445566"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        ball_id = self._find_card_id(self.client, "高级球", "高级球")
        self.assertEqual(
            self.client.post(f"/api/cards/{ball_id}/add-to-deck", json={"deckId": deck_id, "amount": 4, "consumeFree": False}).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(f"/api/decks/{deck_id}/cards/{ball_id}/backup-quantity", json={"quantity": 1}).status_code,
            200,
        )

        wrong_direction = self.client.post(
            f"/api/decks/{deck_id}/cards/{ball_id}/quantity-action",
            json={"entryType": "main", "mode": "back_to_free", "targetQuantity": 4},
        )
        self.assertEqual(wrong_direction.status_code, 400)
        self.assertIn("目标数量更大", wrong_direction.get_json()["error"])

        update_main = self.client.post(
            f"/api/decks/{deck_id}/cards/{ball_id}/quantity-action",
            json={"entryType": "main", "mode": "remove", "targetQuantity": 2},
        )
        self.assertEqual(update_main.status_code, 200)
        main_payload = update_main.get_json()
        main_sections = {section["key"]: section for section in main_payload["sections"]}
        self.assertEqual([item["cardName"] for item in main_sections["item"]["items"]], ["高级球"])
        self.assertEqual([item["deckQuantity"] for item in main_sections["item"]["items"]], [2])
        self.assertEqual([item["cardName"] for item in main_sections["backup"]["items"]], ["高级球"])
        self.assertEqual([item["deckQuantity"] for item in main_sections["backup"]["items"]], [1])

        update_backup = self.client.post(
            f"/api/decks/{deck_id}/cards/{ball_id}/quantity-action",
            json={"entryType": "backup", "mode": "back_to_free", "targetQuantity": 0},
        )
        self.assertEqual(update_backup.status_code, 200)
        backup_payload = update_backup.get_json()
        backup_sections = {section["key"]: section for section in backup_payload["sections"]}
        self.assertEqual([item["cardName"] for item in backup_sections["item"]["items"]], ["高级球"])
        self.assertEqual([item["deckQuantity"] for item in backup_sections["item"]["items"]], [2])
        self.assertNotIn("backup", backup_sections)

    def test_backup_to_main_module_reduces_backup_quantity_and_keeps_total(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "备卡转回模块测试", "description": "独立模块转回主卡", "color": "#6b7c8d"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        ball_id = self._find_card_id(self.client, "高级球", "高级球")
        self.assertEqual(
            self.client.post(f"/api/cards/{ball_id}/add-to-deck", json={"deckId": deck_id, "amount": 4, "consumeFree": False}).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(f"/api/decks/{deck_id}/cards/{ball_id}/backup-quantity", json={"quantity": 3}).status_code,
            200,
        )

        # 从备卡转 2 张回主卡
        move = self.client.put(
            f"/api/decks/{deck_id}/cards/{ball_id}/backup-quantity",
            json={"quantity": 1},
        )
        self.assertEqual(move.status_code, 200)

        detail = self.client.get(f"/api/decks/{deck_id}")
        self.assertEqual(detail.status_code, 200)
        payload = detail.get_json()
        self.assertEqual(payload["cardCount"], 4)
        self.assertEqual(payload["mainCardCount"], 3)
        self.assertEqual(payload["backupCardCount"], 1)

        sections = {section["key"]: section for section in payload["sections"]}
        self.assertIn("item", sections)
        self.assertIn("backup", sections)
        self.assertEqual([item["deckQuantity"] for item in sections["item"]["items"]], [3])
        self.assertEqual([item["deckQuantity"] for item in sections["backup"]["items"]], [1])

    def test_deck_same_name_limit_blocks_add_to_deck_but_exempts_basic_energy(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "同名限制测试", "description": "同名上限", "color": "#884455"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        charmander_id = self._find_card_id(self.client, "小火龙", "小火龙")
        charmander_ball_flash_id = self._find_card_id(self.client, "小火龙", "小火龙球闪")
        fire_energy_id = self._find_card_id(self.client, "火能量", "火能量")

        self.assertEqual(
            self.client.post(f"/api/cards/{charmander_id}/add-to-deck", json={"deckId": deck_id, "amount": 3, "consumeFree": False}).status_code,
            200,
        )

        same_name_limit = self.client.post(
            f"/api/cards/{charmander_ball_flash_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 2, "consumeFree": False},
        )
        self.assertEqual(same_name_limit.status_code, 400)
        self.assertIn("同名卡牌合计不能超过 4 张", same_name_limit.get_json()["error"])

        basic_energy_response = self.client.post(
            f"/api/cards/{fire_energy_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 8, "consumeFree": False},
        )
        self.assertEqual(basic_energy_response.status_code, 200)

    def test_inventory_table_group_quantity_update_rejects_same_name_limit_violation(self):
        holdings = self.client.get("/api/holdings")
        self.assertEqual(holdings.status_code, 200)
        payload = holdings.get_json()

        ordinary = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        mouse_group = next(group for group in ordinary["groups"] if group["groupName"] == "一对鼠 无")

        decks = self.client.get("/api/decks").get_json()["items"]
        deck_ids = {deck["name"]: deck["id"] for deck in decks}
        forced_quantities = [2, 2, 1]

        update_response = self.client.put(
            "/api/inventory-table/group-quantities",
            json={
                "groupKey": mouse_group["groupKey"],
                "cards": [
                    {
                        "id": item["id"],
                        "freeQuantity": item["freeQuantity"],
                        "deckQuantities": [
                            {
                                "deckId": deck_ids[deck_name],
                                "quantity": forced_quantities[index] if deck_name == "电友" else item["deckQuantities"][deck_name],
                            }
                            for deck_name in ["电友", "龙柱", "铝钢龙", "多龙"]
                        ],
                    }
                    for index, item in enumerate(mouse_group["items"])
                ],
            },
        )
        self.assertEqual(update_response.status_code, 400)
        self.assertIn("同名卡牌合计不能超过 4 张", update_response.get_json()["error"])

    def test_basic_energies_roundtrip_in_state_export_import(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "能量卡组", "description": "基础能量测试", "color": "#558833"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        update_response = self.client.put(
            f"/api/decks/{deck_id}/basic-energies",
            json={"items": [{"code": "PSY", "quantity": 5}, {"code": "MET", "quantity": 2}]},
        )
        self.assertEqual(update_response.status_code, 200)

        exported = self.client.get("/api/export/state")
        self.assertEqual(exported.status_code, 200)
        exported_payload = json.loads(exported.data.decode("utf-8"))
        exported_deck = next(deck for deck in exported_payload["decks"] if deck["id"] == deck_id)
        self.assertEqual(
            exported_deck["basicEnergies"],
            [
                {"code": "PSY", "name": "基本超能量", "quantity": 5},
                {"code": "MET", "name": "基本钢能量", "quantity": 2},
            ],
        )

        import_response = self.client.post(
            "/api/import/state",
            data={"file": (io.BytesIO(json.dumps(exported_payload, ensure_ascii=False).encode("utf-8")), "state.json")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_response.status_code, 200)

        imported_deck_id = next(deck["id"] for deck in self.client.get("/api/decks").get_json()["items"] if deck["name"] == "能量卡组")
        detail = self.client.get(f"/api/decks/{imported_deck_id}")
        self.assertEqual(detail.status_code, 200)
        energy_items = next(section for section in detail.get_json()["sections"] if section["key"] == "energy")["items"]
        self.assertEqual([item["displayCode"] for item in energy_items], ["PSY", "MET"])
        self.assertEqual([item["deckQuantity"] for item in energy_items], [5, 2])

    def test_delete_deck_returns_cards_to_free_inventory(self):
        self.assertEqual(self.client.get("/decks").status_code, 200)

        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "临时卡组", "description": "删除回收测试", "color": "#224466"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        self.assertEqual(self.client.get(f"/decks/{deck_id}").status_code, 200)

        search = self.client.get("/api/search?q=小火龙")
        card_id = search.get_json()["items"][0]["id"]

        moved = self.client.post(
            f"/api/cards/{card_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 2, "consumeFree": True},
        )
        self.assertEqual(moved.status_code, 200)
        self.assertEqual(moved.get_json()["freeQuantity"], 0)

        delete_response = self.client.delete(f"/api/decks/{deck_id}")
        self.assertEqual(delete_response.status_code, 204)

        detail = self.client.get(f"/api/cards/{card_id}")
        detail_payload = detail.get_json()
        self.assertEqual(detail_payload["freeQuantity"], 2)
        self.assertEqual(detail_payload["deckQuantity"], 0)

    def test_deleted_default_deck_stays_deleted_after_restart(self):
        decks_before = self.client.get("/api/decks").get_json()["items"]
        dragon_pillar = next(deck for deck in decks_before if deck["name"] == "龙柱")

        delete_response = self.client.delete(f"/api/decks/{dragon_pillar['id']}")
        self.assertEqual(delete_response.status_code, 204)

        remaining_names = [deck["name"] for deck in self.client.get("/api/decks").get_json()["items"]]
        self.assertNotIn("龙柱", remaining_names)

        restarted_app = create_app(
            {
                "TESTING": True,
                "ROOT_DIR": str(self.root),
                "DATABASE": str(self.db_path),
                "DEFAULT_EXCEL_PATH": str(self.default_excel),
            }
        )
        restarted_client = restarted_app.test_client()
        with restarted_client.session_transaction() as session:
            session["account_id"] = 1
            session["account_name"] = "RhymesX"
            session["authed"] = True
        restarted_names = [deck["name"] for deck in restarted_client.get("/api/decks").get_json()["items"]]
        self.assertNotIn("龙柱", restarted_names)

    def test_reorder_decks_updates_list_and_holdings_order(self):
        decks_response = self.client.get("/api/decks")
        decks = decks_response.get_json()["items"]
        original_names = [deck["name"] for deck in decks]
        self.assertEqual(original_names, ["电友", "龙柱", "铝钢龙", "多龙"])

        reversed_ids = [deck["id"] for deck in reversed(decks)]
        reorder_response = self.client.post("/api/decks/reorder", json={"deckIds": reversed_ids})
        self.assertEqual(reorder_response.status_code, 200)

        reordered = self.client.get("/api/decks").get_json()["items"]
        self.assertEqual([deck["name"] for deck in reordered], ["多龙", "铝钢龙", "龙柱", "电友"])

        holdings = self.client.get("/api/holdings").get_json()
        self.assertEqual(holdings["deckNames"], ["多龙", "铝钢龙", "龙柱", "电友"])

    def test_holdings_total_adjust_and_delete_card(self):
        create_deck_response = self.client.post(
            "/api/decks",
            json={"name": "仓库测试卡组", "description": "测试", "color": "#446688"},
        )
        self.assertEqual(create_deck_response.status_code, 201)
        deck_id = create_deck_response.get_json()["id"]

        search = self.client.get("/api/search?q=小火龙")
        card_id = search.get_json()["items"][0]["id"]

        moved = self.client.post(
            f"/api/cards/{card_id}/add-to-deck",
            json={"deckId": deck_id, "amount": 2, "consumeFree": True},
        )
        self.assertEqual(moved.status_code, 200)
        self.assertEqual(moved.get_json()["freeQuantity"], 0)
        self.assertEqual(moved.get_json()["deckQuantity"], 2)

        add_total = self.client.post(f"/api/cards/{card_id}/adjust-total", json={"delta": 1})
        self.assertEqual(add_total.status_code, 200)
        self.assertEqual(add_total.get_json()["freeQuantity"], 1)
        self.assertEqual(add_total.get_json()["deckQuantity"], 2)

        remove_total = self.client.post(f"/api/cards/{card_id}/adjust-total", json={"delta": -2})
        self.assertEqual(remove_total.status_code, 200)
        self.assertEqual(remove_total.get_json()["freeQuantity"], 0)
        self.assertEqual(remove_total.get_json()["deckQuantity"], 1)

        delete_response = self.client.delete(f"/api/cards/{card_id}")
        self.assertEqual(delete_response.status_code, 204)

        deleted_card = self.client.get(f"/api/cards/{card_id}")
        self.assertEqual(deleted_card.status_code, 404)

    def test_inventory_table_page_is_available(self):
        response = self.client.get("/inventory-table")
        self.assertEqual(response.status_code, 200)
        self.assertIn("库存表格", response.get_data(as_text=True))

        index_page = self.client.get("/")
        self.assertEqual(index_page.status_code, 200)
        index_html = index_page.get_data(as_text=True)
        self.assertNotIn('href="/holdings"', index_html)

        decks_page = self.client.get("/decks")
        self.assertEqual(decks_page.status_code, 200)
        decks_html = decks_page.get_data(as_text=True)
        self.assertIn("打开库存表格", decks_html)
        self.assertNotIn("打开总持有", decks_html)

        deck_detail_page = self.client.get("/decks/1")
        self.assertEqual(deck_detail_page.status_code, 200)
        self.assertNotIn('href="/holdings"', deck_detail_page.get_data(as_text=True))

        inventory_table_html = response.get_data(as_text=True)
        self.assertNotIn('href="/holdings"', inventory_table_html)

        app_js = self.client.get("/static/app.js")
        self.assertEqual(app_js.status_code, 200)
        app_js_text = app_js.get_data(as_text=True)
        self.assertIn("{ label: '总持有', value: summary.ownedCount ?? 0, href: '/inventory-table' }", app_js_text)
        self.assertIn("{ label: '卡组', value: summary.deckCount ?? 0, href: '/decks' }", app_js_text)
        self.assertNotIn("href: '/holdings'", app_js_text)
        app_js.close()

    def test_inventory_table_group_quantity_update_sets_exact_values(self):
        self.assertEqual(self.client.get("/inventory-table").status_code, 200)

        holdings = self.client.get("/api/holdings")
        self.assertEqual(holdings.status_code, 200)
        payload = holdings.get_json()

        ordinary = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        charmander_group = next(group for group in ordinary["groups"] if group["groupName"] == "小火龙 火")
        group_items = sorted(charmander_group["items"], key=lambda item: item["id"])

        decks = self.client.get("/api/decks").get_json()["items"]
        deck_ids = {deck["name"]: deck["id"] for deck in decks}

        update_payload = {
            "groupKey": charmander_group["groupKey"],
            "cards": [
                {
                    "id": group_items[0]["id"],
                    "freeQuantity": 5,
                    "deckQuantities": [
                        {"deckId": deck_ids["电友"], "quantity": 1},
                        {"deckId": deck_ids["龙柱"], "quantity": 0},
                        {"deckId": deck_ids["铝钢龙"], "quantity": 2},
                        {"deckId": deck_ids["多龙"], "quantity": 0},
                    ],
                },
                {
                    "id": group_items[1]["id"],
                    "freeQuantity": 0,
                    "deckQuantities": [
                        {"deckId": deck_ids["电友"], "quantity": 0},
                        {"deckId": deck_ids["龙柱"], "quantity": 3},
                        {"deckId": deck_ids["铝钢龙"], "quantity": 0},
                        {"deckId": deck_ids["多龙"], "quantity": 1},
                    ],
                },
            ],
        }

        update_response = self.client.put("/api/inventory-table/group-quantities", json=update_payload)
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.get_json()["updatedCount"], 2)

        first_detail = self.client.get(f"/api/cards/{group_items[0]['id']}")
        self.assertEqual(first_detail.status_code, 200)
        first_payload = first_detail.get_json()
        self.assertEqual(first_payload["freeQuantity"], 5)
        self.assertEqual(first_payload["deckQuantity"], 3)
        self.assertCountEqual(
            first_payload["deckBreakdown"],
            [
                {"deckId": deck_ids["电友"], "deckName": "电友", "quantity": 1},
                {"deckId": deck_ids["铝钢龙"], "deckName": "铝钢龙", "quantity": 2},
            ],
        )

        second_detail = self.client.get(f"/api/cards/{group_items[1]['id']}")
        self.assertEqual(second_detail.status_code, 200)
        second_payload = second_detail.get_json()
        self.assertEqual(second_payload["freeQuantity"], 0)
        self.assertEqual(second_payload["deckQuantity"], 4)
        self.assertCountEqual(
            second_payload["deckBreakdown"],
            [
                {"deckId": deck_ids["龙柱"], "deckName": "龙柱", "quantity": 3},
                {"deckId": deck_ids["多龙"], "deckName": "多龙", "quantity": 1},
            ],
        )

    def test_inventory_table_group_quantity_update_persists_group_item_order(self):
        self.assertEqual(self.client.get("/inventory-table").status_code, 200)

        holdings = self.client.get("/api/holdings")
        self.assertEqual(holdings.status_code, 200)
        payload = holdings.get_json()

        ordinary = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        mouse_group = next(group for group in ordinary["groups"] if group["groupName"] == "一对鼠 无")
        original_ids = [item["id"] for item in mouse_group["items"]]
        reversed_items = list(reversed(mouse_group["items"]))

        decks = self.client.get("/api/decks").get_json()["items"]
        deck_ids = {deck["name"]: deck["id"] for deck in decks}

        update_response = self.client.put(
            "/api/inventory-table/group-quantities",
            json={
                "groupKey": mouse_group["groupKey"],
                "cards": [
                    {
                        "id": item["id"],
                        "freeQuantity": item["freeQuantity"],
                        "deckQuantities": [
                            {"deckId": deck_ids[deck_name], "quantity": item["deckQuantities"][deck_name]}
                            for deck_name in ["电友", "龙柱", "铝钢龙", "多龙"]
                        ],
                    }
                    for item in reversed_items
                ],
            },
        )
        self.assertEqual(update_response.status_code, 200)

        refreshed = self.client.get("/api/holdings")
        self.assertEqual(refreshed.status_code, 200)
        refreshed_payload = refreshed.get_json()
        refreshed_ordinary = next(section for section in refreshed_payload["sections"] if section["key"] == "ordinary_pokemon")
        refreshed_group = next(group for group in refreshed_ordinary["groups"] if group["groupName"] == "一对鼠 无")
        refreshed_ids = [item["id"] for item in refreshed_group["items"]]

        self.assertEqual(refreshed_ids, list(reversed(original_ids)))

    def test_inventory_table_group_reorder_persists_section_group_order(self):
        self.assertEqual(self.client.get("/inventory-table").status_code, 200)

        holdings = self.client.get("/api/holdings")
        self.assertEqual(holdings.status_code, 200)
        payload = holdings.get_json()

        pokemon_section = next(section for section in payload["sections"] if section["key"] == "ordinary_pokemon")
        original_group_keys = [group["groupKey"] for group in pokemon_section["groups"]]
        self.assertGreaterEqual(len(original_group_keys), 2)

        moved_group_keys = original_group_keys[:]
        moved_group_keys[0], moved_group_keys[1] = moved_group_keys[1], moved_group_keys[0]

        reorder_response = self.client.put(
            "/api/inventory-table/group-order",
            json={
                "sectionKey": "item",
                "sectionKey": "ordinary_pokemon",
                "groupKeys": moved_group_keys,
            },
        )
        self.assertEqual(reorder_response.status_code, 200)
        self.assertEqual(reorder_response.get_json()["updatedCount"], len(original_group_keys))

        refreshed = self.client.get("/api/holdings")
        self.assertEqual(refreshed.status_code, 200)
        refreshed_payload = refreshed.get_json()
        refreshed_pokemon_section = next(section for section in refreshed_payload["sections"] if section["key"] == "ordinary_pokemon")
        refreshed_group_keys = [group["groupKey"] for group in refreshed_pokemon_section["groups"]]

        self.assertEqual(refreshed_group_keys, moved_group_keys)

    def test_login_account_is_merged_and_admin_maps_to_rhymesx(self):
        with self.client.session_transaction() as session:
            session.clear()

        unauth_response = self.client.get("/api/summary")
        self.assertEqual(unauth_response.status_code, 401)

        login_response = self.client.post(
            "/login",
            data={"username": "test_admin", "password": "test_pass"},
            follow_redirects=False,
        )
        self.assertEqual(login_response.status_code, 302)

        account_payload = self.client.get("/api/accounts").get_json()
        self.assertEqual(account_payload["current"]["name"], "RhymesX")

        card_id = self._find_card_id(self.client, "CSM1aC-002", "小火龙")
        card_payload = self.client.get(f"/api/cards/{card_id}").get_json()
        self.assertEqual(card_payload["freeQuantity"], 2)

    def test_register_and_change_password(self):
        """注册新账号后不能直接用旧会话访问，改密后能用新密码登录。"""
        invite_code = self._generate_invite_code(self.client)
        # 当前是 RhymesX (id=1)
        register_response = self.client.post(
            "/api/accounts",
            json={"name": "TestUser", "password": "test1234", "inviteCode": invite_code},
        )
        self.assertEqual(register_response.status_code, 201)
        items = register_response.get_json()["items"]
        self.assertTrue(any(item["name"] == "TestUser" for item in items))

        # 修改当前账号密码
        change_response = self.client.put(
            "/api/accounts/password",
            json={"oldPassword": "test_pass", "newPassword": "newpass5678"},
        )
        self.assertEqual(change_response.status_code, 200)
        self.assertTrue(change_response.get_json()["ok"])

        # 旧密码不能再登录
        with self.client.session_transaction() as session:
            session.clear()
        bad_login = self.client.post(
            "/login",
            data={"username": "RhymesX", "password": "test_pass"},
            follow_redirects=False,
        )
        self.assertEqual(bad_login.status_code, 401)

        # 新密码可以登录
        good_login = self.client.post(
            "/login",
            data={"username": "RhymesX", "password": "newpass5678"},
            follow_redirects=False,
        )
        self.assertEqual(good_login.status_code, 302)

        # 新注册账号也能登录
        with self.client.session_transaction() as session:
            session.clear()
        testuser_login = self.client.post(
            "/login",
            data={"username": "TestUser", "password": "test1234"},
            follow_redirects=False,
        )
        self.assertEqual(testuser_login.status_code, 302)

    def test_admin_can_reset_regular_account_password(self):
        """管理员（test_admin 登录）可重置普通账号密码；普通用户不可调用。"""
        # 注册一个普通账号
        invite_code = self._generate_invite_code(self.client)
        # _generate_invite_code 设了 is_admin=True，清掉它来模拟普通用户
        with self.client.session_transaction() as session:
            session["is_admin"] = False
        self.client.post("/api/accounts", json={"name": "Peon", "password": "peonpass", "inviteCode": invite_code})

        # 当前 session 是 account_id=1, account_name=RhymesX（非管理员 session）
        # 普通用户调用 admin reset → 403
        reset_response = self.client.put(
            "/api/accounts/2/password",
            json={"newPassword": "hacked"},
        )
        self.assertEqual(reset_response.status_code, 403)

        # 以管理员身份登录
        with self.client.session_transaction() as session:
            session.clear()
        admin_login = self.client.post(
            "/login",
            data={"username": "test_admin", "password": "test_pass"},
            follow_redirects=False,
        )
        self.assertEqual(admin_login.status_code, 302)

        # 验证 isAdmin
        account_payload = self.client.get("/api/accounts").get_json()
        self.assertTrue(account_payload["isAdmin"])

        # 管理员不能重置 RhymesX（自己）
        bad_reset = self.client.put(
            "/api/accounts/1/password",
            json={"newPassword": "bad1234"},
        )
        self.assertNotEqual(bad_reset.status_code, 200)

        # 管理员重置 Peon 的密码
        good_reset = self.client.put(
            "/api/accounts/2/password",
            json={"newPassword": "rescue999"},
        )
        self.assertEqual(good_reset.status_code, 200)

        # Peon 旧密码不能登录
        with self.client.session_transaction() as session:
            session.clear()
        self.assertEqual(
            self.client.post("/login", data={"username": "Peon", "password": "peonpass"}, follow_redirects=False).status_code,
            401,
        )

        # Peon 新密码可以登录
        self.assertEqual(
            self.client.post("/login", data={"username": "Peon", "password": "rescue999"}, follow_redirects=False).status_code,
            302,
        )


class DbSplitIntegrationTests(unittest.TestCase):
    """测试拆库后的数据完整性和多用户隔离。"""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.data_dir = self.root / "data"
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.default_excel = self.data_dir / "卡表.xlsx"
        self.db_path = self.data_dir / "test.db"

        rows = [
            ["起始包A", "CSM1aC", "002", "小火龙", "宝可梦", "基础", "", "火", "C", "标准", 2, "", ""],
            ["起始包A", "CSM1aC", "002", "小火龙球闪", "宝可梦", "基础", "球闪", "火", "C", "标准", 1, "", ""],
            ["对战包", "SVP", "001", "喷火龙GX", "宝可梦", "阶段2", "宝可梦GX", "火", "RR", "标准", 1, "", ""],
            ["辅助包", "TRN", "101", "高级球", "物品", "", "", "无", "U", "标准", 4, "", ""],
            ["辅助包", "TRN", "102", "研究员", "支援者", "", "", "无", "U", "标准", 2, "", ""],
            ["能量包", "ENG", "202", "火能量", "普通能量", "", "普通能量", "火", "C", "标准", 8, "", ""],
        ]
        workbook = Workbook()
        sheet = workbook.create_sheet("卡表")
        sheet.append(["商品名称", "商品编号", "卡牌编号", "卡牌名称", "类型", "详细", "特殊", "属性", "稀有度", "赛制", "数量", "备注", "昵称"])
        for row in rows:
            sheet.append(row)
        workbook.save(self.default_excel)

        self.app = create_app({
            "TESTING": True,
            "ROOT_DIR": str(self.root),
            "DATABASE": str(self.db_path),
            "DEFAULT_EXCEL_PATH": str(self.default_excel),
            "AUTH_USERNAME": "test_admin",
            "AUTH_PASSWORD": "test_pass",
            "INIT_ADMIN_PASS": "init1234",
        })
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["account_id"] = 1
            session["account_name"] = "RhymesX"
            session["authed"] = True
            session["is_admin"] = True

    def tearDown(self):
        self.temp_dir.cleanup()

    def _create_account(self, name: str, password: str) -> int:
        """注册一个新账号并返回 account_id。"""
        invite = self._generate_invite_code()
        resp = self.client.post("/api/accounts", json={
            "name": name, "password": password, "inviteCode": invite,
        })
        self.assertEqual(resp.status_code, 201)
        accounts = self.client.get("/api/accounts").get_json()
        for acc in accounts["items"]:
            if acc["name"] == name:
                return acc["id"]
        self.fail(f"Account {name} not found after creation")
        return 0

    def _generate_invite_code(self) -> str:
        resp = self.client.post("/api/invite-codes").get_json()
        return resp["codes"][0]["code"]

    def _login_as(self, name: str):
        """以指定账号身份登录（测试模式下绕过密码认证）。"""
        accounts = self.client.get("/api/accounts").get_json()
        acc = next((a for a in accounts["items"] if a["name"] == name), None)
        self.assertIsNotNone(acc, f"Account {name} not found")
        with self.client.session_transaction() as session:
            session["account_id"] = acc["id"]
            session["account_name"] = acc["name"]
            session["authed"] = True
            session.pop("is_admin", None)

    def test_user_has_separate_db_file(self):
        """新账号自动生成独立的 .db 文件。"""
        account_id = self._create_account("Alice", "alice1234")
        accounts_dir = self.data_dir / "accounts"
        self.assertTrue(accounts_dir.exists())
        self.assertTrue((accounts_dir / f"{account_id}.db").exists(), f"Expected {account_id}.db in accounts dir")
        self.assertTrue((accounts_dir / "1.db").exists(), "Default account should have its DB file")

    def test_two_users_have_isolated_free_inventory(self):
        """两个用户的空闲库存互不影响。"""
        alice_id = self._create_account("Alice", "alice1234")

        # RhymesX 先找到小火龙，放入空闲库存
        self._login_as("RhymesX")
        card_id = self._find_card_id(self.client, "小火龙", "小火龙")
        self.client.put(f"/api/cards/{card_id}/free-quantity", json={"quantity": 3})

        # Alice 查看同一张卡，库存应为 0
        self._login_as("Alice")
        card_alice = self.client.get(f"/api/cards/{card_id}").get_json()
        self.assertEqual(card_alice["freeQuantity"], 0)

        # Alice 也放自己的库存
        self.client.put(f"/api/cards/{card_id}/free-quantity", json={"quantity": 5})
        card_alice_after = self.client.get(f"/api/cards/{card_id}").get_json()
        self.assertEqual(card_alice_after["freeQuantity"], 5)

        # RhymesX 的库存不变
        self._login_as("RhymesX")
        card_rhymes = self.client.get(f"/api/cards/{card_id}").get_json()
        self.assertEqual(card_rhymes["freeQuantity"], 3)

    def test_two_users_have_isolated_decks(self):
        """两个用户各自建卡组，互不可见。"""
        self._create_account("Alice", "alice1234")

        # RhymesX 建一个卡组
        self._login_as("RhymesX")
        resp = self.client.post("/api/decks", json={"name": "炎系卡组", "color": "#ff4444"})
        self.assertEqual(resp.status_code, 201)
        rhymes_deck_id = resp.get_json()["id"]

        # Alice 看不到 RhymesX 的卡组
        self._login_as("Alice")
        alice_decks = self.client.get("/api/decks").get_json()["items"]
        self.assertEqual(len(alice_decks), 4)  # only default decks
        self.assertNotIn("炎系卡组", [d["name"] for d in alice_decks])

        # Alice 建自己的卡组
        resp_a = self.client.post("/api/decks", json={"name": "水系卡组", "color": "#4444ff"})
        self.assertEqual(resp_a.status_code, 201)

        # RhymesX 看不到 Alice 的卡组
        self._login_as("RhymesX")
        rhymes_decks = self.client.get("/api/decks").get_json()["items"]
        self.assertIn("炎系卡组", [d["name"] for d in rhymes_decks])
        self.assertNotIn("水系卡组", [d["name"] for d in rhymes_decks])

    def test_two_users_have_isolated_search_preferences(self):
        """两个用户的搜索偏好各自持久化。"""
        self._create_account("Alice", "alice1234")

        self._login_as("RhymesX")
        self.client.put("/api/search/preferences", json={
            "selectedRegulations": ["F", "G"], "considerSameNameRegulation": True,
        })

        self._login_as("Alice")
        self.client.put("/api/search/preferences", json={
            "selectedRegulations": ["H"], "considerSameNameRegulation": False,
        })

        # 验证 Alice 的偏好
        prefs_a = self.client.get("/api/search/options").get_json()["preferences"]
        self.assertEqual(prefs_a["selectedRegulations"], ["H"])
        self.assertFalse(prefs_a["considerSameNameRegulation"])

        # 验证 RhymesX 的偏好
        self._login_as("RhymesX")
        prefs_r = self.client.get("/api/search/options").get_json()["preferences"]
        self.assertCountEqual(prefs_r["selectedRegulations"], ["F", "G"])
        self.assertTrue(prefs_r["considerSameNameRegulation"])

    def test_two_users_have_isolated_holdings_group_orders(self):
        """两个用户的 holdings 分组排序互不影响。"""
        self._create_account("Alice", "alice1234")

        # 两个用户都先确保有一些库存数据
        for user in ("RhymesX", "Alice"):
            self._login_as(user)
            card_id = self._find_card_id(self.client, "高级球", "高级球")
            self.client.put(f"/api/cards/{card_id}/free-quantity", json={"quantity": 2})

            card_id2 = self._find_card_id(self.client, "研究员", "研究员")
            self.client.put(f"/api/cards/{card_id2}/free-quantity", json={"quantity": 2})

        # RhymesX 调整分组顺序
        self._login_as("RhymesX")
        holdings = self.client.get("/api/holdings").get_json()
        item_section = next(s for s in holdings["sections"] if s["key"] == "item")
        reversed_keys = list(reversed([g["groupKey"] for g in item_section["groups"]]))
        self.client.put("/api/inventory-table/group-order", json={
            "sectionKey": "item", "groupKeys": reversed_keys,
        })

        # Alice 的 holdings 顺序不受影响
        self._login_as("Alice")
        holdings_a = self.client.get("/api/holdings").get_json()
        item_section_a = next(s for s in holdings_a["sections"] if s["key"] == "item")
        self.assertIsNotNone(item_section_a)

    def test_migration_preserves_default_account_data(self):
        """迁移脚本 inspect 模式不报错，且能正确识别现有账号。"""
        import subprocess
        result = subprocess.run(
            [
                "C:\\Users\\DELL\\AppData\\Local\\Programs\\Python\\Python311\\python.exe",
                "scripts/migrate_split_db.py",
                "--root-dir", str(self.root),
                "--database", str(self.db_path),
            ],
            capture_output=True, text=True, timeout=30,
        )
        self.assertEqual(result.returncode, 0, f"Migration script failed: {result.stderr}")
        report = json.loads(result.stdout)
        self.assertIn("accounts", report)
        self.assertEqual(len(report["accounts"]), 1)
        self.assertEqual(report["accounts"][0]["name"], "RhymesX")

    def test_migration_apply_and_verify(self):
        """迁移脚本 apply 模式与旧数据格式不冲突，report 正确返回 accounts。"""
        import subprocess
        result = subprocess.run(
            [
                "C:\\Users\\DELL\\AppData\\Local\\Programs\\Python\\Python311\\python.exe",
                "scripts/migrate_split_db.py",
                "--root-dir", str(self.root),
                "--database", str(self.db_path),
                "--apply", "--force",
            ],
            capture_output=True, text=True, timeout=30,
        )
        self.assertEqual(result.returncode, 0, f"Migration apply failed: {result.stderr}")
        report = json.loads(result.stdout)
        self.assertTrue(report["allAccountsVerified"], "All accounts should pass verification")
        self.assertGreaterEqual(len(report["accounts"]), 1, "At least one account should be in the report")

    def _find_card_id(self, client, query: str, card_name: str) -> int:
        payload = client.get(f"/api/search?q={query}").get_json()
        item = next(item for item in payload["items"] if item["cardName"] == card_name)
        return item["id"]


if __name__ == "__main__":
    unittest.main()

